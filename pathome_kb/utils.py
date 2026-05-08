import json
import os
import re
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

import httpx

from .config import PAGE_TEXT_MAX_CHARS

# Outputs land under <repo>/artifacts/pathome_kb/ so they share the
# repo's gitignored artifacts tree.
OUTPUT_DIR = Path(__file__).parent.parent / "artifacts" / "pathome_kb"


def get_crop_dir(crop: str) -> Path:
    """Return (and create) the per-crop output directory: outputs/{Crop}/"""
    d = OUTPUT_DIR / crop.title()
    d.mkdir(parents=True, exist_ok=True)
    return d


def ensure_output_dir() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR


def save_json(filename: str, data: dict | list, output_dir: Path | None = None) -> Path:
    path = (output_dir or ensure_output_dir()) / filename
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"  Saved: {path}")
    return path


def load_json(filename: str, output_dir: Path | None = None) -> dict | list:
    path = (output_dir or ensure_output_dir()) / filename
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_file(filename: str, content: str, output_dir: Path | None = None) -> Path:
    path = (output_dir or ensure_output_dir()) / filename
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  Saved: {path}")
    return path


def chunk_list(lst: list, batch_size: int = 5) -> list[list]:
    """Split a list into chunks of batch_size."""
    return [lst[i : i + batch_size] for i in range(0, len(lst), batch_size)]


def extract_unique_pathogens(extractions: dict) -> list[str]:
    """Pull unique non-null pathogen names from raw extraction data."""
    pathogens = set()
    for source in extractions.get("extractions", []):
        for disease in source.get("extracted_diseases", []):
            pathogen = disease.get("pathogen_scientific_name", {})
            if isinstance(pathogen, dict):
                val = pathogen.get("value")
            else:
                val = pathogen
            if val and val.lower() not in ("null", "unknown", "none"):
                pathogens.add(val)
    return sorted(pathogens)


def registry_to_markdown(registry: dict) -> str:
    """Convert the final registry JSON into a markdown table with hyperlinked cells."""
    crop = registry.get("crop", "Unknown")
    date = registry.get("generated_date", datetime.now(timezone.utc).isoformat())
    diseases = registry.get("diseases", [])

    lines = [
        f"# {crop} Disease Registry",
        f"",
        f"*Generated: {date}*",
        f"",
        "| # | Disease | Pathogen | Type | Affected Parts | Visual Description | Confidence |",
        "|---|---|---|---|---|---|---|",
    ]

    for i, d in enumerate(diseases, 1):
        name = d.get("disease_name", "?")
        pathogen = _hyperlink_field(d.get("pathogen_scientific_name", {}))
        dtype = _hyperlink_field(d.get("type_of_disease", {}))
        parts = _hyperlink_field(d.get("affected_parts", {}))

        vs = d.get("visual_symptoms", {})
        desc_parts = [
            _hyperlink_field(vs.get("summary", {})),
            _hyperlink_field(vs.get("diagnostic_features", {})),
            _hyperlink_field(vs.get("look_alikes", {})),
        ]
        visual_desc = ". ".join(p for p in desc_parts if p and p != "—")
        if not visual_desc:
            visual_desc = "—"

        confidence = d.get("confidence", "?")

        lines.append(
            f"| {i} | {name} | {pathogen} | {dtype} | {parts} | {visual_desc} | {confidence} |"
        )

    # Conflicts section
    conflicts = []
    for d in diseases:
        for c in d.get("conflicts", []):
            conflicts.append(
                f"- **{d['disease_name']}** → `{c['field']}`: "
                f"canonical=`{c.get('canonical_value')}`, "
                f"alt=`{c.get('alternative_value')}` "
                f"({c.get('note', '')})"
            )

    if conflicts:
        lines.append("")
        lines.append("## Conflicts")
        lines.append("")
        lines.extend(conflicts)

    return "\n".join(lines)


def _hyperlink_field(field: dict) -> str:
    """Convert a cited field {value, url, quote} into a markdown hyperlink."""
    if not field or not isinstance(field, dict):
        return "—"

    value = field.get("value")
    if not value:
        return "—"

    # Handle array values
    if isinstance(value, list):
        value = ", ".join(str(v) for v in value)

    url = field.get("url")
    if url:
        # Escape pipes for markdown tables
        value_escaped = str(value).replace("|", "\\|")
        return f"[{value_escaped}]({url})"
    return str(value).replace("|", "\\|")


def fetch_page(url: str, timeout: int = 30) -> str | None:
    """Fetch a URL and return its text content, stripped of HTML tags."""
    try:
        with httpx.Client(follow_redirects=True, timeout=timeout) as client:
            resp = client.get(url, headers={"User-Agent": "Mozilla/5.0 (compatible; DiseaseRegistryBot/1.0)"})
            resp.raise_for_status()
            html = resp.text

        # Strip HTML tags to get plain text (good enough for LLM extraction)
        text = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()

        if len(text) > PAGE_TEXT_MAX_CHARS:
            text = text[:PAGE_TEXT_MAX_CHARS] + "\n\n[TRUNCATED — page too long]"

        return text
    except Exception as e:
        print(f"  WARNING: Failed to fetch {url}: {e}")
        return None


def validate_pathogens_ncbi(pathogens: list[str]) -> list[dict]:
    """Validate pathogen names against NCBI Taxonomy via Entrez API. No LLM needed."""
    results = []
    base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"

    with httpx.Client(timeout=15) as client:
        for name in pathogens:
            try:
                # Search NCBI Taxonomy
                search = client.get(
                    f"{base}/esearch.fcgi",
                    params={"db": "taxonomy", "term": name, "retmode": "xml"},
                )
                search_root = ET.fromstring(search.text)
                id_list = [el.text for el in search_root.findall(".//Id")]

                if not id_list:
                    results.append({
                        "submitted_name": name,
                        "status": "not_found",
                        "current_accepted_name": None,
                        "ncbi_taxonomy_url": None,
                    })
                    continue

                tax_id = id_list[0]

                # Fetch taxonomy record
                fetch = client.get(
                    f"{base}/efetch.fcgi",
                    params={"db": "taxonomy", "id": tax_id, "retmode": "xml"},
                )
                fetch_root = ET.fromstring(fetch.text)
                taxon = fetch_root.find(".//Taxon")

                sci_name = taxon.findtext("ScientificName", "") if taxon is not None else ""
                rank = taxon.findtext("Rank", "") if taxon is not None else ""

                status = "accepted" if sci_name.lower() == name.lower() else "synonym"

                results.append({
                    "submitted_name": name,
                    "status": status,
                    "current_accepted_name": sci_name,
                    "rank": rank,
                    "ncbi_taxonomy_url": f"https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?id={tax_id}",
                    "ncbi_tax_id": tax_id,
                })
                print(f"    {name} → {status} ({sci_name})")

            except Exception as e:
                results.append({
                    "submitted_name": name,
                    "status": "error",
                    "current_accepted_name": None,
                    "ncbi_taxonomy_url": None,
                    "error": str(e),
                })
                print(f"    {name} → error: {e}")

    return results



def load_disease_names_from_dir(disease_dir: str, crop: str) -> list[str]:
    """Extract disease names from image directory folder names.

    Supports both naming conventions:
      {disease_dir}/{Crop} Diseases/{Disease Name}/
      {disease_dir}/{Crop}_Diseases/{Disease_Name}/
    """
    for pattern in [f"{crop.title()} Diseases", f"{crop.title()}_Diseases"]:
        crop_dir = Path(disease_dir) / pattern
        if crop_dir.is_dir():
            break
    else:
        raise FileNotFoundError(
            f"No disease directory found at {disease_dir}/{crop.title()} Diseases "
            f"or {disease_dir}/{crop.title()}_Diseases"
        )
    names = sorted(
        d.name for d in crop_dir.iterdir()
        if d.is_dir() and not d.name.startswith(".")
    )
    return names



def write_enriched_xlsx(registry: dict, base_sheet_path: str | None, output_path: str):
    """Write enriched xlsx from final registry, optionally starting from a base sheet."""
    import openpyxl

    if base_sheet_path and Path(base_sheet_path).exists():
        wb = openpyxl.load_workbook(base_sheet_path)
        ws = wb.active

        # Read existing headers
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

        # Add visual description column if it doesn't already exist
        col_name = "Visual Description"
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers), value=col_name)

        # Find the disease name column
        disease_col = None
        for i, h in enumerate(headers):
            if "disease" in h.lower() and "type" not in h.lower():
                disease_col = i
                break

        if disease_col is None:
            print("  WARNING: Could not find 'Disease' column in base sheet")
        else:
            # Build lookup from registry
            registry_lookup = {}
            for d in registry.get("diseases", []):
                key = d.get("disease_name", "").strip().lower()
                registry_lookup[key] = d

            # Fill in new columns for each existing row
            for row_idx in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=disease_col + 1).value
                if not cell_val:
                    continue
                disease_name = str(cell_val).strip().lower()

                # Fuzzy match: try exact, then substring
                match = registry_lookup.get(disease_name)
                if not match:
                    for key, d in registry_lookup.items():
                        if disease_name in key or key in disease_name:
                            match = d
                            break

                if match:
                    vs = match.get("visual_symptoms", {})
                    parts = [
                        _extract_value(vs.get("summary", {})),
                        _extract_value(vs.get("diagnostic_features", {})),
                        _extract_value(vs.get("look_alikes", {})),
                    ]
                    combined = ". ".join(p for p in parts if p)
                    if combined and "Visual Description" in headers:
                        col_idx = headers.index("Visual Description") + 1
                        ws.cell(row=row_idx, column=col_idx, value=combined)
    else:
        # No base sheet — create from scratch
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Disease Registry"

        headers = [
            "Disease", "Pathogen", "Type of Disease",
            "Affected Parts", "Visual Description",
            "Confidence", "Num Sources",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)

        for row_idx, d in enumerate(registry.get("diseases", []), 2):
            vs = d.get("visual_symptoms", {})
            parts = [
                _extract_value(vs.get("summary", {})),
                _extract_value(vs.get("diagnostic_features", {})),
                _extract_value(vs.get("look_alikes", {})),
            ]
            row_data = [
                d.get("disease_name", ""),
                _extract_value(d.get("pathogen_scientific_name", {})),
                _extract_value(d.get("type_of_disease", {})),
                _extract_value(d.get("affected_parts", {})),
                ". ".join(p for p in parts if p),
                d.get("confidence", ""),
                d.get("num_sources", ""),
            ]
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(output_path)
    print(f"  Saved enriched xlsx: {output_path}")


def _extract_value(field: dict) -> str:
    """Extract the display value from a cited field {value, url, quote}."""
    if not field or not isinstance(field, dict):
        return ""
    value = field.get("value")
    if not value:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value)


def today_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")
