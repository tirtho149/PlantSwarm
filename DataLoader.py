"""
Disease Dataset Curator & Report Generator
===========================================
Handles BOTH local datasets (dataloader.py) AND online Kaggle/Zenodo datasets
(data_loader.py) in one unified pipeline.

FEATURES
─────────
• Interactive prompt: asks how many images per class to save.
• Smart downloading: skips any dataset whose output folder already exists.
• One directory per class:  Curated_Dataset/Images/<Crop>/<Disease>/
  Files named:  1.jpg, 2.png, ... (sequential).
• Deletes raw download after sampling to save disk space.

• Works with LOCAL folders (set LOCAL_SOURCE_ROOT) and/or ONLINE datasets.

REQUIREMENTS
────────────
  pip install kaggle Pillow pandas scikit-learn requests tqdm openpyxl
  pip install datasets          # required for LeafNet (HuggingFace)
  pip install kagglehub         # required for Banana Leaf dataset
  pip install tensorflow tensorflow_datasets  # required for PlantVillage

USAGE
─────
  1.  Set LOCAL_SOURCE_ROOT below (or leave "" to skip local datasets).
  2.  Place ~/.kaggle/kaggle.json for Kaggle datasets.
  3.  python disease_report_generator.py
"""

# ═══════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════

import os
import sys

# Force line-buffered stdout/stderr so output appears immediately in jobs/pipes
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(line_buffering=True)
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(line_buffering=True)

def _log(msg, flush=True):
    print(msg, flush=flush)

_log("DataLoader: starting (importing packages) ...")

# Work root for all data/caches/temp.
# Priority:
#   1) CYAG_WORK_ROOT env var
#   2) historical hardcoded path (if writable)
#   3) local project fallback (always writable on user machines)
_DEFAULT_WORK_ROOT = "/work/mech-ai-scratch/tirtho/CyAg"
_LOCAL_WORK_FALLBACK = os.path.join(os.path.dirname(__file__), "work")
_WORK_ROOT = os.environ.get("CYAG_WORK_ROOT", _DEFAULT_WORK_ROOT)


def _ensure_work_root(path: str) -> str:
    try:
        os.makedirs(path, exist_ok=True)
        return path
    except OSError:
        os.makedirs(_LOCAL_WORK_FALLBACK, exist_ok=True)
        return _LOCAL_WORK_FALLBACK


_WORK_ROOT = _ensure_work_root(_WORK_ROOT)
# Create an ephemeral cache directory per process run so no persistent
# cached artifacts can be reused across executions.
import tempfile
import uuid

_cache_dir = os.path.join(
    _WORK_ROOT,
    f".cache_run_{uuid.uuid4().hex}",
)
os.makedirs(_cache_dir, exist_ok=True)
os.environ["CYAG_WORK_ROOT"] = _WORK_ROOT
# Kaggle credentials: use kaggle.json in work root (KaggleApi and kagglehub read this)
_KAGGLE_JSON = os.path.join(_WORK_ROOT, "kaggle.json")
if os.path.isfile(_KAGGLE_JSON):
    os.environ["KAGGLE_CONFIG_DIR"] = _WORK_ROOT
os.environ["KAGGLEHUB_CACHE"] = os.path.join(_cache_dir, "kagglehub")
os.environ["HF_HOME"] = os.path.join(_cache_dir, "huggingface")
os.environ["HUGGINGFACE_HUB_CACHE"] = os.path.join(_cache_dir, "huggingface", "hub")
os.environ["HF_DATASETS_CACHE"] = os.path.join(_cache_dir, "huggingface", "datasets")
os.environ["TFDS_DATA_DIR"] = os.path.join(_cache_dir, "tfds")
os.environ["KERAS_HOME"] = os.path.join(_cache_dir, "keras")
_tmp = os.path.join(_cache_dir, "tmp")
os.environ["TMPDIR"] = _tmp
os.environ["TEMP"] = _tmp
os.environ["TMP"] = _tmp
try:
    os.makedirs(os.environ["TMPDIR"], exist_ok=True)
except OSError:
    pass


def _load_local_env():
    """
    Load a simple .env file from the work root (if present) and
    propagate HF-related variables so LeafNet can authenticate.
    Expected format per line: KEY=VALUE or KEY="VALUE".
    """
    env_path = os.path.join(_WORK_ROOT, ".env")
    if not os.path.isfile(env_path):
        return
    try:
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                key, val = line.split("=", 1)
                key = key.strip()
                val = val.strip().strip("'").strip('"')
                if key and val and key not in os.environ:
                    os.environ[key] = val
    except OSError:
        pass

    # Normalise HuggingFace token variables if a generic HF token is provided.
    hf_token = os.environ.get("HF")
    if hf_token:
        os.environ.setdefault("HUGGINGFACE_HUB_TOKEN", hf_token)
        os.environ.setdefault("HF_TOKEN", hf_token)

import gc
import zipfile, shutil, difflib, random, json, re, time, hashlib
import subprocess
import multiprocessing
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import requests
from io import BytesIO

import pandas as pd
import numpy as np
from PIL import Image
from sklearn.utils import shuffle as sk_shuffle

# ── HuggingFace / datasets (optional — only needed for LeafNet) ──────────────
try:
    # Patch md5 for environments where usedforsecurity kwarg causes errors
    import hashlib as _hashlib
    _orig_md5 = _hashlib.md5
    def _patched_md5(*a, **kw):
        kw.pop("usedforsecurity", None)
        return _orig_md5(*a, **kw)
    _hashlib.md5 = _patched_md5

    from datasets import load_dataset as hf_load_dataset
    HF_AVAILABLE = True
except ImportError:
    HF_AVAILABLE = False

# Optional: download LeafNet to disk then read parquet row-by-row (avoids stream OOM)
try:
    from huggingface_hub import snapshot_download
    HF_HUB_AVAILABLE = True
except ImportError:
    HF_HUB_AVAILABLE = False
try:
    import pyarrow.parquet as pq
    PYARROW_AVAILABLE = True
except ImportError:
    PYARROW_AVAILABLE = False

# ── tqdm shim ───────────────────────────────────────────────────────────────
try:
    from tqdm import tqdm as _tqdm
    def tqdm(iterable=None, **kw): return _tqdm(iterable, **kw)
except ImportError:
    class tqdm:
        def __init__(self, iterable=None, desc="", total=None, **kw):
            self._it = iter(iterable) if iterable is not None else None
            if desc: print(f"  {desc} ...")
        def __iter__(self):  return self._it
        def __next__(self):  return next(self._it)
        def __enter__(self): return self
        def __exit__(self, *a): pass
        def update(self, n=1): pass

# ── Kaggle ───────────────────────────────────────────────────────────────────
try:
    from kaggle.api.kaggle_api_extended import KaggleApi
    KAGGLE_AVAILABLE = True
except ImportError:
    KAGGLE_AVAILABLE = False

# ── kagglehub (used by Banana Leaf dataset) ──────────────────────────────────
try:
    import kagglehub
    KAGGLEHUB_AVAILABLE = True
except ImportError:
    KAGGLEHUB_AVAILABLE = False

# ── TensorFlow Datasets (optional — only needed for PlantVillage) ────────────
_log("  Importing TensorFlow/tfds (can take 1–2 min) ...")
try:
    import tensorflow_datasets as tfds
    import tensorflow as tf
    TFDS_AVAILABLE = True
    _log("  TensorFlow/tfds ready.")
except ImportError:
    TFDS_AVAILABLE = False
    _log("  TensorFlow/tfds not installed (PlantVillage will be skipped).")

# ── ReportLab ────────────────────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════
#  ★  USER CONFIGURATION  ★
# ═══════════════════════════════════════════════════════════════════════════

# All paths hardcoded under _WORK_ROOT (set at top of file).
_BASE_DIR    = _WORK_ROOT
# Local InternalData root: .../InternalData/
# Expected layout: InternalData/<Category>/<Disease or class folder>/*.JPG
# Example: InternalData/Soybean Diseases/Alfalfa mosaic virus/Soybean Alfalfa mosaic 2 Grau - ....JPG
LOCAL_SOURCE_ROOT = os.path.join(_BASE_DIR, "InternalData")
# CDDM root: .../CDDM-images/images/ (Crop,Disease folders with plant_xxxxx.jpg inside)
# Example: CDDM-images/images/Apple,Alternaria Blotch/plant_69422.jpg
_cddm_base = os.path.join(_BASE_DIR, "CDDM-images")
CDDM_SOURCE_ROOT  = os.path.join(_cddm_base, "images") if os.path.isdir(os.path.join(_cddm_base, "images")) else _cddm_base
# PlantWild v2: .../plantwild_v2/<class_name>/<image>.jpg  (e.g. apple black rot/apple_black_rot_1.jpg)
PLANTWILD_SOURCE_ROOT = os.path.join(_BASE_DIR, "plantwild_v2")
DATA_ROOT    = os.path.join(_BASE_DIR, "data")
DOWNLOAD_ROOT = DATA_ROOT
CURATED_DIR  = os.path.join(_BASE_DIR, "Curated_Dataset")
IMAGES_DIR   = os.path.join(CURATED_DIR, "Images")
SAMPLES_DIR  = CURATED_DIR
OUTPUT_XLSX  = os.path.join(_BASE_DIR, "crop_disease_registry.xlsx")
SUNBURST_DIR = os.path.join(_BASE_DIR, "sunburst_figures")
OUTPUT_SUNBURST_PREFIX = os.path.join(SUNBURST_DIR, "sunburst_crop_disease")
BUGWOOD_IMAGES_ROOT = os.path.join(_BASE_DIR, "Curated_Bugwood_Dataset", "Images")
RANDOM_STATE = 42
IMAGE_EXT    = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp')

# If True, suppress sunburst image-export warnings (Kaleido/Chrome issues).
# HTML output will still be generated when plotly is available.
SUPPRESS_SUNBURST_WARNINGS = True

# Internal: datasets flagged for forced resample on this run.
# If a dataset was previously cached with fewer classes (due to older rules),
# we add it here and rerun its loader once.
FORCE_RESAMPLE_DATASETS = set()

# No CLI: download all images from all sources automatically (no prompts).
# All data is saved to Curated_Dataset/Images/<Crop>/<Disease>/ with source-tagged filenames.

# ── Worker threads: forced single-core/single-worker ──────────────────────────
NUM_WORKERS = 100
_log(f"  [SYSTEM] Work root (all paths under here): {_BASE_DIR}")
_log(f"  [SYSTEM] Ephemeral caches: under {_cache_dir}")
_log(f"  [SYSTEM] Using {NUM_WORKERS} worker threads")

# No interactive prompt — n=None means use all images per class from every source.

# ═══════════════════════════════════════════════════════════════════════════
#  COLOURS
# ═══════════════════════════════════════════════════════════════════════════

DATASET_COLORS = {
    "SBRD":               "#2e9e5e",
    "Mango Leaf":         "#e8794a",
    "Soybean PNAS":       "#4a8ec2",
    "Bean Leaf":          "#9b6fd4",
    "Yellow Rust":        "#d4a82a",
    "FUSARIUM 22":        "#d45858",
    "Banana Leaf":        "#f5c518",
    "Cauliflower":        "#7cb342",
    "Lettuce":            "#00897b",
    "LeafNet":            "#1B5E20",
    "Alfalfa_Diseases":   "#4CAF50",
    "Corn_Diseases":      "#FF9800",
    "Soybean_Diseases":   "#2196F3",
    "Wheat_Diseases":     "#9C27B0",
    "Mango_Leaf_Disease": "#F44336",
}

# Optional metadata for citation tracking in the XLSX registry.
# Maps dataset "name" (as used by loaders) to (URL, paper citation string).
DATASET_METADATA = {
    "SBRD": (
        "https://www.kaggle.com/datasets/isaacritharson/severity-based-rice-leaf-diseases-dataset",
        "Isaac Ritharson. Severity-Based Rice Leaf Diseases Dataset. Kaggle, 2021.",
    ),
    "Mango Leaf": (
        "https://www.kaggle.com/datasets/aryashah2k/mango-leaf-disease-dataset",
        "S. I. Ahmed et al., \"MangoLeafBD: A comprehensive image dataset to classify diseased and healthy mango leaves,\" Data in Brief, vol. 47, p. 108941, 2023. DOI: 10.1016/j.dib.2023.108941.",
    ),
    "Bean Leaf": (
        "https://www.kaggle.com/datasets/marquis03/bean-leaf-lesions-classification",
        "Marquis03. Bean Leaf Lesions Classification Dataset. Kaggle.",
    ),
    "Yellow Rust": (
        "https://www.kaggle.com/datasets/tolgahayit/yellowrust19-yellow-rust-disease-in-wheat",
        "T. Hayit. YellowRust19: Yellow Rust Disease in Wheat. Kaggle, 2020.",
    ),
    "Banana Leaf": (
        "https://www.kaggle.com/datasets/gimrillozarita/banana-leaf-disease-dataset-v1-1",
        "Gimrillo Zarita. Banana Leaf Disease Dataset v1.1. Kaggle, 2022.",
    ),
    "Lettuce": (
        "https://www.kaggle.com/datasets/ashishjstar/lettuce-diseases",
        "Ashish Jstar. Lettuce Diseases Dataset. Kaggle.",
    ),
    "Cucumber": (
        "https://www.kaggle.com/datasets/kareem3egm/cucumber-plant-diseases-dataset",
        "Kareem3egm. Cucumber Plant Diseases Dataset. Kaggle.",
    ),
    "Durian Leaf": (
        "https://www.kaggle.com/datasets/cthng123/durian-leaf-disease-dataset",
        "Cthng123. Durian Leaf Disease Dataset. Kaggle.",
    ),
    "Eggplant Disease": (
        "https://www.kaggle.com/datasets/kamalmoha/eggplant-disease-recognition-dataset",
        "Kamalmoha. Eggplant Disease Recognition Dataset. Kaggle.",
    ),
    "Cotton Disease": (
        "https://www.kaggle.com/datasets/shuvokumarbasak2030/cotton-disease-multi-transformation-dataset",
        "Shuvo Kumar Basak. Cotton Disease Multi Transformation Dataset. Kaggle.",
    ),
    "Pumpkin Leaf": (
        "https://www.kaggle.com/datasets/shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset",
        "Shuvo Kumar Basak. Pumpkin Leaf Disease Multi Transformation Dataset. Kaggle.",
    ),
    "Rose Leaf": (
        "https://www.kaggle.com/datasets/shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset",
        "Shuvo Kumar Basak. Rose Leaf Disease Multi Transformation Dataset. Kaggle.",
    ),
    "StrawberryDiseaseDetection": (
        "https://www.kaggle.com/datasets/usmanafzaal/strawberry-disease-detection-dataset",
        "Usman Afzaal. Strawberry Disease Detection Dataset. Kaggle.",
    ),
    "SugarLeafIDN": (
        "https://www.kaggle.com/datasets/bettydpuspasari/sugarleafidn",
        "Betty D. Puspasari. SugarLeafIDN — Sugarcane Leaf Diseases Dataset. Kaggle.",
    ),
    "FUSARIUM 22": (
        "https://www.kaggle.com/datasets/tolgahayit/fusarium-wilt-disease-in-chickpea-dataset",
        "T. Hayit. Fusarium Wilt Disease in Chickpea Dataset. Kaggle.",
    ),
    "PlantDoc": (
        "https://github.com/pratikkayal/PlantDoc-Dataset",
        "D. Singh et al., \"PlantDoc: A Dataset for Visual Plant Disease Detection,\" Proc. ACM IKDD CoDS and COMAD, 2020. DOI: 10.1145/3371158.3371196.",
    ),
    "PlantVillage": (
        "https://www.tensorflow.org/datasets/catalog/plant_village",
        "D. P. Hughes and M. Salath\u00e9, \"An open access repository of images on plant health to enable the development of mobile disease diagnostics,\" arXiv:1511.08060, 2015.",
    ),
    "LeafNet": (
        "https://arxiv.org/abs/2602.13662",
        "K. N. Quoc, P. D. Dao, and L.-D. Quach, \"LeafNet: A Large-Scale Dataset and Comprehensive Benchmark for Foundational Vision-Language Understanding of Plant Diseases,\" arXiv:2602.13662, 2026.",
    ),
    "Soybean PNAS": (
        "https://zenodo.org/records/12747481",
        "S. Ghosal et al., \"An explainable deep machine vision framework for plant stress phenotyping,\" Proceedings of the National Academy of Sciences (PNAS), vol. 115, no. 18, pp. 4613\u20134618, 2018. DOI: 10.1073/pnas.1716999115.",
    ),
    "PlantWild": (
        "",
        "T. Wei et al., \"Benchmarking In-the-Wild Multimodal Plant Disease Recognition and A Versatile Baseline,\" Proc. ACM Multimedia (MM '24), 2024. DOI: 10.1145/3664647.3680599.",
    ),
}



# ═══════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════════

def safe_name(s):
    return re.sub(r'[^A-Za-z0-9_\-]', '_', s.replace(' ', '_'))

def get_closest_match(name, options):
    m = difflib.get_close_matches(name, options, n=1, cutoff=0.2)
    return m[0] if m else None

def rename_folders(base, expected):
    if not os.path.isdir(base): return
    for f in list(os.listdir(base)):
        fp = os.path.join(base, f)
        if not os.path.isdir(fp): continue
        m = get_closest_match(f, expected)
        if m and m != f:
            np_ = os.path.join(base, m)
            if not os.path.exists(np_): os.rename(fp, np_)

def rename_folders_dict(base, rmap):
    if not os.path.isdir(base): return
    for f in list(os.listdir(base)):
        fp = os.path.join(base, f)
        if os.path.isdir(fp) and f in rmap:
            np_ = os.path.join(base, rmap[f])
            if not os.path.exists(np_): os.rename(fp, np_)

def find_best_class_dir(root, classes, min_match=1):
    """
    Walk `root` recursively and return the directory whose immediate
    subdirectories best match `classes` (case-insensitive prefix/fuzzy).
    Useful when the zip structure is unknown or varies across dataset versions.

    Parameters
    ----------
    root      : str  – top of the extracted archive
    classes   : list – expected class folder names
    min_match : int  – minimum number of class matches required (default 1)

    Returns
    -------
    best directory path, or None if no suitable folder found.
    """
    classes_lower = [c.lower() for c in classes]
    best_path  = None
    best_score = 0

    for dirpath, dirnames, _ in os.walk(root):
        # Skip hidden / system dirs
        dirnames[:] = [d for d in dirnames if not d.startswith('.')]
        subdirs_lower = [d.lower() for d in dirnames if not d.startswith('.')]
        if not subdirs_lower:
            continue
        # Count how many expected classes appear (exact or as a prefix/suffix)
        score = sum(
            1 for cl in classes_lower
            if any(cl in sl or sl in cl for sl in subdirs_lower)
        )
        # Also give credit for directories that contain image files directly
        img_count = sum(
            1 for d in dirnames
            if any(
                f.lower().endswith(IMAGE_EXT)
                for f in os.listdir(os.path.join(dirpath, d))
                if os.path.isfile(os.path.join(dirpath, d, f))
            )
        )
        combined = score * 10 + min(img_count, len(classes))
        if combined > best_score:
            best_score = combined
            best_path  = dirpath

    if best_path and best_score >= min_match * 10:
        return best_path
    return None


def collect_images_flat(folder):
    """All image files directly inside folder (non-recursive)."""
    if not os.path.isdir(folder): return []
    return [os.path.join(folder, f) for f in os.listdir(folder)
            if f.lower().endswith(IMAGE_EXT)
            and os.path.isfile(os.path.join(folder, f))]

def collect_images_df(base_directory):
    """
    Walk class subdirs (recursively) -> DataFrame(path=0, label=1).

    This is intentionally recursive so that if a cache layer or dataset
    update introduces extra nesting (e.g. class/train, class/augmented),
    we still discover all images under each top-level class folder.
    """
    rows = []
    if not os.path.isdir(base_directory):
        return pd.DataFrame(columns=[0, 1])

    for sub in os.listdir(base_directory):
        if sub == ".DS_Store":
            continue
        sp = os.path.join(base_directory, sub)
        if not os.path.isdir(sp):
            continue
        # Recursively walk within each class directory so that cached
        # layouts with extra subdirectories are still fully traversed.
        for dirpath, _, filenames in os.walk(sp):
            for f in filenames:
                if not f.lower().endswith(IMAGE_EXT):
                    continue
                full_path = os.path.join(dirpath, f)
                # Avoid recording broken/corrupted entries that os.walk may surface
                # (e.g., broken symlinks or partially-extracted datasets).
                if not os.path.isfile(full_path):
                    continue
                rows.append({0: full_path, 1: sub})

    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[0, 1])


def _find_class_based_dirs(root_dl, ignore_dirs=None):
    """
    Find all directories under root_dl whose immediate children are class folders
    (each subdir contains image files). Used when train/test splits are not found
    so we can load from whatever structure was downloaded (e.g. single folder or
    nested layout). Returns list of (path, has_images) for dirs that look like
    class-based pools.
    """
    ignore = set((ignore_dirs or [])) | {".DS_Store"}
    found = []
    for dirpath, dirnames, _ in os.walk(root_dl):
        dirnames[:] = [d for d in dirnames if not d.startswith(".") and d not in ignore]
        subs = [d for d in dirnames if os.path.isdir(os.path.join(dirpath, d))]
        if not subs:
            continue
        n_with_images = 0
        for s in subs:
            sp = os.path.join(dirpath, s)
            try:
                files = os.listdir(sp)
            except OSError:
                continue
            if any(f.lower().endswith(IMAGE_EXT) for f in files if os.path.isfile(os.path.join(sp, f))):
                n_with_images += 1
        if n_with_images >= 1 and n_with_images == len(subs):
            found.append(dirpath)
            # Do not descend into this dir's children (we already use this level)
            dirnames.clear()
    return found


# ─── Download helpers ────────────────────────────────────────────────────────

def download_file(url, filename):
    try:
        r = requests.get(url, stream=True, timeout=60)
        total = int(r.headers.get('content-length', 0))
        with open(filename, 'wb') as f, tqdm(desc=os.path.basename(filename),
                total=total, unit='iB', unit_scale=True) as bar:
            for chunk in r.iter_content(1024):
                f.write(chunk); bar.update(len(chunk))
    except Exception as e:
        print(f"  [ERROR] {url}: {e}")

def get_zenodo_urls(record_id):
    from urllib.parse import urlparse, urlunparse, quote
    try:
        r = requests.get(f"https://zenodo.org/api/records/{record_id}", timeout=30)
        if r.status_code == 200:
            out = []
            for f in r.json().get("files", []):
                url = f.get("links", {}).get("self") or ""
                if url:
                    parsed = urlparse(url)
                    url = urlunparse(parsed._replace(path=quote(parsed.path, safe="/")))
                out.append((url, f.get("key", "file")))
            return out
    except Exception as e:
        print(f"  [ERROR] Zenodo: {e}")
    return []

def extract_zip(zpath, dest):
    try:
        with zipfile.ZipFile(zpath, 'r') as zf: zf.extractall(dest)
    except Exception as e:
        print(f"  [ERROR] Extract {zpath}: {e}")

def kaggle_download(dataset_name, path):
    if not KAGGLE_AVAILABLE:
        print(f"  [SKIP] kaggle not installed — {dataset_name}"); return False
    try:
        api = KaggleApi(); api.authenticate()
        print(f"  Downloading: {dataset_name}")
        api.dataset_download_files(dataset_name, path=path, unzip=True, quiet=False)
        return True
    except Exception as e:
        print(f"  [ERROR] Kaggle: {e}"); return False

# ─── Sample & save ───────────────────────────────────────────────────────────

def sample_per_class(data_df, classes, n):
    """
    Take n images per class (or all if n is None).
    When n is None (download-all mode), use all available images per class.
    Classes with fewer than n images are skipped (when n is not None).
    """
    if data_df is None or len(data_df) == 0:
        return pd.DataFrame(columns=[0,1])
    parts = []
    for cls in classes:
        cd = data_df[data_df[1] == cls]
        if n is None:
            if len(cd) == 0:
                continue
            parts.append(cd)
        else:
            if len(cd) < n:
                if len(cd) > 0:
                    print(f"  [SKIP CLASS] '{cls}': only {len(cd)} images, need {n} -- excluded")
                continue
            parts.append(cd.sample(n=n, random_state=RANDOM_STATE))
    if not parts:
        return pd.DataFrame(columns=[0,1])
    return sk_shuffle(pd.concat(parts, ignore_index=True),
                      random_state=RANDOM_STATE).reset_index(drop=True)


def filter_viable_classes(data_df, classes, n):
    """
    Return only classes that have at least n images available (or at least 1 if n is None).
    """
    min_required = 1 if n is None else n
    viable, dropped = [], []
    for cls in classes:
        count = len(data_df[data_df[1] == cls]) if data_df is not None and len(data_df) else 0
        if count >= min_required:
            viable.append(cls)
        else:
            dropped.append((cls, count))
    if dropped and n is not None:
        for cls, count in dropped:
            print(f"  [DROP CLASS] '{cls}': {count} imgs < {n} required -- excluded")
    return viable


# ─── Universal crop/disease resolver (used by save_split and XLSX) ────────
DATASET_CROP = {
    "SBRD":            "Rice",
    "Mango Leaf":      "Mango",
    "Soybean PNAS":    "Soybean",
    "Bean Leaf":       "Bean",
    "Yellow Rust":     "Wheat",
    "FUSARIUM 22":     "Chickpea",
    "Banana Leaf":     "Banana",
    "Cauliflower":     "Cauliflower",
    "Lettuce":         "Lettuce",
    "Cucumber":            "Cucumber",
    "Eggplant Disease":    "Eggplant",
    "Cotton Disease":      "Cotton",
    "Pumpkin Leaf":        "Pumpkin",
    "Rose Leaf":           "Rose",
    "Coconut Disease":     "Coconut",
    "Vanilla Disease":     "Vanilla",
    "SugarLeaf IDN":       "Sugarcane",
    "Cucumber Zenodo":     "Cucumber",
    "Durian Leaf":         "Durian",
    "Strawberry Disease Detection": "Strawberry",
}

CROP_SYNONYMS = {
    "Maize": "Corn",
    "Zea Mays": "Corn",
    "Soyabean": "Soybean",
    "Grapevine": "Grape",
    "Bell": "Bell Pepper",
    "Orange Haunglongbing": "Orange",
    "Orange Huanglongbing": "Orange",
    # Keep crop groups separated (no forced combined canonical groups).
    "Onions": "Onion",
}

# Known canonical crop names, used for heuristics when cleaning mis-parsed
# folder names such as "Grape Esca" or "Orange Huanglongbing".
KNOWN_CROPS = {
    v.strip().title()
    for v in DATASET_CROP.values()
    if isinstance(v, str) and v.strip()
}
KNOWN_CROPS.update(v.strip().title() for v in CROP_SYNONYMS.values())

def _normalize_crop_name(name: str) -> str:
    """
    Map variant crop names to a canonical form so that
    crops like 'Maize' and 'Corn' are merged in the registry.
    """
    if not isinstance(name, str):
        return name
    base = name.strip().title()
    return CROP_SYNONYMS.get(base, base)

# Canonical disease names to deduplicate obvious variants in the registry.
# NOTE: keys should use .title() casing, since _normalize_disease_name()
# normalises via .title() before lookup.
DISEASE_SYNONYMS = {
    # Generic health / leaf variants
    "Healthy Leaf": "Healthy",
    "Healthy Leaves": "Healthy",
    "Leaf Rust": "Rust",
    # "No disease" placeholders (handled as excluded labels downstream)
    "No Disease": "No Disease",
    "No Diseases": "No Disease",
    "No-Disease": "No Disease",
    "No Disease Leaf": "No Disease",
    "No Diseases Leaf": "No Disease",

    # Core issues from issues.txt (critical / high / medium)
    "Frog Eye Leaf Spot": "Frogeye Leaf Spot",
    "Haunglongbing": "Huanglongbing",
    "Haunglongbing Citrus Greening": "Huanglongbing (Citrus Greening)",

    # Section 606–612: generic cross-dataset label merges
    "Bacterial Leaf Spot": "Bacterial Leaf Spot",
    "Bacterial Spot": "Bacterial Leaf Spot",
    "Leaf Bacterial Spot": "Bacterial Leaf Spot",
    "Early Blight": "Early Blight",
    "Early Blight Leaf": "Early Blight",
    "Late Blight": "Late Blight",
    "Leaf Late Blight": "Late Blight",
    "Leaf Mold": "Leaf Mold",
    "Mold Leaf": "Leaf Mold",
    "Leaf Mosaic Virus": "Leaf Mosaic Virus",
    "Mosaic Virus": "Leaf Mosaic Virus",
    "Tomato Mosaic Virus": "Leaf Mosaic Virus",
    "Leaf Yellow Virus": "Leaf Yellow Virus",
    "Yellow Leaf Curl Virus": "Leaf Yellow Virus",
    "Tomato Yellow Leaf Curl Virus": "Leaf Yellow Virus",
    "Spider Mites": "Spider Mites",
    "Spider Mites Two-Spotted Spider Mite": "Spider Mites",
    "Two Spotted Spider Mites Leaf": "Spider Mites",

    # Per-crop mappings from the long synonym section in issues.txt
    # Apple
    "Apple Scab": "Apple Scab",
    "Scab": "Apple Scab",
    "Scab Leaf": "Apple Scab",
    "Rust": "Rust",
    "Rust Leaf": "Rust",
    # Keep "Cedar Apple Rust" distinct (specific rust)

    # Banana
    "Cordana": "Cordana Leaf Spot",
    "Cordana Leaf Spot": "Cordana Leaf Spot",
    "Black Leaf Streak": "Yellow And Black Sigatoka",
    "Yellow And Black Sigatoka": "Yellow And Black Sigatoka",

    # Bean
    "Bean Rust": "Rust",

    # Bell Pepper / Pepper
    "Pepper Leaf Spot": "Pepper Leaf Spot",
    "Pepper Leaf": "Pepper Leaf Spot",
    "Pepper Frogeye Leaf Spot": "Frogeye Leaf Spot",
    "Bell Bacterial Spot": "Bacterial Spot",

    # Broccoli / Cabbage / Cauliflower
    "Alternaria Leaf Spot": "Alternaria Leaf Spot",
    # Downy Mildew variants are already the same string after title()

    # Carrot
    "Alternaria Leaf Blight": "Alternaria Leaf Spot",
    "Cercospora Leaf Blight": "Cercospora Leaf Spot",

    # Cassava
    "Mosaic Virus Disease": "Leaf Mosaic Virus",

    # Chickpea (phenotypes handled via DISEASE_EXCLUDE)

    # Corn
    "Anthracnose Leaf Spot": "Corn Anthracnose",
    "Anthracnose On Seedling": "Corn Anthracnose",
    "Anthracnose Stalk Rot": "Corn Anthracnose",
    "Northern Corn Leaf Blight": "Northern Leaf Blight",
    "Northern Leaf Blight": "Northern Leaf Blight",
    "Rust Leaf": "Rust",
    "Fusarium": "Fusarium Disease",
    "Fusarium Ear Rot": "Fusarium Disease",
    "Fusarium Stalk Rot": "Fusarium Disease",
    "Gibberella": "Gibberella Disease",
    "Gibberella Ear Rot": "Gibberella Disease",
    "Gibberella Stalk Rot": "Gibberella Disease",

    # Cucumber: text-identical after title()

    # Eggplant
    "Leaf Spot Disease": "Cercospora Leaf Spot",

    # Grape
    "Black Rot": "Black Rot",
    "Leaf Black Rot": "Black Rot",

    # Lettuce
    "Downy Mildew On Lettuce": "Downy Mildew",
    "Powdery Mildew On Lettuce": "Powdery Mildew",
    "Septoria Blight On Lettuce": "Septoria Blight",

    # Orange
    "Citrus Greening": "Huanglongbing",
    "Huanglongbing": "Huanglongbing",

    # Pepper / Potato / Pumpkin / Squash
    "Powdery Mildew Leaf": "Powdery Mildew",
    "Leaf Early Blight": "Early Blight",
    "Leaf Late Blight": "Late Blight",

    # Rice severities (treat as one disease each)
    # NOTE: Severity labels must remain distinct classes; handled in
    # _normalize_disease_name() by preserving any label starting with
    # \"Mild \" or \"Severe \". Keep these mappings disabled.

    # Soybean
    "Mosaic": "Soybean Mosaic Virus",
    "Soybean Mosaic Virus": "Soybean Mosaic Virus",
    "Frogeye Leaf Spot": "Frogeye Leaf Spot",

    # Strawberry
    "Anthracnose Fruit Rot": "Anthracnose",
    "Angular Leafspot": "Leaf Spot",

    # Tomato
    "Bacterial Leaf Spot": "Bacterial Leaf Spot",
    "Early Blight Leaf": "Early Blight",
    "Powdery Mildew Leaf": "Powdery Mildew",
    "Powdery Mildew Fruit": "Powdery Mildew",
    "Two Spotted Spider Mites Leaf": "Spider Mites",

    # Wheat
    "Septoria Blotch": "Septoria Leaf Blotch",
    "Septoria Leaf Blotch": "Septoria Leaf Blotch",
    "Septoria Leaf Spot": "Septoria Leaf Blotch",
}

def _normalize_disease_name(name: str) -> str:
    """
    Clean and canonicalise disease names so that:
      - obvious misspellings and variants are merged
      - dataset/category bleed-ins (double-space patterns) are stripped
    """
    if not isinstance(name, str):
        return name
    raw = name.strip()
    if not raw:
        return raw

    # Handle dataset/category bleed-ins that use a double-space separator.
    # Heuristic:
    #   - If prefix looks like a generic category (Ear Rots, Leaf Blight, Maize, Bell, Root Rot),
    #     keep the suffix as the specific disease.
    #   - Otherwise, keep the prefix and ignore the suffix.
    if "  " in raw:
        parts = [p.strip() for p in raw.split("  ") if p.strip()]
        if len(parts) >= 2:
            prefix, suffix = parts[0], parts[1]
            generic_prefixes = {
                "Ear Rots",
                "Leaf Blight",
                "Maize",
                "Bell",
                "Root Rot",
            }
            if prefix.title() in generic_prefixes:
                raw = suffix
            else:
                raw = prefix

    raw = raw.strip()

    # Preserve code-like labels exactly (do not title-case):
    # Examples: "MRMS", "HR", "HS", "SCSMV"
    if re.fullmatch(r"[A-Z0-9]+", raw):
        return raw

    base = raw.title()

    # Preserve severity-qualified labels exactly (do not merge them into a base disease).
    # Example: "Mild Blast" and "Severe Blast" must remain distinct classes.
    if base.startswith("Mild ") or base.startswith("Severe "):
        return base

    return DISEASE_SYNONYMS.get(base, base)

# Diseases that are not true disease classes and should be excluded
# from the registry (e.g. generic "Healthy", "Other", etc.).
DISEASE_EXCLUDE = {
    "Healthy",
    "No Disease",
    "Other",
    "Leaf",
    "Mites",
}

# Safety switch: if True, we only drop labels in DISEASE_EXCLUDE (and EXCLUDED_CLASSES).
# We still normalise spellings/aliases into canonical names, but we do not intentionally
# drop any additional classes.
DROP_ONLY_EXPLICIT_EXCLUDES = True

# Post-processing filters for the XLSX registry.
# Any (crop, disease) pair listed here will be dropped from the
# final Excel output. This is intended as a hard-coded override
# that you can edit as project needs evolve.
#
# Example to exclude all "Unknown" diseases for a crop:
#   ("Corn", "Unknown"),
# Example to exclude a specific noisy or placeholder class:
#   ("Apple", "Background"),
# Hard DROP list — anomalous / non-target classes (curated folders + registry).
EXCLUDED_CLASSES = {
    ("Bell Pepper", "Pepper Healthy"),
    ("Bell Pepper", "Bell Healthy"),
    ("Cucumber", "Fresh Cucumber"),
    ("Cucumber", "Fresh Leaf"),
    ("Cucumber", "Good Cucumber"),
    ("Cucumber", "Ill Cucumber"),
    ("Mango", "Cutting Weevil"),
    ("Mango", "Gall Midge"),
    ("Cotton", "Aphids"),
    ("Cotton", "Army Worm"),
    ("Coconut", "Cci Caterpillars"),
    ("Lettuce", "Shepherd Purse Weeds"),
    ("Lettuce", "Bacterial"),
    ("Lettuce", "Viral"),
    ("Corn", "Misc"),
    ("Corn", "Multiple Foliar Diseases At Once"),
    ("Eggplant", "Insect Pest Disease"),
}

# After normalisation: merge/rename into canonical disease names (same crop).
CLASS_MERGE_RENAME = {
    ("Sugarcane", "Pokkahboeng"): "Pokkah Boeng",
    ("Sugarcane", "Soybean Mosaic Virus"): "Sugarcane Mosaic Virus",
    ("Alfalfa", "Soybean Mosaic Virus"): "Alfalfa Mosaic Virus",
    ("Corn", "Cercospora Leaf Spot Gray Leaf Spot"): "Gray Leaf Spot",
    ("Corn", "Maize Streak Virus - South Africa 2013 - Daren Mueller"): "Maize Streak Virus",
    ("Corn", "Corn Anthracnose"): "Anthracnose Leaf Spot And Top Dieback",
    ("Peach", "Apple Scab"): "Leaf Scab",
    ("Cherry", "Including Sour Powdery Mildew"): "Powdery Mildew",
    # Some sources concatenate words in the folder name
    ("Cherry", "Including Sourpowdery Mildew"): "Powdery Mildew",
    ("Grape", "Isariopsis Leaf Spot"): "Leaf Blight",
}

# Optional safe pair-level merge map generated by matching workflow.
# File format expected: proper_matching_results.csv
# Columns: crop_s1,disease_s1,crop_match,disease_match,status,score
SAFE_MATCHING_FILE = os.path.join(_BASE_DIR, "proper_matching_results.csv")
SAFE_MATCH_STATUSES = {"exact", "fuzzy_high"}
_SAFE_PAIR_MAP_CACHE = None


def _norm_key_text(value: str) -> str:
    s = str(value or "").strip().lower().replace("_", " ")
    s = re.sub(r"[^a-z0-9\\s]+", " ", s)
    return re.sub(r"\\s+", " ", s).strip()


def _load_safe_pair_map() -> dict:
    """
    Build (crop,disease) -> (crop,disease) remap from proper_matching_results.csv.
    Uses only high-confidence statuses to keep merges safe.
    """
    global _SAFE_PAIR_MAP_CACHE
    if _SAFE_PAIR_MAP_CACHE is not None:
        return _SAFE_PAIR_MAP_CACHE

    mapping = {}
    if not os.path.isfile(SAFE_MATCHING_FILE):
        _SAFE_PAIR_MAP_CACHE = mapping
        return mapping

    try:
        df = pd.read_csv(SAFE_MATCHING_FILE)
    except Exception:
        _SAFE_PAIR_MAP_CACHE = mapping
        return mapping

    required = {"crop_s1", "disease_s1", "crop_match", "disease_match", "status"}
    if not required.issubset(set(df.columns)):
        _SAFE_PAIR_MAP_CACHE = mapping
        return mapping

    for _, r in df.iterrows():
        status = str(r.get("status", "")).strip().lower()
        if status not in SAFE_MATCH_STATUSES:
            continue
        c1 = _norm_key_text(r.get("crop_s1", ""))
        d1 = _norm_key_text(r.get("disease_s1", ""))
        c2 = str(r.get("crop_match", "")).strip()
        d2 = str(r.get("disease_match", "")).strip()
        if not c1 or not d1 or not c2 or not d2:
            continue
        mapping[(c1, d1)] = (c2, d2)

    _SAFE_PAIR_MAP_CACHE = mapping
    return mapping


def _normalise_crop_only(crop, strip_suffix=True):
    """Strip dataset-style suffixes, then apply CROP_SYNONYMS / title rules."""
    c = str(crop or "").strip()
    if strip_suffix:
        for suffix in (" Diseases", " Disease"):
            if c.endswith(suffix):
                c = c[:-len(suffix)].strip()
                break
    return _normalize_crop_name(c)


def _finalize_registry_pair(crop, disease):
    """
    Normalise crop and disease, apply CLASS_MERGE_RENAME, re-normalise disease.
    Single entry point for save_split, disk migration, and XLSX/registry views.
    """
    c = _normalise_crop_only(crop)
    d = _normalize_disease_name(str(disease or "").strip())

    # Keep resistance-scale labels separate from true disease names.
    # This follows the merge report recommendation for Chickpea/Wheat.
    if c in {"Chickpea", "Wheat"} and d in {
        "Highly Resistant",
        "Resistant",
        "Moderately Resistant",
        "Mrms",
        "Moderately Susceptible",
        "Susceptible",
        "Highly Susceptible",
        "No Disease",
    }:
        d = f"Resistance Phenotype: {d}"

    # Apply optional safe pair-level remap (exact + high-confidence only).
    safe_map = _load_safe_pair_map()
    key = (_norm_key_text(c), _norm_key_text(d))
    mapped_pair = safe_map.get(key)
    if mapped_pair is not None:
        c = _normalize_crop_name(mapped_pair[0])
        d = _normalize_disease_name(mapped_pair[1])

    mapped = CLASS_MERGE_RENAME.get((c, d))
    if mapped is not None:
        d = _normalize_disease_name(mapped)
    return c, d


def migrate_curated_image_tree(base_dir):
    """
    Normalise Curated_Dataset/Images: crop folder fixes, disease synonyms,
    CLASS_MERGE_RENAME merges, and remove EXCLUDED_CLASSES / DISEASE_EXCLUDE.
    Idempotent. Call when the tree (or registry exports) already exists and
    rules have been updated.
    """
    if not os.path.isdir(base_dir):
        return
    for crop_dir in list(os.listdir(base_dir)):
        crop_path = os.path.join(base_dir, crop_dir)
        if not os.path.isdir(crop_path):
            continue
        raw_crop = crop_dir.replace("_", " ").title()
        normalised_crop = _normalise_crop_only(raw_crop)
        parts = raw_crop.split()
        if len(parts) > 1:
            prefix = _normalise_crop_only(parts[0].title())
            if prefix in KNOWN_CROPS:
                normalised_crop = prefix

        crop_target = os.path.join(base_dir, safe_name(normalised_crop))
        if crop_target != crop_path:
            os.makedirs(crop_target, exist_ok=True)
        else:
            crop_target = crop_path

        for disease_dir in list(os.listdir(crop_path)):
            src = os.path.join(crop_path, disease_dir)
            if not os.path.isdir(src):
                continue
            raw_dis = disease_dir.replace("_", " ").title()
            fin_crop, fin_dis = _finalize_registry_pair(normalised_crop, raw_dis)
            if fin_dis in DISEASE_EXCLUDE or (fin_crop, fin_dis) in EXCLUDED_CLASSES:
                shutil.rmtree(src, ignore_errors=True)
                continue

            dest_crop_dir = os.path.join(base_dir, safe_name(fin_crop))
            os.makedirs(dest_crop_dir, exist_ok=True)
            dst = os.path.join(dest_crop_dir, safe_name(fin_dis))
            if os.path.abspath(src) == os.path.abspath(dst):
                continue
            if os.path.exists(dst):
                for f in os.listdir(src):
                    fsrc = os.path.join(src, f)
                    fdst = os.path.join(dst, f)
                    if not os.path.exists(fdst):
                        shutil.move(fsrc, fdst)
                shutil.rmtree(src, ignore_errors=True)
            else:
                shutil.move(src, dst)

        if crop_target != crop_path:
            try:
                os.rmdir(crop_path)
                print(f"  [MIGRATE] Crop folder {crop_dir!r} -> {normalised_crop!r}")
            except OSError:
                pass


def apply_curated_postprocess_if_present():
    """If Curated_Dataset/Images exists, apply drop/merge/rename rules on disk."""
    if os.path.isdir(IMAGES_DIR):
        migrate_curated_image_tree(IMAGES_DIR)

def _parse_crop_disease_from_label(ds_name, cls_label):
    """
    Universal (crop, disease) resolver used by save_split and XLSX.

    Rules:
      LeafNet / CDDM  — class label is safe_name("Crop_Disease"):
                         first token = crop, remainder = disease.
      Fixed-crop sets — crop from DATASET_CROP, label = disease.
      Local / unknown — dataset category name = crop, label = disease.
    """
    if ds_name == "PlantWild":
        # PlantWild v2 class folders are typically human-readable strings like
        # "apple black rot" or "bell pepper bacterial spot". Use a crop-vocabulary
        # prefix match (supports multi-word crops) to avoid mis-parsing crops like
        # "Bell Pepper" as crop="Bell", disease="Pepper ...".
        raw_label = str(cls_label or "").strip().replace("_", " ")
        if not raw_label:
            return "Unknown", "Unknown"

        # Build a small known-crops vocabulary from configured datasets/synonyms.
        # (Keep lightweight; do not scan disk here.)
        known = set()
        for v in DATASET_CROP.values():
            if isinstance(v, str) and v.strip():
                known.add(v.strip().lower())
        for k, v in CROP_SYNONYMS.items():
            if isinstance(k, str) and k.strip():
                known.add(k.strip().lower())
            if isinstance(v, str) and v.strip():
                known.add(v.strip().lower())

        toks = [t for t in raw_label.split() if t]
        # Try longest prefix (3->2->1) that matches known crop.
        crop = None
        disease = None
        for n_tok in (3, 2, 1):
            if len(toks) <= n_tok:
                continue
            cand = " ".join(toks[:n_tok]).lower()
            if cand in known:
                crop = " ".join(toks[:n_tok])
                disease = " ".join(toks[n_tok:])
                break
        if crop is None:
            # Fallback: first token crop, rest disease
            if len(toks) >= 2:
                crop = toks[0]
                disease = " ".join(toks[1:])
            else:
                crop = raw_label
                disease = "Unknown"
        return _normalize_crop_name(crop.title()), disease.title()

    if ds_name in ("LeafNet", "CDDM", "PlantVillage", "New Plant Diseases"):
        # These sources often encode "Crop + Disease" into a single class label,
        # but separators vary across versions:
        # - PlantVillage/New Plant Diseases commonly use "___"
        # - some loaders use "__"
        # - others use single "_" (safe_name output)
        # - PlantWild/CDDM may have human names with spaces ("Grape Esca")
        raw_label = str(cls_label or "").strip()
        if not raw_label:
            return "Unknown", "Unknown"

        if "___" in raw_label:
            parts = raw_label.split("___", 1)
        elif "__" in raw_label:
            parts = raw_label.split("__", 1)
        elif "_" in raw_label:
            parts = raw_label.split("_", 1)
        else:
            # Fallback for space-delimited labels like "Grape Esca"
            parts = raw_label.split(" ", 1)

        if len(parts) == 2 and parts[0].strip() and parts[1].strip():
            crop = _normalize_crop_name(parts[0].replace("_", " ").strip().title())
            disease = parts[1].replace("_", " ").strip().title()
            return crop, disease

        crop = _normalize_crop_name(raw_label.replace("_", " ").strip().title())
        return crop, "Unknown"

    # PlantDoc: "Apple Scab Leaf" -> crop="Apple", disease="Scab Leaf"
    if ds_name == "PlantDoc":
        parts = cls_label.replace("_", " ").strip().split(" ", 1)
        if len(parts) == 2:
            crop = _normalize_crop_name(parts[0].title())
            return crop, parts[1].title()
        crop = _normalize_crop_name(cls_label.replace("_", " ").title())
        return crop, "Unknown"

    crop = DATASET_CROP.get(ds_name)
    if crop:
        crop = _normalize_crop_name(crop)
        return crop, cls_label.replace("_", " ").strip().title()

    # Local/internal datasets: ds_name is e.g. "Soybean_Diseases"
    # Strip trailing _Diseases/_Disease so it merges with online datasets of the same crop
    raw_crop = ds_name.replace("_", " ").title()
    for suffix in (" Diseases", " Disease"):
        if raw_crop.endswith(suffix):
            raw_crop = raw_crop[:-len(suffix)].strip()
            break
    raw_crop = _normalize_crop_name(raw_crop)
    return raw_crop, cls_label.replace("_", " ").strip().title()

def save_split(sampled_df, dataset_name, split_name):
    """
    Copy images to Curated_Dataset/Images/<Crop>/<Disease>/
    Files are named <src_tag>_<idx>.<ext> so we can track which source each image came from.
    src_tag = first 12 chars of safe_name(dataset_name).
    """
    if sampled_df is None or len(sampled_df) == 0:
        return pd.DataFrame(columns=[0, 1])

    src_tag = safe_name(dataset_name)[:12]
    out_rows = []

    for _, row in sampled_df.iterrows():
        src_path = row[0]
        cls_key  = row[1]

        crop, disease = _parse_crop_disease_from_label(dataset_name, cls_key)
        crop, disease = _finalize_registry_pair(crop, disease)

        # Hard drop: do not write "Healthy" or other non-disease/placeholder labels
        # into the curated dataset at all.
        if DROP_ONLY_EXPLICIT_EXCLUDES:
            if disease in DISEASE_EXCLUDE or (crop, disease) in EXCLUDED_CLASSES:
                continue
        else:
            if disease in DISEASE_EXCLUDE or (crop, disease) in EXCLUDED_CLASSES:
                continue

        dest_dir = os.path.join(IMAGES_DIR, safe_name(crop), safe_name(disease))
        os.makedirs(dest_dir, exist_ok=True)

        ext = os.path.splitext(src_path)[1].lower() or ".jpg"

        # Next index for this source in this folder (files: SBRD_1.jpg, SBRD_2.jpg, ...)
        existing = [f for f in os.listdir(dest_dir)
                    if f.startswith(src_tag + "_") and f.lower().endswith(IMAGE_EXT)]
        idx = len(existing) + 1
        dest = os.path.join(dest_dir, f"{src_tag}_{idx}{ext}")
        while os.path.exists(dest):
            idx += 1
            dest = os.path.join(dest_dir, f"{src_tag}_{idx}{ext}")

        try:
            shutil.copy2(src_path, dest)
            out_rows.append({0: dest, 1: cls_key})
        except Exception as e:
            print(f"  [WARN] Could not copy {src_path}: {e}")

    return pd.DataFrame(out_rows) if out_rows else pd.DataFrame(columns=[0, 1])


def split_save_cleanup(data_df, classes, dataset_name, download_path, n):
    """
    Sample n images per class from a single pool and save to
    Curated_Dataset/Images/<Crop>/<Disease>/.
    """
    viable = filter_viable_classes(data_df, classes, n)
    if not viable:
        print(f"  [DROP ALL] No viable classes for {dataset_name}")
        empty = pd.DataFrame(columns=[0, 1])
        if download_path and os.path.exists(download_path):
            shutil.rmtree(download_path, ignore_errors=True)
        return empty, empty

    sampled  = sample_per_class(data_df, viable, n)
    saved_df = save_split(sampled, dataset_name, "images")   # split_name ignored

    n_saved = len(saved_df)
    print(f"  [SAVED] {n_saved} images "
          f"-> Curated_Dataset/Images/{safe_name(dataset_name)}/")

    if download_path and os.path.exists(download_path):
        shutil.rmtree(download_path, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {download_path}")

    return saved_df, pd.DataFrame(columns=[0, 1])   # second value kept for API compat


def already_sampled(dataset_name, n=None):
    """
    Check if this dataset has already been sampled into the Curated_Dataset.
    Since all datasets now share a flat Crop/Disease hierarchy, we detect
    presence by looking for any image file whose filename contains the
    dataset source tag.

    If n provided, also checks every disease folder that
    belongs to this dataset has enough images.
    """
    src_tag = safe_name(dataset_name)[:12]

    # Never treat PlantWild as cached; always resample it.
    if dataset_name == "PlantWild":
        return False

    # Allow caller to force a re-sample on this run.
    if dataset_name in FORCE_RESAMPLE_DATASETS:
        return False

    # Collect all disease-level folders that have files tagged with src_tag
    found_dirs = {}   # disease_dir_path -> count of matching images
    for base_dir in [IMAGES_DIR]:
        if not os.path.isdir(base_dir):
            continue
        for crop_dir in os.listdir(base_dir):
            crop_path = os.path.join(base_dir, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in os.listdir(crop_path):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                count = sum(
                    1 for f in os.listdir(disease_path)
                    if f.lower().endswith(IMAGE_EXT) and src_tag in f
                )
                if count > 0:
                    found_dirs[disease_path] = count

    if not found_dirs:
        return False   # nothing saved yet for this dataset

    if n is None:
        return True    # just existence check — passed

    # Count-aware: every disease folder for this source must meet threshold
    required = n or 0
    for disease_path, count in found_dirs.items():
        if count < required:
            print(f"  [RESAMPLE] {dataset_name} -> {os.path.basename(disease_path)}: "
                  f"{count} images < {required} required — will re-sample")
            return False
    return True

def load_from_samples(dataset_name, classes):
    """
    Rebuild saved_df / saved_df from already-saved Curated_Dataset folders.
    Images now live at  <IMAGES_DIR>/<Crop>/<Disease>/
    and are tagged with the dataset source in their filename.
    """
    src_tag   = safe_name(dataset_name)[:12]
    img_rows  = []

    for base_dir, rows in [(IMAGES_DIR, img_rows)]:
        if not os.path.isdir(base_dir):
            continue
        for crop_dir in sorted(os.listdir(base_dir)):
            crop_path = os.path.join(base_dir, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in sorted(os.listdir(crop_path)):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                for f in sorted(os.listdir(disease_path)):
                    if f.lower().endswith(IMAGE_EXT) and src_tag in f:
                        # Reconstruct class label as "Crop_Disease" for consistency
                        cls = f"{crop_dir}__{disease_dir}"
                        rows.append({0: os.path.join(disease_path, f), 1: cls})

    saved_df = pd.DataFrame(img_rows) if img_rows else pd.DataFrame(columns=[0,1])
    return saved_df, saved_df

# ═══════════════════════════════════════════════════════════════════════════
#  ONLINE DATASET LOADERS
# ═══════════════════════════════════════════════════════════════════════════

def load_SBRD(n):
    name    = "SBRD"
    classes = ['Healthy','Mild Bacterial Blight','Mild Blast','Mild Brownspot',
               'Mild Tungro','Severe Bacterial Blight','Severe Blast',
               'Severe Brownspot','Severe Tungro']
    desc    = ("Rice leaf disease dataset with severity levels covering Bacterial "
               "Blight, Blast, Brownspot and Tungro at Mild and Severe stages.")
    if already_sampled(name, n):
        # Validate stale caches: if the on-disk classes for this dataset are fewer
        # than expected (after normalisation/exclusion), force a resample. This
        # protects against older runs that merged severity labels (Mild/Severe).
        src_tag = safe_name(name)[:12]

        def _expected_pairs():
            exp = set()
            for cls in classes:
                crop, disease = _parse_crop_disease_from_label(name, cls)
                crop, disease = _finalize_registry_pair(crop, disease)
                if disease in DISEASE_EXCLUDE or (crop, disease) in EXCLUDED_CLASSES:
                    continue
                exp.add((crop, disease))
            return exp

        def _disk_pairs():
            pairs = set()
            if not os.path.isdir(IMAGES_DIR):
                return pairs
            for crop_dir in os.listdir(IMAGES_DIR):
                crop_path = os.path.join(IMAGES_DIR, crop_dir)
                if not os.path.isdir(crop_path):
                    continue
                for disease_dir in os.listdir(crop_path):
                    disease_path = os.path.join(crop_path, disease_dir)
                    if not os.path.isdir(disease_path):
                        continue
                    # consider this pair part of SBRD if any file has src_tag
                    try:
                        has_any = any(
                            f.lower().endswith(IMAGE_EXT) and src_tag in f
                            for f in os.listdir(disease_path)
                        )
                    except OSError:
                        has_any = False
                    if not has_any:
                        continue
                    crop, disease = _finalize_registry_pair(
                        crop_dir.replace("_", " ").title(),
                        disease_dir.replace("_", " ").title(),
                    )
                    if disease in DISEASE_EXCLUDE or (crop, disease) in EXCLUDED_CLASSES:
                        continue
                    pairs.add((crop, disease))
            return pairs

        exp = _expected_pairs()
        disk = _disk_pairs()
        if len(disk) < len(exp):
            print(f"  [RESAMPLE] {name} — disk has {len(disk)} class(es), expected {len(exp)} — re-downloading to restore classes.")

            # Remove existing tagged images so we can re-create cleanly
            if os.path.isdir(IMAGES_DIR):
                for crop_dir in list(os.listdir(IMAGES_DIR)):
                    crop_path = os.path.join(IMAGES_DIR, crop_dir)
                    if not os.path.isdir(crop_path):
                        continue
                    for disease_dir in list(os.listdir(crop_path)):
                        disease_path = os.path.join(crop_path, disease_dir)
                        if not os.path.isdir(disease_path):
                            continue
                        try:
                            for f in list(os.listdir(disease_path)):
                                if f.lower().endswith(IMAGE_EXT) and src_tag in f:
                                    try:
                                        os.remove(os.path.join(disease_path, f))
                                    except OSError:
                                        pass
                        except OSError:
                            pass
                        # prune empty disease dir
                        try:
                            if not any(fn.lower().endswith(IMAGE_EXT) for fn in os.listdir(disease_path)):
                                os.rmdir(disease_path)
                        except OSError:
                            pass
                    # prune empty crop dir
                    try:
                        if not any(os.path.isdir(os.path.join(crop_path, d)) for d in os.listdir(crop_path)):
                            os.rmdir(crop_path)
                    except OSError:
                        pass
        else:
            print(f"  [SKIP] {name} — samples exist, loading from disk.")
            saved, _ = load_from_samples(name, classes)
            return name, saved, saved, classes, desc
    dl   = os.path.join(DATA_ROOT, "SBRD")
    base = os.path.join(dl, "Leaf Disease Dataset", "train")
    if not os.path.exists(dl):
        kaggle_download("isaacritharson/severity-based-rice-leaf-diseases-dataset", dl)
    rename_folders(base, classes)
    data = collect_images_df(base)
    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_MangoLeaf(n):
    name    = "Mango Leaf"
    classes = ['Anthracnose','Bacterial Canker','Cutting Weevil','Die Back',
               'Gall Midge','Healthy','Powdery Mildew','Sooty Mould']
    desc    = ("Mango leaf disease dataset covering 7 diseases plus Healthy. "
               "Includes fungal, bacterial and pest-related leaf conditions.")
    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc
    dl = os.path.join(DATA_ROOT, "mango-leaf-disease-dataset")
    if not os.path.exists(dl):
        kaggle_download("aryashah2k/mango-leaf-disease-dataset", dl)
    rename_folders(dl, classes)
    data = collect_images_df(dl)
    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_SoybeanPNAS(n):
    name    = "Soybean PNAS"
    classes = ['Bacterial Blight','Bacterial Pustule','Frogeye Leaf Spot',
               'Healthy','Herbicide Injury','Iron Deficiency Chlorosis',
               'Potassium Deficiency','Septoria Brown Spot','Sudden Death Syndrome']
    rmap    = {str(i): c for i, c in enumerate(classes)}
    desc    = ("Soybean stress identification from PNAS. Covers 8 stress/disease "
               "conditions plus Healthy including nutrient deficiencies and "
               "fungal, bacterial and environmental stressors.")
    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc
    dl   = os.path.join(DATA_ROOT, "Soybean-PNAS")
    base = os.path.join(dl, "Training Samples")
    if not os.path.exists(dl):
        os.makedirs(dl, exist_ok=True)
        orig = os.getcwd(); os.chdir(dl)
        for url, fname in get_zenodo_urls("12747481"): download_file(url, fname)
        os.chdir(orig)
        for f in os.listdir(dl):
            if f.lower() == "soybean_stress_identification.zip":
                extract_zip(os.path.join(dl, f), dl)
    rename_folders_dict(base, rmap)
    data = collect_images_df(base)
    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_BeanLeaf(n):
    name    = "Bean Leaf"
    classes = ['Angular Leaf Spot','Bean Rust','Healthy']
    desc    = ("Bean leaf lesion classification: Angular Leaf Spot (bacterial), "
               "Bean Rust (fungal) and Healthy. Compact well-balanced dataset.")
    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc
    dl   = os.path.join(DATA_ROOT, "bean-leaf-lesions-classification")
    base = os.path.join(dl, "train")
    if not os.path.exists(dl):
        kaggle_download("marquis03/bean-leaf-lesions-classification", dl)
    rename_folders(base, classes)
    data = collect_images_df(base)
    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_YellowRust(n):
    name = "Yellow Rust"
    # Numeric folder names found in the actual zip (RAW/RAW/0, 1, 2 ...)
    # mapped to meaningful class labels.
    rmap = {
        '0': 'No Disease',
        '1': 'Resistant (R)',
        '2': 'Moderately Resistant (MR)',
        '3': 'MRMS',
        '4': 'Moderately Susceptible (MS)',
        '5': 'Susceptible (S)',
        # legacy letter-named folders (some versions of the dataset)
        'MR':   'Moderately Resistant (MR)',
        'MS':   'Moderately Susceptible (MS)',
        'MRMS': 'MRMS',
        'R':    'Resistant (R)',
        'S':    'Susceptible (S)',
    }
    classes = list(dict.fromkeys(rmap.values()))   # unique, insertion-ordered
    desc    = ("Yellow Rust 19 wheat disease dataset. Labelled by resistance level: "
               "Resistant -> Moderately Resistant -> MRMS -> Moderately Susceptible "
               "-> Susceptible, plus No Disease.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    dl = os.path.join(DATA_ROOT, "yellowrust19")
    if not os.path.exists(dl):
        kaggle_download("tolgahayit/yellowrust19-yellow-rust-disease-in-wheat", dl)

    # The zip extracts to varying paths across versions:
    #   YELLOW-RUST-19/YELLOW-RUST-19/  (old)
    #   RAW/RAW/                        (new — confirmed in screenshot)
    # Use find_best_class_dir to locate whichever folder holds the class subfolders.
    base = find_best_class_dir(dl, list(rmap.keys()))
    if base is None:
        # Fallback: find the deepest directory that has numeric or letter subfolders
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            subs = [d for d in dirnames
                    if d.isdigit() or d in rmap]
            if subs:
                base = dirpath
                break
    if base is None:
        print(f"  [ERROR] Could not locate YellowRust class folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc


    print(f"  [YellowRust] Using folder: {base}")
    print(f"  [YellowRust] Subfolders: {sorted(os.listdir(base))}")

    rename_folders_dict(base, rmap)

    # After renaming, discover which classes actually exist on disk
    actual_classes = [
        c for c in classes
        if os.path.isdir(os.path.join(base, safe_name(c))) or
           os.path.isdir(os.path.join(base, c))
    ]
    if not actual_classes:
        actual_classes = classes   # fall back to full list

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    present = data[1].unique().tolist()
    actual_classes = [c for c in classes if c in present]

    saved, _ = split_save_cleanup(data, actual_classes, name, dl, n)
    return name, saved, saved, actual_classes, desc


def load_BananaLeaf(n):
    """
    Banana Leaf Disease Dataset v1.1 — gimrillozarita/banana-leaf-disease-dataset-v1-1
    4 classes (actual Kaggle structure):
      Cordana, Healthy, Panama Disease, Yellow and Black Sigatoka
    Class folders sit directly in the dataset root — there is no augmented subfolder.
    """
    name    = "Banana Leaf"
    classes = ['Cordana', 'Healthy', 'Panama Disease', 'Yellow and Black Sigatoka']
    desc    = ("Banana leaf disease dataset (v1.1) covering three banana leaf diseases: "
               "Cordana, Panama Disease, Yellow and Black Sigatoka, plus Healthy. "
               "Class images are organised directly under the dataset root folder.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    base = None

    # ── Download via kagglehub (preferred) ──────────────────────────────────
    if KAGGLEHUB_AVAILABLE:
        try:
            print(f"  [kagglehub] Downloading banana-leaf-disease-dataset-v1-1 ...")
            dl_path = kagglehub.dataset_download(
                "gimrillozarita/banana-leaf-disease-dataset-v1-1")
            print(f"  Path to dataset files: {dl_path}")

            # Class folders (Cordana, Healthy, Panama Disease, …) sit directly
            # under the dataset root — there is NO augmented subfolder.
            # Use find_best_class_dir to handle any nesting the cache may add.
            best = find_best_class_dir(dl_path, classes)
            base = best if best else dl_path
        except Exception as e:
            print(f"  [kagglehub ERROR] {e}")

    # ── Fallback: kaggle API ──────────────────────────────────────────────────
    if base is None:
        print(f"  [FALLBACK] Trying kaggle API ...")
        dl = os.path.join(DATA_ROOT, "banana-leaf-disease")
        if not os.path.exists(dl):
            kaggle_download("gimrillozarita/banana-leaf-disease-dataset-v1-1", dl)

        # Use best-match search — no augmented subfolder in this dataset
        base = find_best_class_dir(dl, classes)
        if base is None:
            base = dl   # absolute last resort

    if base is None or not os.path.isdir(base):
        print(f"  [ERROR] Could not locate Banana Leaf dataset folder.")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [BananaLeaf] Using folder: {base}")
    # Show what's inside to aid debugging
    try:
        contents = [d for d in os.listdir(base)
                    if os.path.isdir(os.path.join(base, d)) and not d.startswith('.')]
        print(f"  [BananaLeaf] Subfolders found: {contents}")
    except Exception:
        pass

    # Normalise folder names to match classes list (case-insensitive)
    rename_folders(base, classes)

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # kagglehub downloads to a cache dir — we don't own it, so don't delete it.
    saved_df = sample_per_class(data, classes, n)
    saved_df = save_split(saved_df, name, "images")
    print(f"  [SAVED] {len(saved_df)} images -> Curated_Dataset/Images/...")
    return name, saved_df, saved_df, classes, desc


def load_Lettuce(n):
    """
    Lettuce Diseases dataset — ashishjstar/lettuce-diseases
    Downloaded via kagglehub (falls back to kaggle API).
    Root: Lettuce_disease_datasets/
    8 classes (auto-discovered from subfolders):
      Bacterial, Downy_mildew_on_lettuce, Healthy,
      Powdery_mildew_on_lettuce, Septoria_blight_on_lettuce,
      Shepherd_purse_weed, Viral, Wilt_and_leaf_blight_or_rot
    Images per class vary — uses whatever is available.
    """
    name = "Lettuce"
    desc = ("Lettuce leaf disease dataset with 8 classes covering bacterial, "
            "fungal and viral conditions: Bacterial, Downy Mildew, Powdery Mildew, "
            "Septoria Blight, Viral, Wilt/Leaf Blight, Shepherd's Purse Weed, "
            "and Healthy. Image counts vary across classes.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    base = None

    # ── Download via kagglehub (preferred) ───────────────────────────────────
    if KAGGLEHUB_AVAILABLE:
        try:
            print(f"  [kagglehub] Downloading lettuce-diseases ...")
            dl_path = kagglehub.dataset_download("ashishjstar/lettuce-diseases")
            print(f"  Path to dataset files: {dl_path}")
            # Root folder is Lettuce_disease_datasets/ or the dl_path itself
            for candidate in [
                os.path.join(dl_path, "Lettuce_disease_datasets"),
                dl_path,
            ]:
                if os.path.isdir(candidate):
                    subs = [d for d in os.listdir(candidate)
                            if os.path.isdir(os.path.join(candidate, d))
                            and not d.startswith('.')]
                    if subs:
                        base = candidate
                        break
        except Exception as e:
            print(f"  [kagglehub ERROR] {e}")

    # ── Fallback: kaggle API ──────────────────────────────────────────────────
    if base is None:
        print(f"  [FALLBACK] Trying kaggle API ...")
        dl = os.path.join(DATA_ROOT, "lettuce-diseases")
        if not os.path.exists(dl):
            kaggle_download("ashishjstar/lettuce-diseases", dl)
        for candidate in [
            os.path.join(dl, "Lettuce_disease_datasets"),
            dl,
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    base = candidate
                    break

    if base is None or not os.path.isdir(base):
        print(f"  [ERROR] Could not locate Lettuce dataset folder.")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Auto-discover classes from subfolders — handles any truncation/naming
    classes = sorted([
        d for d in os.listdir(base)
        if os.path.isdir(os.path.join(base, d)) and not d.startswith('.')
    ])
    print(f"  [Lettuce] Found {len(classes)} classes: {classes}")

    data = collect_images_df(base)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {base}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # kagglehub cache — don't delete
    saved_df = sample_per_class(data, classes, n)
    saved_df = save_split(saved_df, name, "images")
    print(f"  [SAVED] {len(saved_df)} images -> Curated_Dataset/Images/...")
    return name, saved_df, saved_df, classes, desc


def load_Cucumber(n):
    """
    Cucumber Plant Diseases Dataset
    kagglehub: kareem3egm/cucumber-plant-diseases-dataset

    Structure:
      Cucumber plant diseases/
        training/  Ill_cucumber/  good_Cucumber/   <- REFERENCE pool
        testing/   Ill_cucumber/  good_Cucumber/   <- BENCHMARK pool
        single_prediction/                         <- ignored
    """
    name    = "Cucumber"
    classes = ["Ill_cucumber", "good_Cucumber"]
    desc    = ("Cucumber plant disease dataset (kaggle: kareem3egm). "
               "Two classes: Ill_cucumber (diseased) and good_Cucumber (healthy). "
               "Training split used for reference images, testing split for benchmark.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    dl = os.path.join(DATA_ROOT, "cucumber-plant-diseases")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading Cucumber dataset ...")
                dl_path = kagglehub.dataset_download(
                    "kareem3egm/cucumber-plant-diseases-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("kareem3egm/cucumber-plant-diseases-dataset", dl)
        else:
            kaggle_download("kareem3egm/cucumber-plant-diseases-dataset", dl)

    def _find_split(split_name):
        for candidate in [
            os.path.join(dl, "Cucumber plant diseases", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("training")
    test_dir  = _find_split("testing")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Cucumber training/testing folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [Cucumber] training: {train_dir}")
    print(f"  [Cucumber] testing:  {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1] == _c]) if len(test_data) else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: {_rc+_bc} imgs < {n} required -- excluded')
    actual_classes = _viable
    if not actual_classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

        # Combine train+test pools, sample n total per class
    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    saved_df = save_split(sample_per_class(combined, actual_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, actual_classes, desc


def load_DurianLeaf(n):
    """
    Durian Leaf Disease Dataset
    kagglehub: cthng123/durian-leaf-disease-dataset

    Structure:
      DLD_FinalDataset_224_sp.../
        train/  ALGAL_LEAF_SPOT/ ALLOCARIDARA_ATT/ HEALTHY_LEAF/
                LEAF_BLIGHT/ PHOMOPSIS_LEAF_SPOT/  <- REFERENCE pool
        test/   (same classes)                     <- BENCHMARK pool
        val/    (same classes)                     <- ignored

    Crop: Durian. Class names are UPPER_SNAKE_CASE.
    """
    name    = "Durian Leaf"
    desc    = ("Durian Leaf Disease Dataset (kaggle: cthng123). "
               "5 classes: Algal Leaf Spot, Allocaridara Attack, Healthy Leaf, "
               "Leaf Blight, Phomopsis Leaf Spot. Train/test/val splits.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        classes_from_disk = sorted(saved[1].unique().tolist()) if len(saved) > 0 else []
        return name, saved, saved, classes_from_disk, desc

    dl = os.path.join(DATA_ROOT, "durian-leaf-disease")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading durian-leaf-disease-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "cthng123/durian-leaf-disease-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("cthng123/durian-leaf-disease-dataset", dl)
        else:
            kaggle_download("cthng123/durian-leaf-disease-dataset", dl)

    def _find_split(split_name):
        """Find train/test/val under the DLD_FinalDataset_... root."""
        # Try direct
        for candidate in [os.path.join(dl, split_name)]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        # One level deep (DLD_FinalDataset_224_sp.../train/)
        for top in os.listdir(dl):
            top_path = os.path.join(dl, top)
            if not os.path.isdir(top_path) or top.startswith('.'):
                continue
            candidate = os.path.join(top_path, split_name)
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        # Walk fallback
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Durian train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [DurianLeaf] train: {train_dir}")
    print(f"  [DurianLeaf] test:  {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    ))

    if not actual_classes:
        print(f"  [ERROR] No classes found in DurianLeaf")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [DurianLeaf] {len(actual_classes)} classes: {actual_classes}")

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1] == _c]) if len(test_data) else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: {_rc+_bc} imgs < {n} required -- excluded')
    actual_classes = _viable
    if not actual_classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

        # Combine train+test pools, sample n total per class
    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    saved_df = save_split(sample_per_class(combined, actual_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")

    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")

    return name, saved_df, saved_df, actual_classes, desc


def load_EggplantDisease(n):
    """
    Eggplant Disease Recognition Dataset
    kagglehub: kamalmoha/eggplant-disease-recognition-dataset

    Structure:
      Eggplant Disease Recogn.../
        Original Images/
          Healthy Leaf/
          Insect Pest Disease/
          Leaf Spot Disease/
          Mosaic Virus Disease/
          Small Leaf Disease/
          White Mold Disease/
          Wilt Disease/
        Augmented Images/   <- ignored

    Crop: Eggplant. Single pool split into ref/bench.
    """
    name = "Eggplant Disease"
    desc = ("Eggplant Disease Recognition Dataset (kaggle: kamalmoha). "
            "7 classes: Healthy Leaf, Insect Pest Disease, Leaf Spot Disease, "
            "Mosaic Virus Disease, Small Leaf Disease, White Mold Disease, Wilt Disease. "
            "Original Images only (Augmented Images folder ignored).")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        classes_from_disk = sorted(saved[1].unique().tolist()) if len(saved) > 0 else []
        return name, saved, saved, classes_from_disk, desc

    dl = os.path.join(DATA_ROOT, "eggplant-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading eggplant-disease-recognition-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "kamalmoha/eggplant-disease-recognition-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("kamalmoha/eggplant-disease-recognition-dataset", dl)
        else:
            kaggle_download("kamalmoha/eggplant-disease-recognition-dataset", dl)

    # ── Locate "Original Images" folder ───────────────────────────────────────
    orig_dir = None
    for candidate in [os.path.join(dl, "Original Images")]:
        if os.path.isdir(candidate):
            orig_dir = candidate
            break
    if orig_dir is None:
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if d != "Augmented Images" and not d.startswith(".")]
            if os.path.basename(dirpath) == "Original Images":
                orig_dir = dirpath
                break
        # Fallback: find any parent that has class-like subfolders (not Augmented)
        if orig_dir is None:
            for dirpath, dirnames, _ in os.walk(dl):
                dirnames[:] = [d for d in dirnames if not d.startswith(".")]
                subs = [d for d in dirnames if d not in ("Augmented Images",)]
                if subs and any("Disease" in d or "Leaf" in d for d in subs):
                    orig_dir = dirpath
                    break

    if orig_dir is None:
        print(f"  [ERROR] Could not locate Original Images folder in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [EggplantDisease] Original Images: {orig_dir}")

    data = collect_images_df(orig_dir)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {orig_dir}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    actual_classes = sorted(data[1].unique().tolist())
    print(f"  [EggplantDisease] {len(actual_classes)} classes: {actual_classes}")

    saved, _ = split_save_cleanup(data, actual_classes, name, dl, n)
    return name, saved, saved, actual_classes, desc


def load_CottonDisease(n):
    """
    Cotton Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/cotton-disease-multi-transformation-dataset

    Structure:
      dataset/dataset/
        train/  Aphids/ Army_Worm/ Bacterial_Blight/ Healthy/
                Powdery_Mildew/ Target_Spot/          <- REFERENCE pool
        test/   (same classes)                        <- BENCHMARK pool
      augmentation/                                   <- ignored entirely

    Crop: Cotton. Train -> reference, test -> benchmark.
    """
    name    = "Cotton Disease"
    classes = ["Aphids", "Army_Worm", "Bacterial_Blight",
               "Healthy", "Powdery_Mildew", "Target_Spot"]
    desc    = ("Cotton Disease Multi Transformation Dataset (kaggle: shuvokumarbasak2030). "
               "6 classes: Aphids, Army Worm, Bacterial Blight, Healthy, "
               "Powdery Mildew, Target Spot. Train/test splits; augmentation folder ignored.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    dl = os.path.join(DATA_ROOT, "cotton-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading cotton-disease-multi-transformation-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/cotton-disease-multi-transformation-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download(
                    "shuvokumarbasak2030/cotton-disease-multi-transformation-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/cotton-disease-multi-transformation-dataset", dl)

    def _find_split(split_name):
        """Locate train/ or test/ under dataset/dataset/, ignoring augmentation/."""
        for candidate in [
            os.path.join(dl, "dataset", "dataset", split_name),
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            # Never descend into augmentation folder
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".") and d.lower() not in ("augmentation", "val")]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Cotton train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [CottonDisease] train: {train_dir}")
    print(f"  [CottonDisease] test:  {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    # Drop classes that cannot meet quota in both pools
    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: train={_rc}/{n}, test={_bc}/{n} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

        # Combine train+test pools, sample n total per class
    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    saved_df = save_split(sample_per_class(combined, actual_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, actual_classes, desc


def load_PumpkinLeaf(n):
    """
    Pumpkin Leaf Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset

    Structure (dataset only; augmentation/ ignored):
      dataset/dataset/
        train/  Bacterial_Leaf_Spot/ Downy_Mildew/ Healthy_Leaf/
                Mosaic_Disease/ Powdery_Mildew/
        test/   (same classes)
        val/    (same classes)  <- ignored
      augmentation/  <- never used

    Crop: Pumpkin.
    """
    name    = "Pumpkin Leaf"
    classes = ["Bacterial_Leaf_Spot", "Downy_Mildew", "Healthy_Leaf",
               "Mosaic_Disease", "Powdery_Mildew"]
    desc    = ("Pumpkin Leaf Disease Multi Transformation Dataset (kaggle: shuvokumarbasak2030). "
               "5 classes: Bacterial Leaf Spot, Downy Mildew, Healthy Leaf, "
               "Mosaic Disease, Powdery Mildew. Uses dataset/ only (augmentation not used).")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    dl = os.path.join(DOWNLOAD_ROOT, "pumpkin-leaf-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading pumpkin-leaf-disease-multi-transformation-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                err_msg = str(e)
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                if "122" in err_msg or "quota" in err_msg.lower() or "Disk" in err_msg:
                    print(f"  [HINT] Disk quota exceeded. Free space under {_BASE_DIR}.")
                kaggle_download(
                    "shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/pumpkin-leaf-disease-multi-transformation-dataset", dl)

    def _find_split(split_name):
        candidates = [
            os.path.join(dl, "dataset", "dataset", split_name),
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]
        # One and two levels down (kagglehub may extract with wrapper folder(s))
        try:
            for top in os.listdir(dl):
                if top.startswith("."): continue
                sub = os.path.join(dl, top)
                if os.path.isdir(sub):
                    candidates.extend([
                        os.path.join(sub, "dataset", "dataset", split_name),
                        os.path.join(sub, "dataset", split_name),
                        os.path.join(sub, split_name),
                    ])
                    for mid in os.listdir(sub):
                        if mid.startswith("."): continue
                        sub2 = os.path.join(sub, mid)
                        if os.path.isdir(sub2):
                            candidates.extend([
                                os.path.join(sub2, "dataset", "dataset", split_name),
                                os.path.join(sub2, "dataset", split_name),
                                os.path.join(sub2, split_name),
                            ])
        except OSError:
            pass
        for candidate in candidates:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        if os.path.isdir(dl):
            for dirpath, dirnames, _ in os.walk(dl):
                dirnames[:] = [d for d in dirnames
                               if not d.startswith(".") and d.lower() not in ("augmentation", "val")]
                if os.path.basename(dirpath).lower() == split_name.lower():
                    subs = [d for d in os.listdir(dirpath)
                            if os.path.isdir(os.path.join(dirpath, d))]
                    if subs:
                        return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    pool_dirs = None
    if train_dir is None and test_dir is None:
        # Fallback: use class-based dirs under dataset/ only (ignore augmentation)
        pool_dirs = _find_class_based_dirs(dl, ignore_dirs=("augmentation", "val"))
        if pool_dirs:
            combined_fallback = pd.concat([collect_images_df(d) for d in pool_dirs], ignore_index=True)
            if len(combined_fallback) > 0:
                print(f"  [PumpkinLeaf] data (all): {len(pool_dirs)} folder(s) combined")
                train_data = combined_fallback
                test_data = pd.DataFrame(columns=[0, 1])
                train_dir = pool_dirs[0]  # for display
            else:
                pool_dirs = None
        if train_dir is None and test_dir is None:
            try:
                contents = os.listdir(dl) if os.path.isdir(dl) else []
                print(f"  [ERROR] Could not locate Pumpkin train/test folders in {dl}")
                print(f"  [DEBUG] Folder contents: {contents[:20]}{'...' if len(contents) > 20 else ''}")
            except OSError:
                print(f"  [ERROR] Could not locate Pumpkin train/test folders in {dl}")
                print(f"  [DEBUG] Folder missing or not readable: {dl}")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc
    if pool_dirs is None:
        print(f"  [PumpkinLeaf] train: {train_dir}")
        print(f"  [PumpkinLeaf] test:  {test_dir}")

    if pool_dirs is None:
        train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
        test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])
    # else train_data already set from fallback; test_data empty

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: train={_rc}/{n}, test={_bc}/{n} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Combine train+test pools, sample n total per class
    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    saved_df = save_split(sample_per_class(combined, actual_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, actual_classes, desc


def load_RoseLeaf(n):
    """
    Rose Leaf Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset

    Structure:
      dataset/dataset/
        train/  Black_Spot/ Downy_Mildew/ Dry_Leaf/ Healthy_Leaf/ Leaf_Holes/
        test/   (same classes)   <- BENCHMARK pool
        val/                     <- ignored
      augmentation/              <- ignored
      raw_mixed/                 <- ignored

    Crop: Rose. Train -> reference, test -> benchmark.
    """
    name    = "Rose Leaf"
    classes = ["Black_Spot", "Downy_Mildew", "Dry_Leaf", "Healthy_Leaf", "Leaf_Holes"]
    desc    = ("Rose Leaf Disease Multi Transformation Dataset (kaggle: shuvokumarbasak2030). "
               "5 classes: Black Spot, Downy Mildew, Dry Leaf, Healthy Leaf, Leaf Holes. "
               "Train/test splits used; augmentation, val, raw_mixed ignored.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    dl = os.path.join(DOWNLOAD_ROOT, "rose-leaf-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading rose-leaf-disease-multi-transformation-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                err_msg = str(e)
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                if "122" in err_msg or "quota" in err_msg.lower() or "Disk" in err_msg:
                    print(f"  [HINT] Disk quota exceeded. Free space under {_BASE_DIR}.")
                kaggle_download(
                    "shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/rose-leaf-disease-multi-transformation-dataset", dl)

    _IGNORE = {"augmentation", "val", "raw_mixed"}

    def _find_split(split_name):
        candidates = [
            os.path.join(dl, "dataset", "dataset", split_name),
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]
        # If kagglehub extracts with an extra top-level folder (e.g. "Version 2"), look one level down
        try:
            for top in os.listdir(dl):
                if top.startswith("."): continue
                sub = os.path.join(dl, top)
                if os.path.isdir(sub):
                    candidates.extend([
                        os.path.join(sub, "dataset", "dataset", split_name),
                        os.path.join(sub, "dataset", split_name),
                        os.path.join(sub, split_name),
                    ])
        except OSError:
            pass
        for candidate in candidates:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".") and d.lower() not in _IGNORE]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    pool_dirs = None
    if train_dir is None and test_dir is None:
        pool_dirs = _find_class_based_dirs(dl, ignore_dirs=_IGNORE)
        if pool_dirs:
            combined_fallback = pd.concat([collect_images_df(d) for d in pool_dirs], ignore_index=True)
            if len(combined_fallback) > 0:
                print(f"  [RoseLeaf] data (all): {len(pool_dirs)} folder(s) combined")
                train_data = combined_fallback
                test_data = pd.DataFrame(columns=[0, 1])
                train_dir = pool_dirs[0]
            else:
                pool_dirs = None
        if train_dir is None and test_dir is None:
            print(f"  [ERROR] Could not locate Rose train/test folders in {dl}")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc
    if pool_dirs is None:
        print(f"  [RoseLeaf] train: {train_dir}")
        print(f"  [RoseLeaf] test:  {test_dir}")

    if pool_dirs is None:
        train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
        test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: train={_rc}/{n}, test={_bc}/{n} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    saved_df = save_split(sample_per_class(combined, actual_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, actual_classes, desc


def load_CoconutDisease(n):
    """
    Coconut Disease Multi Transformation STTV Dataset
    kaggle: shuvokumarbasak2030/coconut-disease-multi-transformation-sttv-dataset

    Structure:
      dataset/
        test/  CCI_Caterpillars/ CCI_Leaflets/ Healthy_Leaves/
               WCLWD_DryingofLeaves/ WCLWD_Flaccidity/ WCLWD_Yellowing/  <- BENCHMARK
        train/ (same classes)                                             <- REFERENCE
        val/                                                              <- ignored

    Crop: Coconut. Train -> reference, test -> benchmark.
    """
    name    = "Coconut Disease"
    classes = ["CCI_Caterpillars", "CCI_Leaflets", "Healthy_Leaves",
               "WCLWD_DryingofLeaves", "WCLWD_Flaccidity", "WCLWD_Yellowing"]
    desc    = ("Coconut Disease Multi Transformation STTV Dataset (kaggle: shuvokumarbasak2030). "
               "6 classes: CCI Caterpillars, CCI Leaflets, Healthy Leaves, "
               "WCLWD Drying of Leaves, WCLWD Flaccidity, WCLWD Yellowing. "
               "Train/test splits used; val ignored.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    dl = os.path.join(DATA_ROOT, "coconut-disease-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading coconut-disease-multi-transformation-sttv-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "shuvokumarbasak2030/coconut-disease-multi-transformation-sttv-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download(
                    "shuvokumarbasak2030/coconut-disease-multi-transformation-sttv-dataset", dl)
        else:
            kaggle_download(
                "shuvokumarbasak2030/coconut-disease-multi-transformation-sttv-dataset", dl)

    _IGNORE = {"val", "augmentation"}

    def _find_split(split_name):
        for candidate in [
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".") and d.lower() not in _IGNORE]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Coconut train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [CoconutDisease] train: {train_dir}")
    print(f"  [CoconutDisease] test:  {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    )) or classes

    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1]   == _c]) if len(test_data)  else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: train={_rc}/{n}, test={_bc}/{n} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

        # Combine train+test pools, sample n total per class
    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    saved_df = save_split(sample_per_class(combined, actual_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, actual_classes, desc


# ═══════════════════════════════════════════════════════════════════════════
#  STRAWBERRY DISEASE DETECTION (KAGGLE)
#  usmanafzaal/strawberry-disease-detection-dataset
#  Layout (observed locally after download/unzip):
#      train/  *.jpg + *.json  (filenames encode disease: e.g. angular_leafspot359.jpg)
#      val/    *.jpg + *.json
#      test/   *.jpg + *.json
#      Test Disease Severity Level/Level 1 + Level 2  (extra severity-labelled samples)
# ═══════════════════════════════════════════════════════════════════════════

def load_StrawberryDiseaseDetection(n):
    name = "Strawberry Disease Detection"
    classes = []  # discovered from filenames
    desc = ("Strawberry disease detection dataset (Kaggle: usmanafzaal/str"
            "awberry-disease-detection-dataset). Images of strawberry leaves "
            "with JSON annotations and explicit train/val/test splits.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        return name, saved, saved, sorted(saved[1].unique().tolist()), desc

    dl = os.path.join(DATA_ROOT, "strawberry-disease-detection-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading strawberry-disease-detection-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "usmanafzaal/strawberry-disease-detection-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("usmanafzaal/strawberry-disease-detection-dataset", dl)
        else:
            kaggle_download("usmanafzaal/strawberry-disease-detection-dataset", dl)

    def _find_split_dir(split_name):
        # Prefer top-level <dl>/<split_name>
        cand = os.path.join(dl, split_name)
        if os.path.isdir(cand):
            return cand
        # Fallback: search anywhere under dl for a folder with this name and images in it
        for dirpath, dirnames, filenames in os.walk(dl):
            if os.path.basename(dirpath).lower() == split_name.lower():
                if any(f.lower().endswith(IMAGE_EXT) for f in filenames):
                    return dirpath
        return None

    def _load_flat_dir(path):
        rows = []
        if not path or not os.path.isdir(path):
            return pd.DataFrame(columns=[0, 1])
        for f in os.listdir(path):
            if not f.lower().endswith(IMAGE_EXT):
                continue
            stem = os.path.splitext(f)[0]
            # Class name = leading alphabetic/underscore prefix before digits, e.g.
            #   angular_leafspot359 -> angular_leafspot
            m = re.match(r"[A-Za-z_]+", stem)
            label = m.group(0) if m else stem
            cls_key = safe_name(label)
            rows.append({0: os.path.join(path, f), 1: cls_key})
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[0, 1])

    train_dir = _find_split_dir("train")
    val_dir   = _find_split_dir("val")
    test_dir  = _find_split_dir("test")

    # Optional: incorporate severity-level images as extra benchmark data
    sev_root = _find_split_dir("Test Disease Severity Level")
    sev_df = pd.DataFrame(columns=[0, 1])
    if sev_root:
        for level in ("Level 1", "Level 2"):
            lp = os.path.join(sev_root, level)
            sev_df = pd.concat([sev_df, _load_flat_dir(lp)], ignore_index=True)

    if not any([train_dir, val_dir, test_dir, sev_root]):
        print(f"  [ERROR] Could not locate Strawberry train/val/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [Strawberry] train: {train_dir}")
    print(f"  [Strawberry] val:   {val_dir}")
    print(f"  [Strawberry] test:  {test_dir}")
    if sev_root:
        print(f"  [Strawberry] severity root: {sev_root}")

    train_df = _load_flat_dir(train_dir)
    val_df   = _load_flat_dir(val_dir)
    test_df  = _load_flat_dir(test_dir)

    ref_data   = train_df
    bench_data = pd.concat([val_df, test_df, sev_df], ignore_index=True) if len(val_df) or len(test_df) or len(sev_df) else pd.DataFrame(columns=[0,1])

    all_classes = sorted(set(
        (ref_data[1].unique().tolist()   if len(ref_data)   else []) +
        (bench_data[1].unique().tolist() if len(bench_data) else [])
    ))

    # Filter classes by quota (using combined ref+bench counts)
    _viable = []
    for _c in all_classes:
        _rc = len(ref_data[ref_data[1] == _c]) if len(ref_data) else 0
        _bc = len(bench_data[bench_data[1] == _c]) if len(bench_data) else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and (_rc + _bc) >= n):
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: total={_rc+_bc} < {n} required -- excluded")
    all_classes = _viable

    if not all_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    combined = pd.concat([ref_data, bench_data], ignore_index=True) if len(ref_data) and len(bench_data) else (ref_data if len(ref_data) else bench_data)
    saved_df = save_split(sample_per_class(combined, all_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, all_classes, desc


def load_VanillaDisease(n):
    """
    Vanilla Disease Multi Transformation Dataset
    kaggle: mihsanpermana/vanilla-plant-disease-image-dataset


    No train/test split — single pool (all variants) split into ref/bench here.
    Crop: Vanilla.
    """
    name = "Vanilla Disease"
    desc = ("Vanilla Disease Multi Transformation Dataset (kaggle: shuvokumarbasak2030). "
            "Classes include black spots, healthy, rotten stem etc. "
            "All image variants across all subfolders used; single pool split into ref/bench.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        classes_from_disk = sorted(saved[1].unique().tolist()) if len(saved) > 0 else []
        return name, saved, saved, classes_from_disk, desc

    dl = os.path.join(DATA_ROOT, "vanilla-plant-disease-image-dataset")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading vanilla-plant-disease-image-dataset ...")
                dl_path = kagglehub.dataset_download(
                    "mihsanpermana/vanilla-plant-disease-image-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download(
                    "mihsanpermana/vanilla-plant-disease-image-dataset", dl)
        else:
            kaggle_download(
                "mihsanpermana/vanilla-plant-disease-image-dataset", dl)

    # ── Locate dataset root — folder whose subdirs are class names ────────────
    # Structure: dl/<root>/<class>/original/*.jpg
    # Find a folder containing subdirs that each have an "original" subfolder
    dataset_root = None
    for dirpath, dirnames, _ in os.walk(dl):
        dirnames[:] = [d for d in dirnames if not d.startswith(".")]
        has_original = [d for d in dirnames
                        if os.path.isdir(os.path.join(dirpath, d, "original"))]
        if has_original:
            dataset_root = dirpath
            break
    # Fallback: any folder containing class-like subdirs with image subfolders
    if dataset_root is None:
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith(".")]
            if len(dirnames) >= 2:
                dataset_root = dirpath
                break

    if dataset_root is None:
        print(f"  [ERROR] Could not locate Vanilla dataset root in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [VanillaDisease] Dataset root: {dataset_root}")

    # ── Collect images from ALL subfolders of each class ────────────────────
    rows = []
    classes_found = []
    for cls_dir in sorted(os.listdir(dataset_root)):
        cls_path = os.path.join(dataset_root, cls_dir)
        if not os.path.isdir(cls_path) or cls_dir.startswith("."):
            continue

        # Collect recursively from all subfolders (all augmentation variants included)
        cls_key = safe_name(cls_dir)
        cls_imgs = collect_images_recursive(cls_path)
        if not cls_imgs:
            print(f"  [SKIP CLASS] '{cls_dir}': no images found")
            continue

        classes_found.append(cls_key)
        for img_path in cls_imgs:
            rows.append({0: img_path, 1: cls_key})


    if not rows:
        print(f"  [ERROR] No images found for VanillaDisease")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    data = pd.DataFrame(rows)
    print(f"  [VanillaDisease] {len(classes_found)} classes, {len(data)} images total")

    saved, _ = split_save_cleanup(data, classes_found, name, dl, n)
    return name, saved, saved, classes_found, desc


def load_SugarLeafIDN(n):
    """
    SugarLeaf-IDN Dataset (sugarcane leaf diseases, Indonesia)
    kagglehub: bettydpuspasari/sugarleafidn

    Structure:
      dataset_Resize_224x224_NB_test100/
        train/  0 Healthy/ 1 Pokkahboeng/ 2 Common Rust (Karat Daun)/
                3 Eye Spot (cincin)/ 4 yellow Spot (noda Kuning)/
                5 red spot (noda Merah)/ 6 Mosaic/
                7 Streak Mosaic SCSMV (bergaris)/ 8 Leaf Scald (blendok)/
        test/   (same 9 classes, 100 images each)  <- BENCHMARK
        validation/                                <- ignored

    Numeric prefix stripped from class names (e.g. "1 Pokkahboeng" -> "Pokkahboeng").
    Crop: Sugarcane. Train -> reference, test -> benchmark.
    """
    name = "SugarLeaf IDN"
    desc = ("SugarLeaf-IDN dataset (kaggle: bettydpuspasari). 9 classes of sugarcane "
            "leaf diseases from East Java, Indonesia. 224x224 px, field conditions. "
            "Classes: Healthy, Pokkahboeng, Common Rust, Eye Spot, Yellow Spot, "
            "Red Spot, Mosaic, Streak Mosaic SCSMV, Leaf Scald.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        classes_from_disk = sorted(saved[1].unique().tolist()) if len(saved) > 0 else []
        return name, saved, saved, classes_from_disk, desc

    dl = os.path.join(DATA_ROOT, "sugarleafidn")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading sugarleafidn ...")
                dl_path = kagglehub.dataset_download("bettydpuspasari/sugarleafidn")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("bettydpuspasari/sugarleafidn", dl)
        else:
            kaggle_download("bettydpuspasari/sugarleafidn", dl)

    _IGNORE = {"validation", "val", "augmentation"}

    def _find_split(split_name):
        # Try one and two levels deep first
        for candidate in [
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        # One level deep (dataset_Resize_.../train/)
        for top in os.listdir(dl):
            top_path = os.path.join(dl, top)
            if not os.path.isdir(top_path) or top.startswith("."): continue
            candidate = os.path.join(top_path, split_name)
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        # Walk fallback
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames
                           if not d.startswith(".") and d.lower() not in _IGNORE]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate SugarLeafIDN train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [SugarLeafIDN] train: {train_dir}")
    print(f"  [SugarLeafIDN] test:  {test_dir}")

    def _collect_strip_prefix(split_dir):
        """Collect images; strip leading numeric prefix from class folder names."""
        if not split_dir or not os.path.isdir(split_dir):
            return pd.DataFrame(columns=[0, 1])
        rows = []
        for cls_dir in os.listdir(split_dir):
            cls_path = os.path.join(split_dir, cls_dir)
            if not os.path.isdir(cls_path) or cls_dir.startswith("."):
                continue
            # Strip leading "N " prefix: "1 Pokkahboeng" -> "Pokkahboeng"
            import re as _re
            clean = _re.sub(r"^\d+\s+", "", cls_dir).strip()
            cls_key = safe_name(clean) if clean else safe_name(cls_dir)
            for f in os.listdir(cls_path):
                if f.lower().endswith(IMAGE_EXT):
                    rows.append({0: os.path.join(cls_path, f), 1: cls_key})
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[0, 1])

    train_data = _collect_strip_prefix(train_dir)
    test_data  = _collect_strip_prefix(test_dir)

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    ))

    if not actual_classes:
        print(f"  [ERROR] No classes found for SugarLeafIDN")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [SugarLeafIDN] {len(actual_classes)} classes: {actual_classes}")

    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1] == _c]) if len(test_data) else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f"  [DROP CLASS] {_c!r}: train={_rc}/{n}, test={_bc}/{n} -- excluded")
    actual_classes = _viable
    if not actual_classes:
        print(f"  [DROP ALL] No viable classes for {name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

        # Combine train+test pools, sample n total per class
    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    saved_df = save_split(sample_per_class(combined, actual_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, actual_classes, desc


def load_CucumberZenodo(n):
    """
    Cucumber Disease and Freshness Classification Dataset – Curated Annotations
    Zenodo DOI: 10.5281/zenodo.16816441
    URL: https://zenodo.org/records/16816441

    Structure (flat class folders, no train/test split):
      Anthracnose/
      Bacterial_Wilt/
      Belly_Rot/
      Downy_Mildew/
      Fresh_Cucumber/
      Fresh_Leaf/
      Pythium_Fruit_Rot/

    Downloaded as a zip via Zenodo REST API.
    Crop: Cucumber. Single pool split into ref/bench.
    """
    name    = "Cucumber Zenodo"
    classes = ["Anthracnose", "Bacterial_Wilt", "Belly_Rot", "Downy_Mildew",
               "Fresh_Cucumber", "Fresh_Leaf", "Pythium_Fruit_Rot"]
    desc    = ("Cucumber Disease and Freshness Classification Dataset – Curated Annotations "
               "(Zenodo DOI: 10.5281/zenodo.16816441). 7 classes: Anthracnose, Bacterial Wilt, "
               "Belly Rot, Downy Mildew, Fresh Cucumber, Fresh Leaf, Pythium Fruit Rot. "
               "Manually annotated with Label Studio.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    dl = os.path.join(DATA_ROOT, "cucumber-zenodo-dataset")
    zip_path = dl + ".zip"

    if not os.path.exists(dl):
        os.makedirs(dl, exist_ok=True)
        # ── Download via Zenodo REST API ──────────────────────────────────────
        RECORD_ID = "16816441"
        api_url   = f"https://zenodo.org/api/records/{RECORD_ID}"
        try:
            import urllib.request
            from urllib.parse import urlparse, urlunparse, quote
            import json as _json
            print(f"  [Zenodo] Fetching record metadata from {api_url} ...")
            with urllib.request.urlopen(api_url, timeout=30) as resp:
                meta = _json.loads(resp.read().decode())
            files = meta.get("files", [])
            zip_files = [f for f in files if f.get("key","").endswith(".zip")]
            dl_files  = zip_files if zip_files else files
            if not dl_files:
                print(f"  [ERROR] No files found in Zenodo record {RECORD_ID}")
                return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc
            for file_info in dl_files:
                file_url  = file_info.get("links", {}).get("self") or file_info.get("download_url")
                file_name = file_info.get("key", "dataset.zip")
                dest      = os.path.join(dl, file_name)
                # Encode URL path so filenames with spaces (e.g. "Image tags.zip") work
                if file_url:
                    parsed = urlparse(file_url)
                    encoded_path = quote(parsed.path, safe="/")
                    file_url = urlunparse(parsed._replace(path=encoded_path))
                print(f"  [Zenodo] Downloading {file_name} ...")
                urllib.request.urlretrieve(file_url, dest)
                if dest.endswith(".zip"):
                    import zipfile
                    print(f"  [Zenodo] Extracting {file_name} ...")
                    with zipfile.ZipFile(dest, "r") as z:
                        z.extractall(dl)
                    os.remove(dest)
        except Exception as e:
            print(f"  [Zenodo ERROR] {e}")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # ── Locate dataset root — folder whose subdirs match class names ──────────
    dataset_root = None
    for dirpath, dirnames, _ in os.walk(dl):
        dirnames[:] = [d for d in dirnames if not d.startswith(".") and d != "__MACOSX"]
        hits = [d for d in dirnames
                if any(kw in d.lower() for kw in
                       ("anthracnose","bacterial","belly","downy","fresh","pythium","cucumber"))]
        if len(hits) >= 3:
            dataset_root = dirpath
            break

    if dataset_root is None:
        dataset_root = dl  # best-effort fallback
    print(f"  [CucumberZenodo] Dataset root: {dataset_root}")

    data = collect_images_df(dataset_root)
    if len(data) == 0:
        print(f"  [ERROR] No images found in {dataset_root}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    actual_classes = sorted(data[1].unique().tolist())
    print(f"  [CucumberZenodo] {len(actual_classes)} classes: {actual_classes}")

    saved, _ = split_save_cleanup(data, actual_classes, name, dl, n)
    return name, saved, saved, actual_classes, desc


def load_Cauliflower(n):
    """
    Cauliflower Disease Multi Transformation Dataset
    kaggle: shuvokumarbasak2030/cauliflower-disease-multi-transformation-dataset
    Structure: dataset/dataset/train/{Bacterial_Spot_Rot, Black_Rot, Downy_Mildew, No_disease}
    4 classes; uses train split for sampling.
    """
    name    = "Cauliflower"
    classes = ['Bacterial_Spot_Rot', 'Black_Rot', 'Downy_Mildew', 'No_disease']
    desc    = ("Cauliflower leaf disease dataset with multi-transformation augmentation. "
               "Covers Bacterial Spot Rot, Black Rot, Downy Mildew and No Disease "
               "(healthy). Train/val/test splits available; train split used here.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    dl   = os.path.join(DATA_ROOT, "cauliflower-disease")

    if not os.path.exists(dl):
        kaggle_download(
            "shuvokumarbasak2030/cauliflower-disease-multi-transformation-dataset", dl)

    # Confirmed structure (Kaggle Data Explorer):
    #   dataset/dataset/train/  <-- REFERENCE images
    #   dataset/dataset/test/   <-- BENCHMARK images
    #   dataset/dataset/val/    <-- ignored
    #   augmentation/           <-- ignored entirely

    def _find_cauliflower_split(split_name):
        # Try explicit paths first
        for candidate in [
            os.path.join(dl, "dataset", "dataset", split_name),
            os.path.join(dl, "dataset", split_name),
            os.path.join(dl, split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith('.')]
                if subs:
                    return candidate
        # Walk fallback — only match folder named split_name inside a dataset parent
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            if (os.path.basename(dirpath) == split_name
                    and "dataset" in dirpath.lower()):
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))
                        and not d.startswith('.')]
                if subs:
                    return dirpath
        return None

    train_dir = _find_cauliflower_split("train")
    test_dir  = _find_cauliflower_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate Cauliflower train or test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    print(f"  [Cauliflower] train: {train_dir}")
    print(f"  [Cauliflower] test:  {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    if len(train_data) == 0 and len(test_data) == 0:
        print(f"  [ERROR] No images found in train or test folders")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), classes, desc

    # Combine train+test, sample n per class, save to one directory per class
    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    _viable = []
    for _c in classes:
        count = len(combined[combined[1] == _c]) if len(combined) else 0
        if (n is None and count >= 1) or (n is not None and count >= n):
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: {count} imgs < {n} required -- excluded')
    classes = _viable
    if not classes:
        print(f'  [DROP ALL] No classes meet quota in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    saved_df = sample_per_class(combined, classes, n)
    saved_df = save_split(saved_df, name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, classes, desc


def load_NewPlantDiseases(n):
    """
    New Plant Diseases Dataset (vipoooool/new-plant-diseases-dataset) — 1.43 GB
    kagglehub: vipoooool/new-plant-diseases-dataset

    Structure:
      New Plant Diseases Dataset/New Plant Diseases Da.../
        train/  <Crop>___<Disease>/   <- class subfolders, REFERENCE pool
        test/test/                    <- flat folder, filenames encode class, BENCHMARK pool
          AppleCedarRust1.JPG
          AppleScab1.JPG

    Class format: "Apple___Apple_scab" (triple underscore, same as PlantVillage tfds).
    Train split -> reference. Flat test folder -> benchmark via filename prefix matching.
    """
    name = "New Plant Diseases"
    desc = ("New Plant Diseases Dataset (Kaggle: vipoooool). 1.43 GB, 38 classes "
            "across 14 crop species using triple-underscore Crop___Disease naming. "
            "Train split used for reference; flat test folder for benchmark.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    dl = os.path.join(DATA_ROOT, "new-plant-diseases")
    if not os.path.exists(dl):
        if KAGGLEHUB_AVAILABLE:
            try:
                print(f"  [kagglehub] Downloading New Plant Diseases dataset ...")
                dl_path = kagglehub.dataset_download(
                    "vipoooool/new-plant-diseases-dataset")
                print(f"  Path to dataset files: {dl_path}")
                os.makedirs(dl, exist_ok=True)
                shutil.copytree(dl_path, dl, dirs_exist_ok=True)
            except Exception as e:
                print(f"  [kagglehub ERROR] {e} — trying kaggle API fallback")
                kaggle_download("vipoooool/new-plant-diseases-dataset", dl)
        else:
            kaggle_download("vipoooool/new-plant-diseases-dataset", dl)

    # ── Locate train folder (has Crop___Disease subfolders) ───────────────────
    train_dir = None
    for candidate in [
        os.path.join(dl, "New Plant Diseases Dataset", "New Plant Diseases Dataset", "train"),
        os.path.join(dl, "New Plant Diseases Dataset", "train"),
        os.path.join(dl, "train"),
    ]:
        if os.path.isdir(candidate):
            train_dir = candidate; break
    if train_dir is None:
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            if os.path.basename(dirpath) == "train":
                if any("___" in d for d in dirnames):
                    train_dir = dirpath; break

    if train_dir is None:
        print(f"  [ERROR] Could not locate train folder in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
    print(f"  [NewPlantDiseases] train: {train_dir}")

    # ── Locate flat test folder ────────────────────────────────────────────────
    test_dir = None
    for candidate in [
        os.path.join(dl, "New Plant Diseases Dataset", "New Plant Diseases Dataset", "test", "test"),
        os.path.join(dl, "New Plant Diseases Dataset", "test", "test"),
        os.path.join(dl, "test", "test"),
        os.path.join(dl, "test"),
    ]:
        if os.path.isdir(candidate):
            imgs = [f for f in os.listdir(candidate) if f.lower().endswith(IMAGE_EXT)]
            if imgs:
                test_dir = candidate; break
    if test_dir is None:
        for dirpath, dirnames, filenames in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]
            imgs = [f for f in filenames if f.lower().endswith(IMAGE_EXT)]
            if imgs and not dirnames:   # leaf folder with images only
                test_dir = dirpath; break
    print(f"  [NewPlantDiseases] test:  {test_dir}")

    # ── Reference DataFrame from train subfolders ─────────────────────────────
    train_data = collect_images_df(train_dir)
    def _norm(lbl):
        if "___" in lbl:
            parts = lbl.split("___", 1)
            return safe_name(f"{parts[0].strip()}_{parts[1].strip()}")
        return safe_name(lbl)
    if len(train_data):
        train_data[1] = train_data[1].apply(_norm)
    classes = sorted(train_data[1].unique().tolist()) if len(train_data) else []

    # ── Benchmark DataFrame from flat test folder (filename prefix matching) ───
    img_rows = []
    if test_dir and classes:
        cls_lookup = {c.lower().replace("_","").replace(" ",""): c for c in classes}
        for fname in os.listdir(test_dir):
            if not fname.lower().endswith(IMAGE_EXT):
                continue
            stem_norm = os.path.splitext(fname)[0].lower().rstrip("0123456789").replace("_","").replace(" ","")
            matched = next(
                (cls_key for cls_bare, cls_key in cls_lookup.items()
                 if stem_norm == cls_bare or stem_norm.startswith(cls_bare)),
                None
            )
            if matched:
                img_rows.append({0: os.path.join(test_dir, fname), 1: matched})
    bench_data = pd.DataFrame(img_rows) if img_rows else pd.DataFrame(columns=[0,1])
    print(f"  [NewPlantDiseases] {len(classes)} classes | "
          f"{len(train_data)} train imgs | {len(bench_data)} test imgs matched")

    if not classes:
        print(f"  [ERROR] No classes found")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(bench_data[bench_data[1] == _c]) if len(bench_data) else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: {_rc+_bc} imgs < {n} required -- excluded')
    classes = _viable
    if not classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    combined = pd.concat([train_data, bench_data], ignore_index=True) if len(train_data) and len(bench_data) else (train_data if len(train_data) else bench_data)
    saved_df = sample_per_class(combined, classes, n)
    saved_df = save_split(saved_df, name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, classes, desc


def load_PlantDoc(n):
    """
    PlantDoc Dataset — GitHub: pratikkayal/PlantDoc-Dataset
    https://github.com/pratikkayal/PlantDoc-Dataset.git

    Structure:
      PlantDoc-Dataset/
        train/  <Class Name with spaces>/  <- REFERENCE pool
        test/   <Class Name with spaces>/  <- BENCHMARK pool

    Class names e.g. "Apple Scab Leaf" — first word = crop, rest = disease.
    Train -> reference, test -> benchmark. Cloned via git.
    """
    name = "PlantDoc"
    desc = ("PlantDoc Dataset (GitHub: pratikkayal). 27 classes across 13 crop "
            "species. Real-world plant disease images with natural lighting. "
            "Train split for reference, test split for benchmark.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    dl = os.path.join(DATA_ROOT, "PlantDoc-Dataset")
    if not os.path.exists(dl):
        print(f"  [git] Cloning PlantDoc-Dataset ...")
        ret = os.system(
            f'git clone --depth=1 '
            f'https://github.com/pratikkayal/PlantDoc-Dataset.git "{dl}"')
        if ret != 0 or not os.path.isdir(dl):
            print(f"  [ERROR] git clone failed (exit {ret}). "
                  "Check internet access and that git is installed.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        print(f"  [git] Cloned to {dl}")

    def _find_split(split_name):
        for candidate in [
            os.path.join(dl, split_name),
            os.path.join(dl, "PlantDoc-Dataset", split_name),
        ]:
            if os.path.isdir(candidate):
                subs = [d for d in os.listdir(candidate)
                        if os.path.isdir(os.path.join(candidate, d))
                        and not d.startswith(".")]
                if subs:
                    return candidate
        for dirpath, dirnames, _ in os.walk(dl):
            dirnames[:] = [d for d in dirnames if not d.startswith(".")]
            if os.path.basename(dirpath).lower() == split_name.lower():
                subs = [d for d in os.listdir(dirpath)
                        if os.path.isdir(os.path.join(dirpath, d))]
                if subs:
                    return dirpath
        return None

    train_dir = _find_split("train")
    test_dir  = _find_split("test")

    if train_dir is None and test_dir is None:
        print(f"  [ERROR] Could not locate PlantDoc train/test folders in {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [PlantDoc] train: {train_dir}")
    print(f"  [PlantDoc] test:  {test_dir}")

    train_data = collect_images_df(train_dir) if train_dir else pd.DataFrame(columns=[0,1])
    test_data  = collect_images_df(test_dir)  if test_dir  else pd.DataFrame(columns=[0,1])

    actual_classes = sorted(set(
        (train_data[1].unique().tolist() if len(train_data) else []) +
        (test_data[1].unique().tolist()  if len(test_data)  else [])
    ))

    if not actual_classes:
        print(f"  [ERROR] No classes found in PlantDoc")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [PlantDoc] {len(actual_classes)} classes discovered")

    # Drop classes that cannot meet quota in EITHER pool
    _viable = []
    for _c in actual_classes:
        _rc = len(train_data[train_data[1] == _c]) if len(train_data) else 0
        _bc = len(test_data[test_data[1] == _c]) if len(test_data) else 0
        if (n is None and (_rc + _bc) >= 1) or (n is not None and _rc + _bc >= n):
            _viable.append(_c)
        else:
            print(f'  [DROP CLASS] {_c!r}: {_rc+_bc} imgs < {n} required -- excluded')
    actual_classes = _viable
    if not actual_classes:
        print(f'  [DROP ALL] No classes meet quota for both pools in {name}')
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

        # Combine train+test pools, sample n total per class
    combined = pd.concat([train_data, test_data], ignore_index=True) if len(train_data) and len(test_data) else (train_data if len(train_data) else test_data)
    saved_df = save_split(sample_per_class(combined, actual_classes, n), name, "images")
    print(f"  [SAVED] {name}: {len(saved_df)} images")
    if os.path.exists(dl):
        shutil.rmtree(dl, ignore_errors=True)
        print(f"  [CLEANUP] Deleted raw data: {dl}")
    return name, saved_df, saved_df, actual_classes, desc



def load_FUSARIUM22(n):
    name    = "FUSARIUM 22"
    rmap    = {'1(HR)':'Highly Resistant','9(HS)':'Highly Susceptible',
               '5(MR)':'Moderately Resistant','3(R)':'Resistant','7(S)':'Susceptible'}
    classes = list(rmap.values())
    desc    = ("Fusarium Wilt disease severity in chickpea. Classes range from Highly "
               "Resistant to Highly Susceptible — ideal for ordinal severity tasks.")
    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc
    dl   = os.path.join(DATA_ROOT, "fusarium22")
    base = os.path.join(dl, "FUSARIUM-22", "dataset_raw")
    if not os.path.exists(dl):
        kaggle_download("tolgahayit/fusarium-wilt-disease-in-chickpea-dataset", dl)
    rename_folders_dict(base, rmap)
    data = collect_images_df(base)
    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc

# ═══════════════════════════════════════════════════════════════════════════
#  LEAFNET LOADER  (HuggingFace streaming — enalis/LeafNet)
# ═══════════════════════════════════════════════════════════════════════════

def _extract_crop_disease(caption):
    """Parse crop + disease label from a LeafNet caption string."""
    c  = str(caption).strip()
    cl = c.lower()

    # "a image of <crop> healthy leaves"
    m = re.search(r'a image of ([a-z]+) healthy leaves?', cl)
    if m:
        return m.group(1).capitalize(), "Healthy"

    # "a image of <crop> leaves diseased by <Disease> [with ...]"
    m = re.search(r'a image of ([a-z]+) leaves? diseased by ([a-z][a-z\s]*?)(?:\s+with|\s+disease|$)', cl)
    if m:
        return m.group(1).capitalize(), m.group(2).strip().title()

    # "a image of <crop> leaves with <Disease> [with symptoms ...]"
    m = re.search(r'a image of ([a-z]+) leaves? with ([a-z][a-z\s]*?)(?:\s+with|\s+symptoms|$)', cl)
    if m:
        disease = m.group(2).strip().title()
        if len(disease.split()) <= 5:
            return m.group(1).capitalize(), disease

    # Fallback: "<crop> leaves diseased by <Disease>"
    m = re.search(r'([a-z]+) leaves? diseased by ([a-z][a-z\s]*?)(?:\s+with|$)', cl)
    if m:
        return m.group(1).capitalize(), m.group(2).strip().title()

    return None


def _save_pil_image(pil_img, dest_path):
    """Convert a PIL image (any mode) to JPEG and save to dest_path."""
    try:
        img = pil_img.convert("RGB").resize((224, 224), Image.BILINEAR)
        img.save(dest_path, "JPEG", quality=85)
        return True
    except Exception as e:
        print(f"  [IMG SAVE] {e}")
        return False


def load_RadyPlantDiseases(n):
    """
    Load Rady10/Plant-Diseases-Image-Text-Pairs from HuggingFace.
    Extract crop/disease from filename: [crop]_[disease]_[number].jpg
    Requires: datasets; for download-then-process: huggingface_hub, pyarrow.
    """
    name = "RadyPlantDiseases"
    desc = ("Rady Plant Diseases — HuggingFace dataset (Rady10/Plant-Diseases-Image-Text-Pairs). "
            "Image-text pairs with crop/disease encoded in filenames. "
            "Extract labels from filename format: crop_disease_number.jpg")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    if not HF_AVAILABLE:
        print("  [SKIP] RadyPlantDiseases — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    RADY_REPO_DIR = os.path.join(DATA_ROOT, "rady-plant-diseases-repo")
    dl = os.path.join(DATA_ROOT, "rady-plant-diseases-raw")

    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    try:
        os.makedirs(dl, exist_ok=True)
    except OSError as e:
        if e.errno == 122:
            print(f"  [RADY] Disk quota exceeded creating {dl}")
            print(f"  [RADY] Free space under {_BASE_DIR} or request more quota. Skipping RadyPlantDiseases.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        raise

    for stale_base in [os.path.join(IMAGES_DIR, name)]:
        if os.path.isdir(stale_base):
            shutil.rmtree(stale_base, ignore_errors=True)
            print(f"  [CLEANUP] Removed stale {name} output: {stale_base}")

    if "HUGGINGFACE_HUB_TOKEN" in os.environ:
        os.environ["HF_TOKEN"] = os.environ["HUGGINGFACE_HUB_TOKEN"]

    needed = (n if n is not None else 999999)
    scanned = 0
    new_classes_seen = set()
    t0 = time.time()
    GC_EVERY = 100

    def _row_to_pil(row_image):
        """Get PIL Image from parquet row or streaming row."""
        if row_image is None:
            return None
        if hasattr(row_image, "convert"):
            return row_image
        if isinstance(row_image, np.ndarray):
            return Image.fromarray(row_image)
        if isinstance(row_image, dict) and "bytes" in row_image:
            return Image.open(BytesIO(row_image["bytes"]))
        if isinstance(row_image, bytes):
            return Image.open(BytesIO(row_image))
        return None

    def _extract_crop_disease_from_filename(filename):
        """Extract crop and disease from filename: [crop]_[disease]_[number].jpg"""
        if not filename:
            return None
        basename = os.path.basename(filename)
        name_no_ext = os.path.splitext(basename)[0]
        parts = name_no_ext.rsplit("_", 1)
        if len(parts) < 2:
            return None
        rest = parts[0]
        rest_parts = rest.split("_")
        if len(rest_parts) < 2:
            return None
        crop = safe_name(rest_parts[0])
        disease = safe_name("_".join(rest_parts[1:]))
        return crop, disease

    use_local_parquet = HF_HUB_AVAILABLE and PYARROW_AVAILABLE
    parquet_files = []
    if use_local_parquet:
        parquet_files_check = []
        if os.path.isdir(RADY_REPO_DIR):
            for root, _, files in os.walk(RADY_REPO_DIR):
                parquet_files_check.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))

        if parquet_files_check:
            print(f"  [RADY] Found {len(parquet_files_check)} cached parquet file(s) — skipping re-download.")
            parquet_files = parquet_files_check
        else:
            try:
                os.makedirs(RADY_REPO_DIR, exist_ok=True)
                print(f"  [RADY] Downloading full dataset to disk -> {RADY_REPO_DIR}")
                snapshot_download("Rady10/Plant-Diseases-Image-Text-Pairs", repo_type="dataset", local_dir=RADY_REPO_DIR)
                print(f"  [RADY] Download complete.")
                for root, _, files in os.walk(RADY_REPO_DIR):
                    parquet_files.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))
            except Exception as e:
                err_str = str(e)
                is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
                print(f"  [RADY] Download failed: {e}")
                if is_quota:
                    print(f"  [RADY] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
                    print(f"  [RADY] Skipping RadyPlantDiseases.")
                    return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
                use_local_parquet = False

        if not parquet_files:
            print(f"  [RADY] No parquet files found, falling back to streaming.")
            use_local_parquet = False

    if use_local_parquet and parquet_files:
        print(f"  [RADY] Processing {len(parquet_files)} parquet file(s) from disk -> {dl}")
        for p_path in parquet_files:
            try:
                pf = pq.ParquetFile(p_path)
            except Exception:
                continue
            for batch in pf.iter_batches(batch_size=1):
                scanned += 1
                rows = batch.to_pylist()
                if not rows:
                    continue
                row = rows[0]
                del rows
                img_obj = row.get("image")
                file_path = ""
                if isinstance(img_obj, dict):
                    file_path = img_obj.get("path", "")
                elif isinstance(img_obj, str):
                    file_path = img_obj
                result = _extract_crop_disease_from_filename(file_path)
                if not result:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                crop, disease = result
                cls_label = safe_name(f"{crop}_{disease}")
                cls_dir = os.path.join(dl, cls_label)
                os.makedirs(cls_dir, exist_ok=True)
                try:
                    img = _row_to_pil(img_obj)
                    if img is None:
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                    if count >= needed:
                        del img
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                    img.convert("RGB").save(dest, "JPEG", quality=85)
                    del img
                except Exception:
                    pass
                del row
                if cls_label not in new_classes_seen:
                    new_classes_seen.add(cls_label)
                    print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
                if scanned % GC_EVERY == 0:
                    gc.collect()
                if scanned % 20_000 == 0:
                    print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")
    else:
        print(f"  [RADY] Streaming from HuggingFace -> {dl}")
        try:
            ds = hf_load_dataset("Rady10/Plant-Diseases-Image-Text-Pairs", split="train", streaming=True)
        except Exception as e:
            err_str = str(e)
            is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
            print(f"  [ERROR] Could not load RadyPlantDiseases: {e}")
            if is_quota:
                print(f"  [RADY] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        for i, row in enumerate(ds):
            scanned = i + 1
            img_obj = row.get("image")
            file_path = ""
            if isinstance(img_obj, dict):
                file_path = img_obj.get("path", "")
            elif isinstance(img_obj, str):
                file_path = img_obj
            result = _extract_crop_disease_from_filename(file_path)
            if not result:
                del row
                if scanned % GC_EVERY == 0:
                    gc.collect()
                continue
            crop, disease = result
            cls_label = safe_name(f"{crop}_{disease}")
            cls_dir = os.path.join(dl, cls_label)
            os.makedirs(cls_dir, exist_ok=True)
            try:
                img = _row_to_pil(img_obj)
                if img is None:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                if count >= needed:
                    del img
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                img.convert("RGB").save(dest, "JPEG", quality=85)
                del img
            except Exception:
                pass
            del row
            if cls_label not in new_classes_seen:
                new_classes_seen.add(cls_label)
                print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
            if scanned % GC_EVERY == 0:
                gc.collect()
            if scanned % 20_000 == 0:
                print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")

    print(f"\n  [RADY] Done: {len(new_classes_seen)} classes, {scanned:,} rows in {time.time()-t0:.1f}s")

    data    = collect_images_df(dl)
    classes = sorted(data[1].unique().tolist()) if len(data) else []

    if len(data) == 0:
        print(f"  [ERROR] No images saved to {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_A2H0H0R1PlantDisease(n):
    """
    Load A2H0H0R1/plant-disease-new from HuggingFace.
    Simple structure: image + text label (formatted as Crop__Disease).
    Requires: datasets; for download-then-process: huggingface_hub, pyarrow.
    """
    name = "A2H0H0R1PlantDisease"
    desc = ("A2H0H0R1 Plant Disease — HuggingFace dataset (A2H0H0R1/plant-disease-new). "
            "Clean classification dataset with pre-formatted crop__disease labels. "
            "1635 images with diverse crop and disease categories.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    if not HF_AVAILABLE:
        print("  [SKIP] A2H0H0R1PlantDisease — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    A2H_REPO_DIR = os.path.join(DATA_ROOT, "a2h0h0r1-plant-disease-repo")
    dl = os.path.join(DATA_ROOT, "a2h0h0r1-plant-disease-raw")

    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    try:
        os.makedirs(dl, exist_ok=True)
    except OSError as e:
        if e.errno == 122:
            print(f"  [A2H] Disk quota exceeded creating {dl}")
            print(f"  [A2H] Free space under {_BASE_DIR} or request more quota. Skipping A2H0H0R1PlantDisease.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        raise

    for stale_base in [os.path.join(IMAGES_DIR, name)]:
        if os.path.isdir(stale_base):
            shutil.rmtree(stale_base, ignore_errors=True)
            print(f"  [CLEANUP] Removed stale {name} output: {stale_base}")

    if "HUGGINGFACE_HUB_TOKEN" in os.environ:
        os.environ["HF_TOKEN"] = os.environ["HUGGINGFACE_HUB_TOKEN"]

    needed = (n if n is not None else 999999)
    scanned = 0
    new_classes_seen = set()
    t0 = time.time()
    GC_EVERY = 100

    def _row_to_pil(row_image):
        """Get PIL Image from parquet row or streaming row."""
        if row_image is None:
            return None
        if hasattr(row_image, "convert"):
            return row_image
        if isinstance(row_image, np.ndarray):
            return Image.fromarray(row_image)
        if isinstance(row_image, dict) and "bytes" in row_image:
            return Image.open(BytesIO(row_image["bytes"]))
        if isinstance(row_image, bytes):
            return Image.open(BytesIO(row_image))
        return None

    use_local_parquet = HF_HUB_AVAILABLE and PYARROW_AVAILABLE
    parquet_files = []
    if use_local_parquet:
        parquet_files_check = []
        if os.path.isdir(A2H_REPO_DIR):
            for root, _, files in os.walk(A2H_REPO_DIR):
                parquet_files_check.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))

        if parquet_files_check:
            print(f"  [A2H] Found {len(parquet_files_check)} cached parquet file(s) — skipping re-download.")
            parquet_files = parquet_files_check
        else:
            try:
                os.makedirs(A2H_REPO_DIR, exist_ok=True)
                print(f"  [A2H] Downloading full dataset to disk -> {A2H_REPO_DIR}")
                snapshot_download("A2H0H0R1/plant-disease-new", repo_type="dataset", local_dir=A2H_REPO_DIR)
                print(f"  [A2H] Download complete.")
                for root, _, files in os.walk(A2H_REPO_DIR):
                    parquet_files.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))
            except Exception as e:
                err_str = str(e)
                is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
                print(f"  [A2H] Download failed: {e}")
                if is_quota:
                    print(f"  [A2H] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
                    print(f"  [A2H] Skipping A2H0H0R1PlantDisease.")
                    return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
                use_local_parquet = False

        if not parquet_files:
            print(f"  [A2H] No parquet files found, falling back to streaming.")
            use_local_parquet = False

    if use_local_parquet and parquet_files:
        print(f"  [A2H] Processing {len(parquet_files)} parquet file(s) from disk -> {dl}")
        for p_path in parquet_files:
            try:
                pf = pq.ParquetFile(p_path)
            except Exception:
                continue
            for batch in pf.iter_batches(batch_size=1):
                scanned += 1
                rows = batch.to_pylist()
                if not rows:
                    continue
                row = rows[0]
                del rows
                label = str(row.get("label") or "").strip()
                if not label:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                label = safe_name(label)
                cls_dir = os.path.join(dl, label)
                os.makedirs(cls_dir, exist_ok=True)
                try:
                    img = _row_to_pil(row.get("image"))
                    if img is None:
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                    if count >= needed:
                        del img
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                    img.convert("RGB").save(dest, "JPEG", quality=85)
                    del img
                except Exception:
                    pass
                del row
                if label not in new_classes_seen:
                    new_classes_seen.add(label)
                    print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {label}  ({time.time()-t0:.0f}s)")
                if scanned % GC_EVERY == 0:
                    gc.collect()
                if scanned % 20_000 == 0:
                    print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")
    else:
        print(f"  [A2H] Streaming from HuggingFace -> {dl}")
        try:
            ds = hf_load_dataset("A2H0H0R1/plant-disease-new", split="train", streaming=True)
        except Exception as e:
            err_str = str(e)
            is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
            print(f"  [ERROR] Could not load A2H0H0R1PlantDisease: {e}")
            if is_quota:
                print(f"  [A2H] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        for i, row in enumerate(ds):
            scanned = i + 1
            label = str(row.get("label") or "").strip()
            if not label:
                del row
                if scanned % GC_EVERY == 0:
                    gc.collect()
                continue
            label = safe_name(label)
            cls_dir = os.path.join(dl, label)
            os.makedirs(cls_dir, exist_ok=True)
            try:
                img = _row_to_pil(row.get("image"))
                if img is None:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                if count >= needed:
                    del img
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                img.convert("RGB").save(dest, "JPEG", quality=85)
                del img
            except Exception:
                pass
            del row
            if label not in new_classes_seen:
                new_classes_seen.add(label)
                print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {label}  ({time.time()-t0:.0f}s)")
            if scanned % GC_EVERY == 0:
                gc.collect()
            if scanned % 20_000 == 0:
                print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")

    print(f"\n  [A2H] Done: {len(new_classes_seen)} classes, {scanned:,} rows in {time.time()-t0:.1f}s")

    data    = collect_images_df(dl)
    classes = sorted(data[1].unique().tolist()) if len(data) else []

    if len(data) == 0:
        print(f"  [ERROR] No images saved to {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_AvinashPlantDisease(n):
    """
    Load avinashhm/plant-disease-classification-complete from HuggingFace.
    Structure: image + numeric label + file path (contains crop/disease).
    Extract crop/disease from paths like [crop]/[disease]/file.jpg
    Requires: datasets; for download-then-process: huggingface_hub, pyarrow.
    """
    name = "AvinashPlantDisease"
    desc = ("Avinash Plant Disease — HuggingFace dataset (avinashhm/plant-disease-classification-complete). "
            "Complete plant disease classification with numeric labels and hierarchical paths. "
            "72.8k images with crop and disease extracted from file paths.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    if not HF_AVAILABLE:
        print("  [SKIP] AvinashPlantDisease — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    AVINASH_REPO_DIR = os.path.join(DATA_ROOT, "avinash-plant-disease-repo")
    dl = os.path.join(DATA_ROOT, "avinash-plant-disease-raw")

    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    try:
        os.makedirs(dl, exist_ok=True)
    except OSError as e:
        if e.errno == 122:
            print(f"  [AVINASH] Disk quota exceeded creating {dl}")
            print(f"  [AVINASH] Free space under {_BASE_DIR} or request more quota. Skipping AvinashPlantDisease.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        raise

    for stale_base in [os.path.join(IMAGES_DIR, name)]:
        if os.path.isdir(stale_base):
            shutil.rmtree(stale_base, ignore_errors=True)
            print(f"  [CLEANUP] Removed stale {name} output: {stale_base}")

    if "HUGGINGFACE_HUB_TOKEN" in os.environ:
        os.environ["HF_TOKEN"] = os.environ["HUGGINGFACE_HUB_TOKEN"]

    needed = (n if n is not None else 999999)
    scanned = 0
    new_classes_seen = set()
    t0 = time.time()
    GC_EVERY = 100

    def _row_to_pil(row_image):
        """Get PIL Image from parquet row or streaming row."""
        if row_image is None:
            return None
        if hasattr(row_image, "convert"):
            return row_image
        if isinstance(row_image, np.ndarray):
            return Image.fromarray(row_image)
        if isinstance(row_image, dict) and "bytes" in row_image:
            return Image.open(BytesIO(row_image["bytes"]))
        if isinstance(row_image, bytes):
            return Image.open(BytesIO(row_image))
        return None

    def _extract_crop_disease_from_path(file_path):
        """Extract crop and disease from path like [crop]/[disease]/file.jpg"""
        if not file_path:
            return None
        path_str = file_path.strip().replace("\\", "/")
        parts = [p.strip() for p in path_str.split("/") if p.strip()]
        if len(parts) >= 2:
            disease = safe_name(parts[-2])
            crop = safe_name(parts[-3] if len(parts) >= 3 else "Unknown")
            return crop, disease
        return None

    use_local_parquet = HF_HUB_AVAILABLE and PYARROW_AVAILABLE
    parquet_files = []
    if use_local_parquet:
        parquet_files_check = []
        if os.path.isdir(AVINASH_REPO_DIR):
            for root, _, files in os.walk(AVINASH_REPO_DIR):
                parquet_files_check.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))

        if parquet_files_check:
            print(f"  [AVINASH] Found {len(parquet_files_check)} cached parquet file(s) — skipping re-download.")
            parquet_files = parquet_files_check
        else:
            try:
                os.makedirs(AVINASH_REPO_DIR, exist_ok=True)
                print(f"  [AVINASH] Downloading full dataset to disk -> {AVINASH_REPO_DIR}")
                snapshot_download("avinashhm/plant-disease-classification-complete", repo_type="dataset", local_dir=AVINASH_REPO_DIR)
                print(f"  [AVINASH] Download complete.")
                for root, _, files in os.walk(AVINASH_REPO_DIR):
                    parquet_files.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))
            except Exception as e:
                err_str = str(e)
                is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
                print(f"  [AVINASH] Download failed: {e}")
                if is_quota:
                    print(f"  [AVINASH] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
                    print(f"  [AVINASH] Skipping AvinashPlantDisease.")
                    return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
                use_local_parquet = False

        if not parquet_files:
            print(f"  [AVINASH] No parquet files found, falling back to streaming.")
            use_local_parquet = False

    if use_local_parquet and parquet_files:
        print(f"  [AVINASH] Processing {len(parquet_files)} parquet file(s) from disk -> {dl}")
        for p_path in parquet_files:
            try:
                pf = pq.ParquetFile(p_path)
            except Exception:
                continue
            for batch in pf.iter_batches(batch_size=1):
                scanned += 1
                rows = batch.to_pylist()
                if not rows:
                    continue
                row = rows[0]
                del rows
                file_path = row.get("path", "")
                result = _extract_crop_disease_from_path(file_path)
                if not result:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                crop, disease = result
                cls_label = safe_name(f"{crop}_{disease}")
                cls_dir = os.path.join(dl, cls_label)
                os.makedirs(cls_dir, exist_ok=True)
                try:
                    img = _row_to_pil(row.get("image"))
                    if img is None:
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                    if count >= needed:
                        del img
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                    img.convert("RGB").save(dest, "JPEG", quality=85)
                    del img
                except Exception:
                    pass
                del row
                if cls_label not in new_classes_seen:
                    new_classes_seen.add(cls_label)
                    print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
                if scanned % GC_EVERY == 0:
                    gc.collect()
                if scanned % 20_000 == 0:
                    print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")
    else:
        print(f"  [AVINASH] Streaming from HuggingFace -> {dl}")
        try:
            ds = hf_load_dataset("avinashhm/plant-disease-classification-complete", split="train", streaming=True)
        except Exception as e:
            err_str = str(e)
            is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
            print(f"  [ERROR] Could not load AvinashPlantDisease: {e}")
            if is_quota:
                print(f"  [AVINASH] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        for i, row in enumerate(ds):
            scanned = i + 1
            file_path = row.get("path", "")
            result = _extract_crop_disease_from_path(file_path)
            if not result:
                del row
                if scanned % GC_EVERY == 0:
                    gc.collect()
                continue
            crop, disease = result
            cls_label = safe_name(f"{crop}_{disease}")
            cls_dir = os.path.join(dl, cls_label)
            os.makedirs(cls_dir, exist_ok=True)
            try:
                img = _row_to_pil(row.get("image"))
                if img is None:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                if count >= needed:
                    del img
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                img.convert("RGB").save(dest, "JPEG", quality=85)
                del img
            except Exception:
                pass
            del row
            if cls_label not in new_classes_seen:
                new_classes_seen.add(cls_label)
                print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
            if scanned % GC_EVERY == 0:
                gc.collect()
            if scanned % 20_000 == 0:
                print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")

    print(f"\n  [AVINASH] Done: {len(new_classes_seen)} classes, {scanned:,} rows in {time.time()-t0:.1f}s")

    data    = collect_images_df(dl)
    classes = sorted(data[1].unique().tolist()) if len(data) else []

    if len(data) == 0:
        print(f"  [ERROR] No images saved to {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_SakethPlantDisease(n):
    """
    Load sakethdevx/plant-disease-dataset from HuggingFace.
    File path-based structure: image column contains paths like [crop]/[disease]/file.jpg
    Requires: datasets; for download-then-process: huggingface_hub, pyarrow.
    """
    name = "SakethPlantDisease"
    desc = ("Saketh Plant Disease — HuggingFace dataset (sakethdevx/plant-disease-dataset). "
            "Hierarchical structure with crop and disease extracted from file paths. "
            "87.3k images organized by crop type and disease category.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    if not HF_AVAILABLE:
        print("  [SKIP] SakethPlantDisease — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    SAKETH_REPO_DIR = os.path.join(DATA_ROOT, "saketh-plant-disease-repo")
    dl = os.path.join(DATA_ROOT, "saketh-plant-disease-raw")

    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    try:
        os.makedirs(dl, exist_ok=True)
    except OSError as e:
        if e.errno == 122:
            print(f"  [SAKETH] Disk quota exceeded creating {dl}")
            print(f"  [SAKETH] Free space under {_BASE_DIR} or request more quota. Skipping SakethPlantDisease.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        raise

    for stale_base in [os.path.join(IMAGES_DIR, name)]:
        if os.path.isdir(stale_base):
            shutil.rmtree(stale_base, ignore_errors=True)
            print(f"  [CLEANUP] Removed stale {name} output: {stale_base}")

    if "HUGGINGFACE_HUB_TOKEN" in os.environ:
        os.environ["HF_TOKEN"] = os.environ["HUGGINGFACE_HUB_TOKEN"]

    needed = (n if n is not None else 999999)
    scanned = 0
    new_classes_seen = set()
    t0 = time.time()
    GC_EVERY = 100

    def _row_to_pil(row_image):
        """Get PIL Image from parquet row or streaming row."""
        if row_image is None:
            return None
        if hasattr(row_image, "convert"):
            return row_image
        if isinstance(row_image, np.ndarray):
            return Image.fromarray(row_image)
        if isinstance(row_image, dict) and "bytes" in row_image:
            return Image.open(BytesIO(row_image["bytes"]))
        if isinstance(row_image, bytes):
            return Image.open(BytesIO(row_image))
        return None

    def _extract_crop_disease_from_path(file_path):
        """Extract crop and disease from path like [crop]/[disease]/file.jpg"""
        if not file_path:
            return None
        parts = file_path.strip().replace("\\", "/").split("/")
        parts = [p for p in parts if p]
        if len(parts) >= 2:
            crop = safe_name(parts[-3] if len(parts) >= 3 else "Unknown")
            disease = safe_name(parts[-2] if len(parts) >= 2 else "Unknown")
            return crop, disease
        return None

    use_local_parquet = HF_HUB_AVAILABLE and PYARROW_AVAILABLE
    parquet_files = []
    if use_local_parquet:
        parquet_files_check = []
        if os.path.isdir(SAKETH_REPO_DIR):
            for root, _, files in os.walk(SAKETH_REPO_DIR):
                parquet_files_check.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))

        if parquet_files_check:
            print(f"  [SAKETH] Found {len(parquet_files_check)} cached parquet file(s) — skipping re-download.")
            parquet_files = parquet_files_check
        else:
            try:
                os.makedirs(SAKETH_REPO_DIR, exist_ok=True)
                print(f"  [SAKETH] Downloading full dataset to disk -> {SAKETH_REPO_DIR}")
                snapshot_download("sakethdevx/plant-disease-dataset", repo_type="dataset", local_dir=SAKETH_REPO_DIR)
                print(f"  [SAKETH] Download complete.")
                for root, _, files in os.walk(SAKETH_REPO_DIR):
                    parquet_files.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))
            except Exception as e:
                err_str = str(e)
                is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
                print(f"  [SAKETH] Download failed: {e}")
                if is_quota:
                    print(f"  [SAKETH] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
                    print(f"  [SAKETH] Skipping SakethPlantDisease.")
                    return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
                use_local_parquet = False

        if not parquet_files:
            print(f"  [SAKETH] No parquet files found, falling back to streaming.")
            use_local_parquet = False

    if use_local_parquet and parquet_files:
        print(f"  [SAKETH] Processing {len(parquet_files)} parquet file(s) from disk -> {dl}")
        for p_path in parquet_files:
            try:
                pf = pq.ParquetFile(p_path)
            except Exception:
                continue
            for batch in pf.iter_batches(batch_size=1):
                scanned += 1
                rows = batch.to_pylist()
                if not rows:
                    continue
                row = rows[0]
                del rows
                file_path = row.get("image_file_path", row.get("path", ""))
                result = _extract_crop_disease_from_path(file_path)
                if not result:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                crop, disease = result
                cls_label = safe_name(f"{crop}_{disease}")
                cls_dir = os.path.join(dl, cls_label)
                os.makedirs(cls_dir, exist_ok=True)
                try:
                    img = _row_to_pil(row.get("image"))
                    if img is None:
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                    if count >= needed:
                        del img
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                    img.convert("RGB").save(dest, "JPEG", quality=85)
                    del img
                except Exception:
                    pass
                del row
                if cls_label not in new_classes_seen:
                    new_classes_seen.add(cls_label)
                    print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
                if scanned % GC_EVERY == 0:
                    gc.collect()
                if scanned % 20_000 == 0:
                    print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")
    else:
        print(f"  [SAKETH] Streaming from HuggingFace -> {dl}")
        try:
            ds = hf_load_dataset("sakethdevx/plant-disease-dataset", split="train", streaming=True)
        except Exception as e:
            err_str = str(e)
            is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
            print(f"  [ERROR] Could not load SakethPlantDisease: {e}")
            if is_quota:
                print(f"  [SAKETH] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        for i, row in enumerate(ds):
            scanned = i + 1
            file_path = row.get("image_file_path", row.get("path", ""))
            result = _extract_crop_disease_from_path(file_path)
            if not result:
                del row
                if scanned % GC_EVERY == 0:
                    gc.collect()
                continue
            crop, disease = result
            cls_label = safe_name(f"{crop}_{disease}")
            cls_dir = os.path.join(dl, cls_label)
            os.makedirs(cls_dir, exist_ok=True)
            try:
                img = _row_to_pil(row.get("image"))
                if img is None:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                if count >= needed:
                    del img
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                img.convert("RGB").save(dest, "JPEG", quality=85)
                del img
            except Exception:
                pass
            del row
            if cls_label not in new_classes_seen:
                new_classes_seen.add(cls_label)
                print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
            if scanned % GC_EVERY == 0:
                gc.collect()
            if scanned % 20_000 == 0:
                print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")

    print(f"\n  [SAKETH] Done: {len(new_classes_seen)} classes, {scanned:,} rows in {time.time()-t0:.1f}s")

    data    = collect_images_df(dl)
    classes = sorted(data[1].unique().tolist()) if len(data) else []

    if len(data) == 0:
        print(f"  [ERROR] No images saved to {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_VQAPlantDisease(n):
    """
    Load raghavendrad60/vqa_plant-disease-classification-merged-dataset from HuggingFace.
    VQA structure: image + question + answer + pre-computed class columns.
    Requires: datasets; for download-then-process: huggingface_hub, pyarrow.
    """
    name = "VQAPlantDisease"
    desc = ("VQA Plant Disease — HuggingFace dataset (raghavendrad60/vqa_plant-disease-classification-merged-dataset). "
            "Visual Question Answering for plant disease identification with pre-computed class labels. "
            "21 plant-disease classes from merged dataset.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    if not HF_AVAILABLE:
        print("  [SKIP] VQAPlantDisease — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    VQA_REPO_DIR = os.path.join(DATA_ROOT, "vqa-plant-disease-repo")
    dl = os.path.join(DATA_ROOT, "vqa-plant-disease-raw")

    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    try:
        os.makedirs(dl, exist_ok=True)
    except OSError as e:
        if e.errno == 122:
            print(f"  [VQAPLANT] Disk quota exceeded creating {dl}")
            print(f"  [VQAPLANT] Free space under {_BASE_DIR} or request more quota. Skipping VQAPlantDisease.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        raise

    for stale_base in [os.path.join(IMAGES_DIR, name)]:
        if os.path.isdir(stale_base):
            shutil.rmtree(stale_base, ignore_errors=True)
            print(f"  [CLEANUP] Removed stale {name} output: {stale_base}")

    if "HUGGINGFACE_HUB_TOKEN" in os.environ:
        os.environ["HF_TOKEN"] = os.environ["HUGGINGFACE_HUB_TOKEN"]

    needed = (n if n is not None else 999999)
    scanned = 0
    new_classes_seen = set()
    t0 = time.time()
    GC_EVERY = 100

    def _row_to_pil(row_image):
        """Get PIL Image from parquet row or streaming row."""
        if row_image is None:
            return None
        if hasattr(row_image, "convert"):
            return row_image
        if isinstance(row_image, np.ndarray):
            return Image.fromarray(row_image)
        if isinstance(row_image, dict) and "bytes" in row_image:
            return Image.open(BytesIO(row_image["bytes"]))
        if isinstance(row_image, bytes):
            return Image.open(BytesIO(row_image))
        return None

    use_local_parquet = HF_HUB_AVAILABLE and PYARROW_AVAILABLE
    parquet_files = []
    if use_local_parquet:
        parquet_files_check = []
        if os.path.isdir(VQA_REPO_DIR):
            for root, _, files in os.walk(VQA_REPO_DIR):
                parquet_files_check.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))

        if parquet_files_check:
            print(f"  [VQAPLANT] Found {len(parquet_files_check)} cached parquet file(s) — skipping re-download.")
            parquet_files = parquet_files_check
        else:
            try:
                os.makedirs(VQA_REPO_DIR, exist_ok=True)
                print(f"  [VQAPLANT] Downloading full dataset to disk -> {VQA_REPO_DIR}")
                snapshot_download("raghavendrad60/vqa_plant-disease-classification-merged-dataset", repo_type="dataset", local_dir=VQA_REPO_DIR)
                print(f"  [VQAPLANT] Download complete.")
                for root, _, files in os.walk(VQA_REPO_DIR):
                    parquet_files.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))
            except Exception as e:
                err_str = str(e)
                is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
                print(f"  [VQAPLANT] Download failed: {e}")
                if is_quota:
                    print(f"  [VQAPLANT] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
                    print(f"  [VQAPLANT] Skipping VQAPlantDisease.")
                    return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
                use_local_parquet = False

        if not parquet_files:
            print(f"  [VQAPLANT] No parquet files found, falling back to streaming.")
            use_local_parquet = False

    if use_local_parquet and parquet_files:
        print(f"  [VQAPLANT] Processing {len(parquet_files)} parquet file(s) from disk -> {dl}")
        for p_path in parquet_files:
            try:
                pf = pq.ParquetFile(p_path)
            except Exception:
                continue
            for batch in pf.iter_batches(batch_size=1):
                scanned += 1
                rows = batch.to_pylist()
                if not rows:
                    continue
                row = rows[0]
                del rows
                cls_label = row.get("class", "").strip()
                if not cls_label:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                cls_label = safe_name(cls_label)
                cls_dir = os.path.join(dl, cls_label)
                os.makedirs(cls_dir, exist_ok=True)
                try:
                    img = _row_to_pil(row.get("image"))
                    if img is None:
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                    if count >= needed:
                        del img
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                    img.convert("RGB").save(dest, "JPEG", quality=85)
                    del img
                except Exception:
                    pass
                del row
                if cls_label not in new_classes_seen:
                    new_classes_seen.add(cls_label)
                    print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
                if scanned % GC_EVERY == 0:
                    gc.collect()
                if scanned % 20_000 == 0:
                    print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")
    else:
        print(f"  [VQAPLANT] Streaming from HuggingFace -> {dl}")
        try:
            ds = hf_load_dataset("raghavendrad60/vqa_plant-disease-classification-merged-dataset", split="train", streaming=True)
        except Exception as e:
            err_str = str(e)
            is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
            print(f"  [ERROR] Could not load VQAPlantDisease: {e}")
            if is_quota:
                print(f"  [VQAPLANT] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        for i, row in enumerate(ds):
            scanned = i + 1
            cls_label = row.get("class", "").strip()
            if not cls_label:
                del row
                if scanned % GC_EVERY == 0:
                    gc.collect()
                continue
            cls_label = safe_name(cls_label)
            cls_dir = os.path.join(dl, cls_label)
            os.makedirs(cls_dir, exist_ok=True)
            try:
                img = _row_to_pil(row.get("image"))
                if img is None:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                if count >= needed:
                    del img
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                img.convert("RGB").save(dest, "JPEG", quality=85)
                del img
            except Exception:
                pass
            del row
            if cls_label not in new_classes_seen:
                new_classes_seen.add(cls_label)
                print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
            if scanned % GC_EVERY == 0:
                gc.collect()
            if scanned % 20_000 == 0:
                print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")

    print(f"\n  [VQAPLANT] Done: {len(new_classes_seen)} classes, {scanned:,} rows in {time.time()-t0:.1f}s")

    data    = collect_images_df(dl)
    classes = sorted(data[1].unique().tolist()) if len(data) else []

    if len(data) == 0:
        print(f"  [ERROR] No images saved to {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_BDCropVegetable(n):
    """
    Load Saon110/bd-crop-vegetable-plant-disease-dataset from HuggingFace.
    Simple structure: image + label_name columns.
    Requires: datasets; for download-then-process: huggingface_hub, pyarrow.
    """
    name = "BDCropVegetable"
    desc = ("BD Crop Vegetable — HuggingFace dataset (Saon110/bd-crop-vegetable-plant-disease-dataset). "
            "Bengali vegetable disease classification with direct label names. "
            "26 crop-disease classes covering banana, brinjal, and other vegetables.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    if not HF_AVAILABLE:
        print("  [SKIP] BDCropVegetable — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    BD_REPO_DIR = os.path.join(DATA_ROOT, "bd-crop-vegetable-repo")
    dl = os.path.join(DATA_ROOT, "bd-crop-vegetable-raw")

    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    try:
        os.makedirs(dl, exist_ok=True)
    except OSError as e:
        if e.errno == 122:
            print(f"  [BDCROPVEG] Disk quota exceeded creating {dl}")
            print(f"  [BDCROPVEG] Free space under {_BASE_DIR} or request more quota. Skipping BDCropVegetable.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        raise

    for stale_base in [os.path.join(IMAGES_DIR, name)]:
        if os.path.isdir(stale_base):
            shutil.rmtree(stale_base, ignore_errors=True)
            print(f"  [CLEANUP] Removed stale {name} output: {stale_base}")

    if "HUGGINGFACE_HUB_TOKEN" in os.environ:
        os.environ["HF_TOKEN"] = os.environ["HUGGINGFACE_HUB_TOKEN"]

    needed = (n if n is not None else 999999)
    scanned = 0
    new_classes_seen = set()
    t0 = time.time()
    GC_EVERY = 100

    def _row_to_pil(row_image):
        """Get PIL Image from parquet row or streaming row."""
        if row_image is None:
            return None
        if hasattr(row_image, "convert"):
            return row_image
        if isinstance(row_image, np.ndarray):
            return Image.fromarray(row_image)
        if isinstance(row_image, dict) and "bytes" in row_image:
            return Image.open(BytesIO(row_image["bytes"]))
        if isinstance(row_image, bytes):
            return Image.open(BytesIO(row_image))
        return None

    use_local_parquet = HF_HUB_AVAILABLE and PYARROW_AVAILABLE
    parquet_files = []
    if use_local_parquet:
        parquet_files_check = []
        if os.path.isdir(BD_REPO_DIR):
            for root, _, files in os.walk(BD_REPO_DIR):
                parquet_files_check.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))

        if parquet_files_check:
            print(f"  [BDCROPVEG] Found {len(parquet_files_check)} cached parquet file(s) — skipping re-download.")
            parquet_files = parquet_files_check
        else:
            try:
                os.makedirs(BD_REPO_DIR, exist_ok=True)
                print(f"  [BDCROPVEG] Downloading full dataset to disk -> {BD_REPO_DIR}")
                snapshot_download("Saon110/bd-crop-vegetable-plant-disease-dataset", repo_type="dataset", local_dir=BD_REPO_DIR)
                print(f"  [BDCROPVEG] Download complete.")
                for root, _, files in os.walk(BD_REPO_DIR):
                    parquet_files.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))
            except Exception as e:
                err_str = str(e)
                is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
                print(f"  [BDCROPVEG] Download failed: {e}")
                if is_quota:
                    print(f"  [BDCROPVEG] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
                    print(f"  [BDCROPVEG] Skipping BDCropVegetable.")
                    return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
                use_local_parquet = False

        if not parquet_files:
            print(f"  [BDCROPVEG] No parquet files found, falling back to streaming.")
            use_local_parquet = False

    if use_local_parquet and parquet_files:
        print(f"  [BDCROPVEG] Processing {len(parquet_files)} parquet file(s) from disk -> {dl}")
        for p_path in parquet_files:
            try:
                pf = pq.ParquetFile(p_path)
            except Exception:
                continue
            for batch in pf.iter_batches(batch_size=1):
                scanned += 1
                rows = batch.to_pylist()
                if not rows:
                    continue
                row = rows[0]
                del rows
                cls_label = row.get("label_name", "").strip()
                if not cls_label:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                cls_label = safe_name(cls_label)
                cls_dir = os.path.join(dl, cls_label)
                os.makedirs(cls_dir, exist_ok=True)
                try:
                    img = _row_to_pil(row.get("image"))
                    if img is None:
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                    if count >= needed:
                        del img
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                    img.convert("RGB").save(dest, "JPEG", quality=85)
                    del img
                except Exception:
                    pass
                del row
                if cls_label not in new_classes_seen:
                    new_classes_seen.add(cls_label)
                    print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
                if scanned % GC_EVERY == 0:
                    gc.collect()
                if scanned % 20_000 == 0:
                    print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")
    else:
        print(f"  [BDCROPVEG] Streaming from HuggingFace -> {dl}")
        try:
            ds = hf_load_dataset("Saon110/bd-crop-vegetable-plant-disease-dataset", split="train", streaming=True)
        except Exception as e:
            err_str = str(e)
            is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
            print(f"  [ERROR] Could not load BDCropVegetable: {e}")
            if is_quota:
                print(f"  [BDCROPVEG] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        for i, row in enumerate(ds):
            scanned = i + 1
            cls_label = row.get("label_name", "").strip()
            if not cls_label:
                del row
                if scanned % GC_EVERY == 0:
                    gc.collect()
                continue
            cls_label = safe_name(cls_label)
            cls_dir = os.path.join(dl, cls_label)
            os.makedirs(cls_dir, exist_ok=True)
            try:
                img = _row_to_pil(row.get("image"))
                if img is None:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                if count >= needed:
                    del img
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                img.convert("RGB").save(dest, "JPEG", quality=85)
                del img
            except Exception:
                pass
            del row
            if cls_label not in new_classes_seen:
                new_classes_seen.add(cls_label)
                print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
            if scanned % GC_EVERY == 0:
                gc.collect()
            if scanned % 20_000 == 0:
                print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")

    print(f"\n  [BDCROPVEG] Done: {len(new_classes_seen)} classes, {scanned:,} rows in {time.time()-t0:.1f}s")

    data    = collect_images_df(dl)
    classes = sorted(data[1].unique().tolist()) if len(data) else []

    if len(data) == 0:
        print(f"  [ERROR] No images saved to {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


def load_LeafNet(n):
    """
    Load enalis/LeafNet: all paths under work root (data/leafnet-repo, data/leafnet-raw).
    With huggingface_hub + pyarrow: download once, then process one row at a time.
    Fallback: stream from API (may OOM).
    Requires: datasets; for download-then-process: huggingface_hub, pyarrow.
    """
    name = "LeafNet"
    desc = ("LeafNet — HuggingFace dataset (enalis/LeafNet). "
            "Multi-crop leaf disease classification derived from natural-language "
            "captions. Classes are (Crop, Disease) pairs spanning diverse species "
            "and conditions including Healthy controls.")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    if not HF_AVAILABLE:
        print("  [SKIP] LeafNet — 'datasets' package not installed. "
              "Run:  pip install datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # All LeafNet paths under hardcoded work root.
    LEAFNET_REPO_DIR = os.path.join(DATA_ROOT, "leafnet-repo")
    dl = os.path.join(DATA_ROOT, "leafnet-raw")

    # Wipe any previous staging run so class folders start clean
    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    try:
        os.makedirs(dl, exist_ok=True)
    except OSError as e:
        if e.errno == 122:  # Disk quota exceeded
            print(f"  [LEAFNET] Disk quota exceeded creating {dl}")
            print(f"  [LEAFNET] Free space under {_BASE_DIR} or request more quota. Skipping LeafNet.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        raise

    # Wipe stale curated output (wrong structure from previous runs)
    for stale_base in [os.path.join(IMAGES_DIR, "LeafNet")]:
        if os.path.isdir(stale_base):
            shutil.rmtree(stale_base, ignore_errors=True)
            print(f"  [CLEANUP] Removed stale LeafNet output: {stale_base}")

    if "HUGGINGFACE_HUB_TOKEN" in os.environ:
        os.environ["HF_TOKEN"] = os.environ["HUGGINGFACE_HUB_TOKEN"]

    needed = (n if n is not None else 999999)
    scanned = 0
    new_classes_seen = set()
    t0 = time.time()
    GC_EVERY = 100

    def _row_to_pil(row_image):
        """Get PIL Image from parquet row 'image' (struct with bytes) or streaming row (PIL/ndarray)."""
        if row_image is None:
            return None
        if hasattr(row_image, "convert"):
            return row_image
        if isinstance(row_image, np.ndarray):
            return Image.fromarray(row_image)
        if isinstance(row_image, dict) and "bytes" in row_image:
            return Image.open(BytesIO(row_image["bytes"]))
        if isinstance(row_image, bytes):
            return Image.open(BytesIO(row_image))
        return None

    # Prefer: download all data to disk, then process one-by-one from local parquet (avoids stream OOM).
    use_local_parquet = HF_HUB_AVAILABLE and PYARROW_AVAILABLE
    parquet_files = []
    if use_local_parquet:
        if not os.path.isdir(LEAFNET_REPO_DIR):
            try:
                os.makedirs(LEAFNET_REPO_DIR, exist_ok=True)
                print(f"  [LEAFNET] Downloading full dataset to disk -> {LEAFNET_REPO_DIR}")
                snapshot_download("enalis/LeafNet", repo_type="dataset", local_dir=LEAFNET_REPO_DIR)
                print(f"  [LEAFNET] Download complete.")
            except Exception as e:
                err_str = str(e)
                is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
                print(f"  [LEAFNET] Download failed: {e}")
                if is_quota:
                    print(f"  [LEAFNET] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
                    print(f"  [LEAFNET] Skipping LeafNet.")
                    return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
                use_local_parquet = False
        if use_local_parquet:
            for root, _, files in os.walk(LEAFNET_REPO_DIR):
                parquet_files.extend(os.path.join(root, f) for f in files if f.endswith(".parquet"))
            if not parquet_files:
                print(f"  [LEAFNET] No parquet files in {LEAFNET_REPO_DIR}, falling back to streaming.")
                use_local_parquet = False

    if use_local_parquet and parquet_files:
        print(f"  [LEAFNET] Processing {len(parquet_files)} parquet file(s) from disk (one row at a time) -> {dl}")
        for p_path in parquet_files:
            try:
                pf = pq.ParquetFile(p_path)
            except Exception:
                continue
            for batch in pf.iter_batches(batch_size=1):
                scanned += 1
                rows = batch.to_pylist()
                if not rows:
                    continue
                row = rows[0]
                del rows
                caption = row.get("caption") or row.get("text") or ""
                result = _extract_crop_disease(caption)
                if not result:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                crop, disease = result
                cls_label = safe_name(f"{crop}_{disease}")
                cls_dir = os.path.join(dl, cls_label)
                os.makedirs(cls_dir, exist_ok=True)
                try:
                    img = _row_to_pil(row.get("image"))
                    if img is None:
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                    if count >= needed:
                        del img
                        del row
                        if scanned % GC_EVERY == 0:
                            gc.collect()
                        continue
                    dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                    img.convert("RGB").save(dest, "JPEG", quality=85)
                    del img
                except Exception:
                    pass
                del row
                if cls_label not in new_classes_seen:
                    new_classes_seen.add(cls_label)
                    print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
                if scanned % GC_EVERY == 0:
                    gc.collect()
                if scanned % 20_000 == 0:
                    print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")
    else:
        # Fallback: stream from HuggingFace API (one row at a time; may OOM on low memory).
        print(f"  [LEAFNET] Streaming from HuggingFace -> {dl}")
        try:
            ds = hf_load_dataset("enalis/LeafNet", split="train", streaming=True)
        except Exception as e:
            err_str = str(e)
            is_quota = "122" in err_str or "quota" in err_str.lower() or "Disk" in err_str
            print(f"  [ERROR] Could not load LeafNet: {e}")
            if is_quota:
                print(f"  [LEAFNET] Disk quota exceeded. Free space under {_BASE_DIR} or request more quota.")
            return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc
        for i, row in enumerate(ds):
            scanned = i + 1
            result = _extract_crop_disease(row.get("caption", ""))
            if not result:
                del row
                if scanned % GC_EVERY == 0:
                    gc.collect()
                continue
            crop, disease = result
            cls_label = safe_name(f"{crop}_{disease}")
            cls_dir = os.path.join(dl, cls_label)
            os.makedirs(cls_dir, exist_ok=True)
            try:
                img = _row_to_pil(row.get("image"))
                if img is None:
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                count = len([f for f in os.listdir(cls_dir) if f.lower().endswith(IMAGE_EXT)])
                if count >= needed:
                    del img
                    del row
                    if scanned % GC_EVERY == 0:
                        gc.collect()
                    continue
                dest = os.path.join(cls_dir, f"{count+1:05d}.jpg")
                img.convert("RGB").save(dest, "JPEG", quality=85)
                del img
            except Exception:
                pass
            del row
            if cls_label not in new_classes_seen:
                new_classes_seen.add(cls_label)
                print(f"  [{scanned:>8,}]  New class {len(new_classes_seen):>3}: {cls_label}  ({time.time()-t0:.0f}s)")
            if scanned % GC_EVERY == 0:
                gc.collect()
            if scanned % 20_000 == 0:
                print(f"  [{scanned:>8,}]  {len(new_classes_seen)} classes found  ({time.time()-t0:.0f}s)")

    print(f"\n  [LEAFNET] Done: {len(new_classes_seen)} classes, {scanned:,} rows in {time.time()-t0:.1f}s")

    # ── Standard pipeline from here — identical to all other loaders ──────────
    data    = collect_images_df(dl)   # walks dl/<ClassName>/*.jpg
    classes = sorted(data[1].unique().tolist()) if len(data) else []

    if len(data) == 0:
        print(f"  [ERROR] No images saved to {dl}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


# ═══════════════════════════════════════════════════════════════════════════
#  LOCAL DATASET LOADER
# ═══════════════════════════════════════════════════════════════════════════

def collect_images_recursive(folder):
    """
    Return ALL image files found anywhere under `folder` (recursive).
    Used for local datasets where images may be spread across sub-subfolders.
    """
    found = []
    if not os.path.isdir(folder):
        return found
    for dirpath, _, filenames in os.walk(folder):
        for f in filenames:
            if f.lower().endswith(IMAGE_EXT):
                found.append(os.path.join(dirpath, f))
    return found


def build_local_class_map(root):
    """
    Recursively discover all leaf-level disease classes under `root`.

    Strategy
    --------
    The top-level subdirectories of `root` become **categories**
    (e.g. "Corn Diseases", "Soybean Diseases").  Within each category
    every subdirectory — at ANY depth — that directly contains image
    files is treated as a **class**.  The class name is built by joining
    all folder-name segments below the category root with '__', so nested
    paths become unique, flat keys:

        Corn Diseases/Ear rots/Aspergillus ear rot  →  Ear_rots__Aspergillus_ear_rot

    Folders that contain only sub-folders (no images of their own) are
    NOT included as classes — only the deepest image-bearing folders are.
    This avoids double-counting parent folders.

    Returns
    -------
    dict  category_name -> { class_key: absolute_path }
    """
    if not os.path.isdir(root):
        print(f"  [ERROR] LOCAL_SOURCE_ROOT does not exist: {root}")
        return {}

    class_map = {}

    for cat_dir in sorted(os.listdir(root)):
        cat_path = os.path.join(root, cat_dir)
        if not os.path.isdir(cat_path) or cat_dir.startswith('.'):
            continue

        cat_key = safe_name(cat_dir)   # e.g. "Corn_Diseases"
        classes = {}

        for dirpath, dirnames, filenames in os.walk(cat_path):
            # Skip hidden dirs
            dirnames[:] = [d for d in dirnames if not d.startswith('.')]

            # Check if this folder contains images directly
            imgs_here = [f for f in filenames if f.lower().endswith(IMAGE_EXT)]
            if not imgs_here:
                continue  # no images here — keep walking deeper

            # Build class key: relative path from cat_path, separators → '__'
            rel = os.path.relpath(dirpath, cat_path)
            if rel == '.':
                # Images sitting directly in the category root — skip,
                # they don't belong to a named class folder
                continue
            cls_key = safe_name(rel.replace(os.sep, '__'))
            classes[cls_key] = dirpath

        if classes:
            class_map[cat_key] = classes
            print(f"  [LOCAL] {cat_key}: {len(classes)} class folders discovered")
        else:
            print(f"  [LOCAL] {cat_key}: no image-bearing subfolders found — skipped")

    return class_map


def load_local_category(category_name, cls_map, n, source_root):
    """
    Process one auto-discovered local category.
    Images are collected recursively from each class folder.
    Returns (name, saved_df, saved_df, classes, desc).
    """
    name = category_name
    desc = (f"Local disease dataset — {category_name.replace('_', ' ')}. "
            f"Source: {source_root}")

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        classes = list(cls_map.keys())
        saved, _ = load_from_samples(name, classes)
        return name, saved, saved, classes, desc

    need = 1 if n is None else n
    rows, included = [], []
    # cls_map is { class_key: absolute_path } from build_local_class_map — use path, not source_root+key
    for cls_key, src_path in cls_map.items():
        imgs = collect_images_recursive(src_path)
        if len(imgs) < need:
            print(f"  [SKIP CLASS] {cls_key}: {len(imgs)} imgs (need {need})")
            continue
        included.append(cls_key)
        for p in imgs:
            rows.append({0: p, 1: cls_key})

    if not rows:
        print(f"  [WARN] No valid classes in {category_name}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [LOCAL] {len(included)} classes with >= {need} images")
    data     = pd.DataFrame(rows)
    saved_df = sample_per_class(data, included, n)
    saved_df = save_split(saved_df, name, "images")
    print(f"  [SAVED] {len(saved_df)} images -> Curated_Dataset/Images/...")
    return name, saved_df, saved_df, included, desc


BUGWOOD_EXACT_CROP_MAP = {
    # exact canonicals already aligned with DataLoader.py output
    "alfalfa": "Alfalfa",
    "apple": "Apple",
    "banana": "Banana",
    "bananas": "Banana",
    "basil": "Basil",
    "bean": "Bean",
    "bell pepper": "Bell Pepper",
    "bell_pepper": "Bell Pepper",
    "blueberry": "Blueberry",
    "broccoli": "Broccoli",
    "cabbage": "Cabbage",
    "carrot": "Carrot",
    "cashew": "Cashew",
    "cassava": "Cassava",
    "cauliflower": "Cauliflower",
    "celery": "Celery",
    "cherry": "Cherry",
    "chickpea": "Chickpea",
    "citrus": "Citrus",
    "coconut": "Coconut",
    "coconut palm": "Coconut",
    "coffee": "Coffee",
    "corn": "Corn",
    "cotton": "Cotton",
    "cucumber": "Cucumber",
    "durian": "Durian",
    "eggplant": "Eggplant",
    "garlic": "Garlic",
    "ginger": "Ginger",
    "grape": "Grape",
    "grapevine": "Grape",
    "wine grape": "Grape",
    "lettuce": "Lettuce",
    "mango": "Mango",
    "maple": "Maple",
    "orange": "Orange",
    "orange haunglongbing": "Orange",
    "orange huanglongbing": "Orange",
    "peach": "Peach",
    "pear": "Pear",
    "pepper": "Pepper",
    "plum": "Plum",
    "potato": "Potato",
    "pumpkin": "Pumpkin",
    "raspberry": "Raspberry",
    "rice": "Rice",
    "rose": "Rose",
    "rye": "Rye",
    "soybean": "Soybean",
    "squash": "Squash",
    "strawberry": "Strawberry",
    "sugarcane": "Sugarcane",
    "tea": "Tea",
    "tobacco": "Tobacco",
    "tomato": "Tomato",
    "vanilla": "Vanilla",
    "wheat": "Wheat",
    "zucchini": "Zucchini",
}

BUGWOOD_TOMATO_DISEASE_KEYS = {
    "early blight", "late blight", "leaf mold", "septoria", "yellow leaf curl", "mosaic"
}
BUGWOOD_PEPPER_DISEASE_KEYS = {
    "bacterial spot", "leaf spot", "powdery mildew", "pepper"
}
# Bugwood taxonomy entries that are not crop hosts and should not become crop folders.
BUGWOOD_NON_CROP_KEYS = {
    "wood decay fungi",
    "wood decay fungus",
    "canker complex",
    "shelf fungi",
    "bark beetle",
    "powdery mildew",
    "downy mildew",
}


def _normalize_key_for_bugwood(value: str) -> str:
    s = str(value or "").strip().lower().replace("_", " ")
    s = re.sub(r"[^a-z0-9\\s]+", " ", s)
    s = re.sub(r"\\s+", " ", s).strip()
    return s


def _fallback_bugwood_crop(raw_crop: str) -> str | None:
    """
    Preserve Bugwood-only crop hosts that are not yet in the explicit map.
    This prevents valid host crops from being silently dropped.
    """
    crop_key = _normalize_key_for_bugwood(raw_crop)
    if not crop_key or crop_key in BUGWOOD_NON_CROP_KEYS:
        return None
    candidate = re.sub(r"\\s+", " ", str(raw_crop or "").replace("_", " ")).strip().title()
    if not candidate:
        return None
    return _normalize_crop_name(candidate)


def _map_bugwood_crop_exact(raw_crop: str, raw_disease: str) -> str | None:
    crop_key = _normalize_key_for_bugwood(raw_crop)
    disease_key = _normalize_key_for_bugwood(raw_disease)

    if crop_key in ("tomato pepper", "tomato_pepper"):
        if any(k in disease_key for k in BUGWOOD_PEPPER_DISEASE_KEYS):
            return "Pepper"
        if any(k in disease_key for k in BUGWOOD_TOMATO_DISEASE_KEYS):
            return "Tomato"
        # default split fallback: keep tomato for ambiguous shared terms
        return "Tomato"

    mapped = BUGWOOD_EXACT_CROP_MAP.get(crop_key)
    if mapped:
        return mapped
    return _fallback_bugwood_crop(raw_crop)


def load_BugwoodMerged(n):
    """
    Merge Bugwood curated image tree into the final unified Curated_Dataset tree.

    Expected source layout:
      Curated_Bugwood_Dataset/Images/<Crop>/<Disease>/*.jpg

    This loader treats Bugwood as another dataset source and copies images into:
      Curated_Dataset/Images/<Crop>/<Disease>/Bugwood_<idx>.jpg
    while applying the same crop/disease normalization and exclusion rules used
    elsewhere in DataLoader.py.
    """
    name = "Bugwood"
    desc = (
        "Bugwood merged local dataset. Source tree: "
        f"{BUGWOOD_IMAGES_ROOT}. Crops/diseases normalized into unified registry."
    )

    if not os.path.isdir(BUGWOOD_IMAGES_ROOT):
        print(f"  [SKIP] Bugwood — source folder not found: {BUGWOOD_IMAGES_ROOT}")
        return name, pd.DataFrame(columns=[0, 1]), pd.DataFrame(columns=[0, 1]), [], desc

    # Reuse cache detection logic (source-tag based).
    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        classes_from_disk = sorted(saved[1].unique().tolist()) if len(saved) > 0 else []
        return name, saved, saved, classes_from_disk, desc

    # Discover all crop/disease folders in Bugwood tree.
    discovered = []
    for crop_dir in sorted(os.listdir(BUGWOOD_IMAGES_ROOT)):
        crop_path = os.path.join(BUGWOOD_IMAGES_ROOT, crop_dir)
        if not os.path.isdir(crop_path):
            continue
        for disease_dir in sorted(os.listdir(crop_path)):
            disease_path = os.path.join(crop_path, disease_dir)
            if not os.path.isdir(disease_path):
                continue
            imgs = collect_images_recursive(disease_path)
            if not imgs:
                continue
            discovered.append((crop_dir, disease_dir, imgs))

    if not discovered:
        print(f"  [SKIP] Bugwood — no crop/disease image folders found in {BUGWOOD_IMAGES_ROOT}")
        return name, pd.DataFrame(columns=[0, 1]), pd.DataFrame(columns=[0, 1]), [], desc

    src_tag = safe_name(name)[:12]
    out_rows = []
    classes = set()
    copied = 0
    skipped_unmapped = 0

    for raw_crop, raw_disease, imgs in discovered:
        mapped_crop = _map_bugwood_crop_exact(raw_crop, raw_disease)
        if not mapped_crop:
            skipped_unmapped += 1
            continue

        crop, disease = _finalize_registry_pair(
            mapped_crop,
            raw_disease.replace("_", " ").title(),
        )
        if disease in DISEASE_EXCLUDE or (crop, disease) in EXCLUDED_CLASSES:
            continue

        # Optional class quota support to keep behavior consistent with other loaders.
        use_imgs = imgs if n is None else imgs[:n]
        if not use_imgs:
            continue

        dest_dir = os.path.join(IMAGES_DIR, safe_name(crop), safe_name(disease))
        os.makedirs(dest_dir, exist_ok=True)

        existing = [
            f for f in os.listdir(dest_dir)
            if f.lower().endswith(IMAGE_EXT) and f.startswith(src_tag + "_")
        ]
        idx = len(existing) + 1

        for src_path in use_imgs:
            ext = os.path.splitext(src_path)[1].lower() or ".jpg"
            dest = os.path.join(dest_dir, f"{src_tag}_{idx}{ext}")
            while os.path.exists(dest):
                idx += 1
                dest = os.path.join(dest_dir, f"{src_tag}_{idx}{ext}")
            try:
                shutil.copy2(src_path, dest)
                out_rows.append({0: dest, 1: f"{safe_name(crop)}__{safe_name(disease)}"})
                copied += 1
                idx += 1
            except Exception as e:
                print(f"  [WARN] Bugwood copy failed: {src_path} -> {e}")

        classes.add(f"{safe_name(crop)}__{safe_name(disease)}")

    saved_df = pd.DataFrame(out_rows) if out_rows else pd.DataFrame(columns=[0, 1])
    print(f"  [SAVED] Bugwood: {copied} images -> Curated_Dataset/Images/...")
    if skipped_unmapped:
        print(f"  [BUGWOOD] Skipped {skipped_unmapped} unmapped crop folders (strict exact mapping).")
    return name, saved_df, saved_df, sorted(classes), desc

# ═══════════════════════════════════════════════════════════════════════════
#  XLSX EXPORT — grouped by Crop → Disease
# ═══════════════════════════════════════════════════════════════════════════


def generate_xlsx(all_datasets, output_path):
    """
    Sheet 1 — "By Crop & Disease":
        Grouped by Crop (merged cell) → Disease (merged cell) →
        one row per source dataset with image count.

    Sheet 2 — "Summary":
        One row per unique (Crop, Disease) pair, totals across all sources.

    Counts are read from disk (tagged filenames) — single unified pool,
    no ref/bench split. src_tag is passed explicitly to avoid closure bug.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed — skipping xlsx export.")
        return

    # ── Build registry ────────────────────────────────────────────────────────
    from collections import defaultdict

    def _is_internal(ds_name):
        return (ds_name.endswith("_Diseases") or ds_name.endswith("_Disease")
                or ds_name in ("Alfalfa_Diseases", "Corn_Diseases",
                               "Soybean_Diseases", "Wheat_Diseases", "Rye_Diseases"))

    def _display_source(ds_name):
        return "Internal" if _is_internal(ds_name) else ds_name

    # FIX: pass src_tag explicitly so it is never captured from the outer loop.
    def _disk_counts_for(base_dir, tag):
        """Count on-disk images whose filename contains `tag` (the dataset source tag)."""
        counts = {}
        if not os.path.isdir(base_dir):
            return counts
        for cr in os.listdir(base_dir):
            cp = os.path.join(base_dir, cr)
            if not os.path.isdir(cp):
                continue
            for di in os.listdir(cp):
                dp = os.path.join(cp, di)
                if not os.path.isdir(dp):
                    continue
                n_imgs = sum(
                    1 for f in os.listdir(dp)
                    if f.lower().endswith(IMAGE_EXT) and tag in f
                )
                if n_imgs > 0:
                    counts[(cr, di)] = n_imgs
        return counts

    # crop -> disease -> [{"dataset": str, "images": int}]
    grouped  = defaultdict(lambda: defaultdict(list))
    registry = {}   # (crop_lower, disease_lower) -> {crop, disease, sources, total_images}

    for ds_name, saved_df, _unused_bench, classes, _ in all_datasets:
        if not classes:
            continue

        display_src = _display_source(ds_name)
        # FIX: capture src_tag in this iteration's local scope and pass it explicitly.
        tag = safe_name(ds_name)[:12]

        disk_counts = _disk_counts_for(IMAGES_DIR, tag)

        if disk_counts:
            # Primary path: count directly from the curated folder on disk.
            for (crop_dir, disease_dir), n_imgs in disk_counts.items():
                _raw_crop = crop_dir.replace("_", " ").title()
                _raw_dis  = disease_dir.replace("_", " ").title()
                crop, disease = _finalize_registry_pair(_raw_crop, _raw_dis)
                if (crop, disease) in EXCLUDED_CLASSES or disease in DISEASE_EXCLUDE:
                    continue
                grouped[crop][disease].append({"dataset": display_src, "images": n_imgs})
                pk = (crop.lower(), disease.lower())
                if pk not in registry:
                    registry[pk] = {"crop": crop, "disease": disease,
                                    "sources": set(), "total_images": 0}
                registry[pk]["sources"].add(display_src)
                registry[pk]["total_images"] += n_imgs
        else:
            # Fallback: disk not yet written (fresh run) — use the unified saved_df.
            # FIX: only iterate saved_df once (no ref+bench double-count).
            if saved_df is None or len(saved_df) == 0:
                continue
            cnt = {}
            for cls_label, grp in saved_df.groupby(1):
                cnt[cls_label] = len(grp)
            for cls in classes:
                n_imgs = cnt.get(cls, 0)
                if n_imgs == 0:
                    continue
                crop, disease = _parse_crop_disease_from_label(ds_name, cls)
                crop, disease = _finalize_registry_pair(crop, disease)
                if (crop, disease) in EXCLUDED_CLASSES or disease in DISEASE_EXCLUDE:
                    continue
                grouped[crop][disease].append({"dataset": display_src, "images": n_imgs})
                pk = (crop.lower(), disease.lower())
                if pk not in registry:
                    registry[pk] = {"crop": crop, "disease": disease,
                                    "sources": set(), "total_images": 0}
                registry[pk]["sources"].add(display_src)
                registry[pk]["total_images"] += n_imgs


    grouped_sorted = {
        crop: dict(sorted(diseases.items()))
        for crop, diseases in sorted(grouped.items())
    }
    sorted_pairs = sorted(registry.items(), key=lambda x: x[0])

    # ── Shared styles ─────────────────────────────────────────────────────────
    TNR = "Times New Roman"

    HDR_FILL    = PatternFill("solid", start_color="1A1A1A")   # near-black
    HDR_FONT    = Font(name=TNR, bold=True, color="FFFFFF", size=11)
    HDR_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    CROP_FILL   = PatternFill("solid", start_color="333333")   # dark grey
    CROP_FONT   = Font(name=TNR, bold=True, color="FFFFFF", size=11)

    DIS_FILL    = PatternFill("solid", start_color="AAAAAA")   # mid grey
    DIS_FONT    = Font(name=TNR, bold=True, color="000000", size=10)

    BODY_FONT   = Font(name=TNR, size=10)
    BODY_ALIGN  = Alignment(vertical="center")
    ALT_FILL_A  = PatternFill("solid", start_color="F5F5F5")
    ALT_FILL_B  = PatternFill("solid", start_color="FFFFFF")

    TOTAL_FILL  = PatternFill("solid", start_color="1A1A1A")
    TOTAL_FONT  = Font(name=TNR, bold=True, color="FFFFFF", size=10)

    thin   = Side(style="thin",   color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    WRAP_ALIGN  = Alignment(vertical="center", wrap_text=True)
    CTR_ALIGN   = Alignment(horizontal="center", vertical="center")

    def _hdr(ws, headers, col_widths):
        ws.append(headers)
        for col, (cell, w) in enumerate(zip(ws[1], col_widths), start=1):
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = HDR_ALIGN
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    def _style(ws, row_idx, n_cols, fill, font, align=None):
        # Excel maximum row limit: 1048576
        if row_idx > 1048576:
            _log(f"⚠ WARNING: Row {row_idx} exceeds Excel limit (1048576). Skipping style.")
            return
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill      = fill
            cell.font      = font
            cell.border    = BORDER
            cell.alignment = align or BODY_ALIGN

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 1 — By Crop & Disease
    # ════════════════════════════════════════════════════════════════════════
    ws1       = wb.active
    ws1.title = "By Crop & Disease"
    headers1  = ["Crop", "Disease", "Source Dataset", "Images"]
    widths1   = [22, 30, 28, 14]
    _hdr(ws1, headers1, widths1)

    row_idx  = 2
    alt_body = 0

    for crop, diseases in grouped_sorted.items():
        # Count total rows for this crop (for merging)
        crop_row_start = row_idx
        crop_row_count = sum(len(entries) for entries in diseases.values())

        for disease, entries in diseases.items():
            dis_row_start = row_idx
            for entry in entries:
                # Check Excel row limit before appending
                if row_idx >= 1048576:
                    _log(f"⚠ WARNING: Reached Excel row limit. Dataset truncated at {row_idx-1} rows.")
                    break
                alt_body += 1
                fill = ALT_FILL_A if alt_body % 2 == 0 else ALT_FILL_B
                ws1.append(["", "", entry["dataset"], entry["images"]])
                _style(ws1, row_idx, 4, fill, BODY_FONT)
                ws1.cell(row_idx, 3).alignment = BODY_ALIGN
                ws1.cell(row_idx, 4).alignment = CTR_ALIGN
                ws1.row_dimensions[row_idx].height = 16
                row_idx += 1

            # Disease merged cell (col 2, rows dis_row_start..row_idx-1)
            dis_end = row_idx - 1
            if dis_row_start == dis_end:
                ws1.cell(dis_row_start, 2).value = disease
            else:
                ws1.merge_cells(start_row=dis_row_start, start_column=2,
                                end_row=dis_end,        end_column=2)
                ws1.cell(dis_row_start, 2).value = disease
            for r in range(dis_row_start, row_idx):
                c = ws1.cell(r, 2)
                c.fill      = DIS_FILL
                c.font      = DIS_FONT
                c.border    = BORDER
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)

        # Crop merged cell (col 1, rows crop_row_start..row_idx-1)
        crop_end = row_idx - 1
        if crop_row_start == crop_end:
            ws1.cell(crop_row_start, 1).value = crop
        else:
            ws1.merge_cells(start_row=crop_row_start, start_column=1,
                            end_row=crop_end,         end_column=1)
            ws1.cell(crop_row_start, 1).value = crop
        for r in range(crop_row_start, row_idx):
            c = ws1.cell(r, 1)
            c.fill      = CROP_FILL
            c.font      = CROP_FONT
            c.border    = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)

    # Totals row
    n_data = row_idx - 2
    ws1.append(["TOTAL", "", "", f"=SUM(D2:D{row_idx-1})"])
    _style(ws1, row_idx, 4, TOTAL_FILL, TOTAL_FONT, CTR_ALIGN)
    ws1.row_dimensions[row_idx].height = 20

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 2 — Summary (one row per unique Crop × Disease)
    # ════════════════════════════════════════════════════════════════════════
    ws2       = wb.create_sheet("Summary")
    headers2  = ["#", "Crop", "Disease", "# Sources", "Source Datasets", "Total Images"]
    widths2   = [5, 20, 30, 10, 44, 14]
    _hdr(ws2, headers2, widths2)

    for i, (pk, info) in enumerate(sorted_pairs, start=1):
        # Check Excel row limit before appending
        if i + 1 >= 1048576:
            _log(f"⚠ WARNING: Sheet 2 reached Excel row limit at {i} rows.")
            break
        sources_str = ", ".join(sorted(info["sources"]))
        n_src       = len(info["sources"])
        ws2.append([i, info["crop"], info["disease"], n_src,
                    sources_str, info["total_images"]])
        r = i + 1
        fill = ALT_FILL_A if i % 2 == 0 else ALT_FILL_B
        _style(ws2, r, 6, fill, BODY_FONT)
        ws2.cell(r, 5).alignment = WRAP_ALIGN
        ws2.cell(r, 1).alignment = CTR_ALIGN
        ws2.cell(r, 4).alignment = CTR_ALIGN
        ws2.cell(r, 6).alignment = CTR_ALIGN
        ws2.row_dimensions[r].height = 16

    # Totals row
    n2 = len(sorted_pairs)
    ws2.append(["", "TOTAL", f"=COUNTA(C2:C{n2+1})", "", "", f"=SUM(F2:F{n2+1})"])
    _style(ws2, n2+2, 6, TOTAL_FILL, TOTAL_FONT, CTR_ALIGN)
    ws2.cell(n2+2, 2).alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[n2+2].height = 20

    # Build a small DataFrame from the registry for downstream visualisation.
    summary_rows = []
    for (_pk, info) in sorted_pairs:
        summary_rows.append({
            "Crop": info["crop"],
            "Disease": info["disease"],
            "Total Images": info["total_images"],
        })
    summary_df = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(
        columns=["Crop", "Disease", "Total Images"]
    )

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 3 — Image Sources (every image with its source dataset)
    # ════════════════════════════════════════════════════════════════════════
    tag_to_name = {}
    for ds_name, _ref, _bench, _cls, _ in all_datasets:
        tag = safe_name(ds_name)[:12]
        tag_to_name[tag] = ds_name

    ws3 = wb.create_sheet("Image Sources")
    headers3 = ["Crop", "Disease", "Source Dataset", "Filename"]
    widths3  = [22, 35, 32, 28]
    _hdr(ws3, headers3, widths3)

    row3 = 2
    if os.path.isdir(IMAGES_DIR):
        for crop_dir in sorted(os.listdir(IMAGES_DIR)):
            crop_path = os.path.join(IMAGES_DIR, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in sorted(os.listdir(crop_path)):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                crop_display, disease_display = _finalize_registry_pair(
                    crop_dir.replace("_", " ").title(),
                    disease_dir.replace("_", " ").title(),
                )
                for f in sorted(os.listdir(disease_path)):
                    if not f.lower().endswith(IMAGE_EXT):
                        continue
                    # Check Excel row limit before appending
                    if row3 >= 1048576:
                        _log(f"⚠ WARNING: Sheet 3 (Image Sources) reached Excel row limit at {row3-1} images. Truncating.")
                        break
                    base, ext = os.path.splitext(f)
                    parts = base.rsplit("_", 1)
                    if len(parts) == 2 and parts[1].isdigit():
                        tag = parts[0]
                        source = tag_to_name.get(tag, tag)
                    else:
                        source = "Unknown"
                    ws3.append([crop_display, disease_display, source, f])
                    fill = ALT_FILL_A if row3 % 2 == 0 else ALT_FILL_B
                    _style(ws3, row3, 4, fill, BODY_FONT)
                    ws3.row_dimensions[row3].height = 16
                    row3 += 1

    ws3.append(["TOTAL", "", "", f"=COUNTA(D2:D{row3})"])
    _style(ws3, row3, 4, TOTAL_FILL, TOTAL_FONT, CTR_ALIGN)
    ws3.row_dimensions[row3].height = 20
    print(f"         Image Sources: {row3 - 2} rows (one per image)")

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 4 — Datasets & Papers (one row per dataset)
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Datasets & Papers")
    headers4 = ["Dataset Name", "URL", "Paper / Citation"]
    widths4  = [26, 60, 80]
    _hdr(ws4, headers4, widths4)

    # Collect all dataset names seen plus any extra keys in DATASET_METADATA
    seen_ds = set(ds_name for ds_name, _ref, _bench, _cls, _ in all_datasets)
    for extra in DATASET_METADATA.keys():
        seen_ds.add(extra)

    row4 = 2
    for ds_name in sorted(seen_ds):
        # Check Excel row limit before appending
        if row4 >= 1048576:
            _log(f"⚠ WARNING: Sheet 4 (Datasets) reached Excel row limit at {row4-1} datasets.")
            break
        url, citation = DATASET_METADATA.get(ds_name, ("", ""))
        ws4.append([ds_name, url, citation])
        fill = ALT_FILL_A if row4 % 2 == 0 else ALT_FILL_B
        _style(ws4, row4, 3, fill, BODY_FONT)
        ws4.cell(row4, 1).alignment = WRAP_ALIGN
        ws4.cell(row4, 2).alignment = WRAP_ALIGN
        ws4.cell(row4, 3).alignment = WRAP_ALIGN
        ws4.row_dimensions[row4].height = 22
        row4 += 1

    # ════════════════════════════════════════════════════════════════════════
    #  EXPORT TO CSV ONLY (skip Excel generation)
    # ════════════════════════════════════════════════════════════════════════

    base_name = os.path.splitext(output_path)[0]

    # Export comprehensive summary CSV
    try:
        registry_csv = base_name + ".csv"
        summary_df.to_csv(registry_csv, index=False)
        _log(f"  CSV (Summary) -> {os.path.abspath(registry_csv)}")
        _log(f"         {len(sorted_pairs)} unique crop-disease pairs")
    except Exception as e:
        _log(f"  [WARN] Could not write summary CSV: {e}")

    # Sheet 1: By Crop & Disease
    try:
        sheet1_csv = f"{base_name}_sheet1_crop_disease.csv"
        sheet1_data = []
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                sheet1_data.append({
                    "Crop": row[0],
                    "Disease": row[1],
                    "Source Dataset": row[2],
                    "Images": row[3]
                })
        if sheet1_data:
            pd.DataFrame(sheet1_data).to_csv(sheet1_csv, index=False)
            _log(f"  CSV (Sheet 1) -> {sheet1_csv}")
    except Exception as e:
        _log(f"  [WARN] Could not export Sheet 1 to CSV: {e}")

    # Sheet 3: Image Sources (most detailed)
    try:
        sheet3_csv = f"{base_name}_sheet3_image_sources.csv"
        sheet3_data = []
        for row in ws3.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                sheet3_data.append({
                    "Crop": row[0],
                    "Disease": row[1],
                    "Source": row[2],
                    "Filename": row[3]
                })
        if sheet3_data:
            pd.DataFrame(sheet3_data).to_csv(sheet3_csv, index=False)
            _log(f"  CSV (Sheet 3) -> {sheet3_csv}")
    except Exception as e:
        _log(f"  [WARN] Could not export Sheet 3 to CSV: {e}")

    return


# ═══════════════════════════════════════════════════════════════════════════
#  PLANT VILLAGE LOADER  (tensorflow_datasets)
# ═══════════════════════════════════════════════════════════════════════════

def load_PlantVillage(n):
    """
    Loads PlantVillage via tensorflow_datasets.
    Label format:  "Crop___Disease"  (triple underscore)
    Each label becomes one class; crop and disease are split on "___".

    Requires:  pip install tensorflow tensorflow_datasets
    """
    name = "PlantVillage"
    desc = ("PlantVillage dataset loaded via tensorflow_datasets. "
            "Covers 38 crop-disease classes across 14 crop species including "
            "healthy controls. Images are RGB leaf photos taken under controlled "
            "conditions.")

    if not TFDS_AVAILABLE:
        print("  [SKIP] PlantVillage — tensorflow_datasets not installed. "
              "Run:  pip install tensorflow tensorflow_datasets")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        # Recover class list from saved folders
        src_tag  = safe_name(name)[:12]
        classes  = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp) if f.lower().endswith(IMAGE_EXT)):
                    classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(classes), desc

    dl = os.path.join(DATA_ROOT, "plantvillage-raw")
    if os.path.isdir(dl):
        shutil.rmtree(dl, ignore_errors=True)
    os.makedirs(dl, exist_ok=True)

    # Force TFDS download/cache under work root (data_dir overrides any default)
    tfds_data_dir = os.path.join(_BASE_DIR, "data", "tfds")
    os.makedirs(tfds_data_dir, exist_ok=True)
    print(f"  Loading PlantVillage via tensorflow_datasets -> {dl}  (TFDS data: {tfds_data_dir})")
    try:
        ds, ds_info = tfds.load(
            "plant_village",
            split="train",
            as_supervised=True,
            with_info=True,
            data_dir=tfds_data_dir,
        )
    except Exception as e:
        print(f"  [ERROR] Could not load PlantVillage: {e}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    labels      = ds_info.features["label"].names   # e.g. ["Apple___Apple_scab", ...]
    needed      = (n if n is not None else 999999)  # all when DOWNLOAD_ALL
    counters    = {}   # cls_key -> count saved

    print(f"  PlantVillage: {len(labels)} classes, streaming images ...")

    for img_tensor, label_idx in ds:
        label_raw = labels[int(label_idx)]          # "Apple___Apple_scab"
        # Normalise: split on "___" -> (crop, disease), safe_name each part
        if "___" in label_raw:
            parts   = label_raw.split("___", 1)
            crop    = safe_name(parts[0].replace("_", " ").strip())
            disease = safe_name(parts[1].replace("_", " ").strip())
        else:
            crop    = safe_name(label_raw)
            disease = "Unknown"

        cls_key = f"{crop}__{disease}"             # double underscore — matches load_from_samples
        cls_dir = os.path.join(dl, cls_key)
        os.makedirs(cls_dir, exist_ok=True)

        if counters.get(cls_key, 0) >= needed:
            continue                                 # already have enough for this class

        try:
            img_np  = img_tensor.numpy()
            pil_img = Image.fromarray(img_np).convert("RGB")
            count   = counters.get(cls_key, 0)
            dest    = os.path.join(cls_dir, f"{count+1:05d}.jpg")
            pil_img.save(dest, "JPEG", quality=85)
            counters[cls_key] = count + 1
        except Exception:
            continue

        # Stop early once every class has enough images
        if all(v >= needed for v in counters.values()) and len(counters) == len(labels):
            break

    total_saved = sum(counters.values())
    print(f"  PlantVillage: {len(counters)} classes, {total_saved} images staged")

    if total_saved == 0:
        print(f"  [ERROR] No images saved for PlantVillage")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    data    = collect_images_df(dl)
    classes = sorted(data[1].unique().tolist()) if len(data) else []
    saved, _ = split_save_cleanup(data, classes, name, dl, n)
    return name, saved, saved, classes, desc


# ═══════════════════════════════════════════════════════════════════════════
#  CDDM LOCAL LOADER
#  Structure: CDDM-images/images/<Crop,Disease>/  (comma-separated folder names)
#  Each folder is both a crop and a disease — no sub-hierarchy needed.
# ═══════════════════════════════════════════════════════════════════════════

# ── Auto-download helpers ────────────────────────────────────────────────────

def _ensure_gdown():
    """Install gdown if not available."""
    try:
        import gdown  # noqa: F401
        return True
    except ImportError:
        _log("  [AUTO-DL] gdown not found — installing ...")
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "-q", "gdown"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
            import gdown  # noqa: F401
            _log("  [AUTO-DL] gdown installed.")
            return True
        except Exception as e:
            _log(f"  [AUTO-DL] Could not install gdown: {e}")
            return False


def _download_plantwild(dest_dir: str) -> bool:
    """
    Download plantwild_v2.zip from Google Drive and unzip to dest_dir.
    File ID: 1wJEMRaNNuYGDqq2IQraQVtkCRHTOIDLD  (~1.5 GB)
    Falls back to wget if gdown fails.
    """
    GDRIVE_FILE_ID = "1wJEMRaNNuYGDqq2IQraQVtkCRHTOIDLD"
    zip_path = os.path.join(_BASE_DIR, "plantwild_v2.zip")

    if os.path.isdir(dest_dir) and any(
        os.path.isdir(os.path.join(dest_dir, d)) for d in (os.listdir(dest_dir) if os.path.isdir(dest_dir) else [])
    ):
        _log(f"  [AUTO-DL] PlantWild already present at {dest_dir}, skipping download.")
        return True

    _log(f"  [AUTO-DL] PlantWild v2 not found — downloading (~1.5 GB) ...")

    downloaded = False

    # Try gdown first
    if _ensure_gdown():
        try:
            import gdown
            _log(f"  [AUTO-DL] Trying gdown for PlantWild ...")
            gdown.download(id=GDRIVE_FILE_ID, output=zip_path, quiet=False)
            if os.path.isfile(zip_path) and os.path.getsize(zip_path) > 1_000_000:
                downloaded = True
                _log(f"  [AUTO-DL] gdown download complete: {zip_path}")
            else:
                _log(f"  [AUTO-DL] gdown produced empty/missing file, falling back to wget.")
        except Exception as e:
            _log(f"  [AUTO-DL] gdown failed: {e} — trying wget ...")

    # Fallback: wget with the direct Drive URL (mirrors the working command you showed)
    if not downloaded:
        gdrive_url = (
            f"https://drive.usercontent.google.com/download"
            f"?id={GDRIVE_FILE_ID}&export=download&confirm=t"
        )
        _log(f"  [AUTO-DL] Downloading via wget: {gdrive_url}")
        try:
            ret = subprocess.call(
                ["wget", "--no-check-certificate", "-q", "--show-progress",
                 "-O", zip_path, gdrive_url]
            )
            if ret == 0 and os.path.isfile(zip_path) and os.path.getsize(zip_path) > 1_000_000:
                downloaded = True
                _log(f"  [AUTO-DL] wget download complete: {zip_path}")
            else:
                _log(f"  [AUTO-DL] wget failed (exit code {ret}).")
        except FileNotFoundError:
            _log("  [AUTO-DL] wget not available on this system.")

    if not downloaded:
        _log("  [AUTO-DL] PlantWild download failed — skipping.")
        return False

    # Unzip
    _log(f"  [AUTO-DL] Extracting {zip_path} -> {_BASE_DIR} ...")
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(_BASE_DIR)
        _log(f"  [AUTO-DL] Extraction complete -> {dest_dir}")
        os.remove(zip_path)
        _log(f"  [AUTO-DL] Removed zip: {zip_path}")
        return os.path.isdir(dest_dir)
    except Exception as e:
        _log(f"  [AUTO-DL] Extraction error: {e}")
        return False


def _download_cddm(dest_dir: str) -> bool:
    """
    Download CDDM dataset images from the UnicomBenchmark GitHub repo
    using sparse git checkout (pulls only CDDMBench/dataset/images/).
    Falls back to gdown if a Drive mirror is available.
    """
    GITHUB_REPO = "https://github.com/UnicomAI/UnicomBenchmark.git"
    SPARSE_PATH = "CDDMBench/dataset/images"
    clone_dir = os.path.join(_BASE_DIR, "_cddm_clone_tmp")

    if os.path.isdir(dest_dir) and any(
        os.path.isdir(os.path.join(dest_dir, d)) for d in (os.listdir(dest_dir) if os.path.isdir(dest_dir) else [])
    ):
        _log(f"  [AUTO-DL] CDDM already present at {dest_dir}, skipping download.")
        return True

    _log(f"  [AUTO-DL] CDDM not found — fetching via sparse git checkout ...")
    _log(f"  [AUTO-DL] Repo: {GITHUB_REPO}  path: {SPARSE_PATH}")

    try:
        if os.path.isdir(clone_dir):
            shutil.rmtree(clone_dir)
        os.makedirs(clone_dir, exist_ok=True)

        _log("  [AUTO-DL] git init + sparse-checkout ...")
        subprocess.check_call(["git", "init", clone_dir],
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.check_call(
            ["git", "-C", clone_dir, "remote", "add", "origin", GITHUB_REPO],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        subprocess.check_call(
            ["git", "-C", clone_dir, "config", "core.sparseCheckout", "true"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        sparse_file = os.path.join(clone_dir, ".git", "info", "sparse-checkout")
        with open(sparse_file, "w") as f:
            f.write(SPARSE_PATH + "\n")

        _log("  [AUTO-DL] git pull (--depth 1) — this may take a few minutes ...")
        subprocess.check_call(
            ["git", "-C", clone_dir, "pull", "--depth", "1", "origin", "main"],
        )

        src = os.path.join(clone_dir, SPARSE_PATH)
        if not os.path.isdir(src):
            _log(f"  [AUTO-DL] Expected path not found after clone: {src}")
            shutil.rmtree(clone_dir, ignore_errors=True)
            return False

        # Move to expected location: _BASE_DIR/CDDM-images/images/
        os.makedirs(dest_dir, exist_ok=True)
        for item in os.listdir(src):
            s = os.path.join(src, item)
            d = os.path.join(dest_dir, item)
            shutil.move(s, d)

        shutil.rmtree(clone_dir, ignore_errors=True)
        _log(f"  [AUTO-DL] CDDM images ready at {dest_dir}")
        return True

    except subprocess.CalledProcessError as e:
        _log(f"  [AUTO-DL] git sparse checkout failed: {e}")
        shutil.rmtree(clone_dir, ignore_errors=True)
        return False
    except Exception as e:
        _log(f"  [AUTO-DL] CDDM download error: {e}")
        shutil.rmtree(clone_dir, ignore_errors=True)
        return False


def load_CDDM(n):
    """
    Loads the CDDM (Comma-Delimited Disease Map) local dataset.
    Folder structure:  <root>/<Crop,Disease>/plant_xxxxx.jpg
    The comma in the folder name separates Crop from Disease, e.g.:
      "Apple,Brown Spot"  ->  crop="Apple", disease="Brown Spot"
    Each folder becomes one class. Classes with fewer than (n + n)
    images are skipped with a warning.
    """
    name = "CDDM"
    # Resolve root at runtime: prefer .../CDDM-images/images/ when present (Crop,Disease folders live there)
    cddm_root = CDDM_SOURCE_ROOT
    if cddm_root:
        images_sub = os.path.join(cddm_root.rstrip(os.sep), "images")
        if os.path.isdir(images_sub):
            cddm_root = images_sub

    # Auto-download if not present
    if not cddm_root or not os.path.isdir(cddm_root):
        _log(f"  [AUTO-DL] CDDM path missing ({cddm_root}) — attempting auto-download ...")
        dl_target = os.path.join(_BASE_DIR, "CDDM-images", "images")
        if _download_cddm(dl_target):
            cddm_root = dl_target
        else:
            cddm_root = None

    desc = ("CDDM local dataset. Classes encoded as 'Crop,Disease' folder names. "
            f"Source: {cddm_root}")

    if not cddm_root or not os.path.isdir(cddm_root):
        print(f"  [SKIP] CDDM — path not found and auto-download failed: {cddm_root or CDDM_SOURCE_ROOT}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [CDDM] Scanning: {cddm_root}")
    # Discover all class folders (any folder directly inside root, e.g. "Apple,Brown Spot")
    all_cls_dirs = sorted([
        d for d in os.listdir(cddm_root)
        if os.path.isdir(os.path.join(cddm_root, d))
        and not d.startswith('.')
    ])

    if not all_cls_dirs:
        print(f"  [SKIP] CDDM — no class folders found in {cddm_root}")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    # Use safe_name of each folder as the class key stored on disk
    classes = [safe_name(d) for d in all_cls_dirs]

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp): continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp): continue
                if any(src_tag in f for f in os.listdir(dp)
                       if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        # Stale cache: disk has fewer classes than source (e.g. old run saved 1 class) — re-sample
        if len(saved_classes) < len(all_cls_dirs):
            print(f"  [RESAMPLE] CDDM — disk has {len(saved_classes)} class(es), source has {len(all_cls_dirs)} — re-sampling from source.")
        else:
            return name, saved, saved, sorted(saved_classes), desc

    need = 1 if n is None else n
    rows, included = [], []
    for raw_dir in all_cls_dirs:
        cls_key  = safe_name(raw_dir)
        src_path = os.path.join(cddm_root, raw_dir)
        imgs     = collect_images_recursive(src_path)
        if len(imgs) < need:
            print(f"  [SKIP CLASS] {raw_dir}: {len(imgs)} imgs (need {need})")
            continue
        included.append(cls_key)
        for p in imgs:
            rows.append({0: p, 1: cls_key})

    if not rows:
        print(f"  [WARN] No valid classes in CDDM")
        return name, pd.DataFrame(columns=[0,1]), pd.DataFrame(columns=[0,1]), [], desc

    print(f"  [CDDM] {len(included)} classes with >= {need} images")
    data     = pd.DataFrame(rows)
    saved_df = sample_per_class(data, included, n)
    saved_df = save_split(saved_df, name, "images")
    print(f"  [SAVED] {len(saved_df)} images -> Curated_Dataset/Images/...")
    return name, saved_df, saved_df, included, desc


# ═══════════════════════════════════════════════════════════════════════════
#  PLANTWILD V2 LOCAL LOADER
#  Structure: plantwild_v2/<class_name>/<image>.jpg  (e.g. apple black rot/apple_black_rot_1.jpg)
#  Class = subdirectory name; images directly inside (or recursive).
# ═══════════════════════════════════════════════════════════════════════════

def load_PlantWild(n):
    """
    Loads the PlantWild v2 local dataset.
    Folder structure:  <root>/<class_name>/<image>.jpg
    Example: plantwild_v2/apple black rot/apple_black_rot_1.jpg
    Each subdirectory is one class. Classes with fewer than n images are skipped.
    """
    name = "PlantWild"
    root = PLANTWILD_SOURCE_ROOT

    # Auto-download if not present
    if not root or not os.path.isdir(root):
        _log(f"  [AUTO-DL] PlantWild path missing ({root}) — attempting auto-download ...")
        if _download_plantwild(root or os.path.join(_BASE_DIR, "plantwild_v2")):
            root = PLANTWILD_SOURCE_ROOT  # re-check after download
        else:
            root = None

    desc = (f"PlantWild v2 local dataset. Classes = subdirectory names. Source: {root}")

    if not root or not os.path.isdir(root):
        print(f"  [SKIP] PlantWild — path not found and auto-download failed: {root}")
        return name, pd.DataFrame(columns=[0, 1]), pd.DataFrame(columns=[0, 1]), [], desc

    print(f"  [PlantWild] Scanning: {root}")
    all_cls_dirs = sorted([
        d for d in os.listdir(root)
        if os.path.isdir(os.path.join(root, d)) and not d.startswith(".")
    ])

    if not all_cls_dirs:
        print(f"  [SKIP] PlantWild — no class folders in {root}")
        return name, pd.DataFrame(columns=[0, 1]), pd.DataFrame(columns=[0, 1]), [], desc

    if already_sampled(name, n):
        print(f"  [SKIP] {name} — samples exist, loading from disk.")
        saved, _ = load_from_samples(name, [])
        src_tag = safe_name(name)[:12]
        saved_classes = []
        for crop_d in (os.listdir(IMAGES_DIR) if os.path.isdir(IMAGES_DIR) else []):
            cp = os.path.join(IMAGES_DIR, crop_d)
            if not os.path.isdir(cp):
                continue
            for dis_d in os.listdir(cp):
                dp = os.path.join(cp, dis_d)
                if not os.path.isdir(dp):
                    continue
                if any(src_tag in f for f in os.listdir(dp) if f.lower().endswith(IMAGE_EXT)):
                    saved_classes.append(f"{crop_d}__{dis_d}")
        return name, saved, saved, sorted(saved_classes), desc

    need = 1 if n is None else n
    rows, included = [], []
    for raw_dir in all_cls_dirs:
        cls_key = safe_name(raw_dir)
        src_path = os.path.join(root, raw_dir)
        imgs = collect_images_recursive(src_path)
        if len(imgs) < need:
            print(f"  [SKIP CLASS] {raw_dir}: {len(imgs)} imgs (need {need})")
            continue
        included.append(cls_key)
        for p in imgs:
            rows.append({0: p, 1: cls_key})

    if not rows:
        print(f"  [WARN] PlantWild — no valid classes")
        return name, pd.DataFrame(columns=[0, 1]), pd.DataFrame(columns=[0, 1]), [], desc

    print(f"  [PlantWild] {len(included)} classes with >= {need} images")
    data = pd.DataFrame(rows)
    saved_df = sample_per_class(data, included, n)
    saved_df = save_split(saved_df, name, "images")
    print(f"  [SAVED] {len(saved_df)} images -> Curated_Dataset/Images/...")
    return name, saved_df, saved_df, included, desc


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    _log("DataLoader: main() started.")

    # Load local .env (HF tokens etc.) before any remote dataset access.
    _load_local_env()

    # Hard "erase" for each execution: never reuse previously processed outputs.
    for p in [DATA_ROOT, CURATED_DIR, SUNBURST_DIR]:
        if os.path.isdir(p):
            shutil.rmtree(p, ignore_errors=True)
    if os.path.isfile(OUTPUT_XLSX):
        try:
            os.remove(OUTPUT_XLSX)
        except OSError:
            pass

    os.makedirs(DATA_ROOT,  exist_ok=True)
    os.makedirs(IMAGES_DIR, exist_ok=True)

    # ── Migration: normalise existing crop/disease directories on disk ───────
    migrate_curated_image_tree(IMAGES_DIR)

    n = None   # no CLI: download all images per class from every source
    _log("\n  [MODE] Download all images from all sources (no prompts).\n")

    all_datasets = []

    def _print_cropwise_classes(dataset_name, classes):
        """
        Print crop-wise disease class names for this dataset using the same
        parsing + normalisation rules used by the registry, and emit debug
        information for any classes that were dropped.
        """
        if not classes:
            return
        from collections import defaultdict

        grouped = defaultdict(set)  # crop -> set(diseases)
        dropped = []                # (raw_label, reason)

        for cls_label in classes:
            raw = str(cls_label)
            crop, disease = _parse_crop_disease_from_label(dataset_name, raw)
            crop_n, dis_n = _finalize_registry_pair(crop, disease)

            if dis_n in DISEASE_EXCLUDE:
                dropped.append((raw, f"excluded by DISEASE_EXCLUDE as '{dis_n}'"))
                continue
            if (crop_n, dis_n) in EXCLUDED_CLASSES:
                dropped.append((raw, f"excluded by EXCLUDED_CLASSES as ('{crop_n}','{dis_n}')"))
                continue

            grouped[crop_n].add(dis_n)

        if not grouped:
            return

        _log("   [CLASSES] Crop-wise disease classes:")
        for crop in sorted(grouped.keys()):
            diseases = sorted(grouped[crop])
            _log(f"     - {crop} ({len(diseases)}): {', '.join(diseases)}")

        if dropped:
            _log("   [DEBUG] Dropped class labels for this dataset:")
            for raw, reason in dropped:
                _log(f"     - '{raw}' -> {reason}")

    def _expected_pairs_for_dataset(dataset_name, classes):
        exp = set()
        for cls in (classes or []):
            crop, disease = _parse_crop_disease_from_label(dataset_name, str(cls))
            crop, disease = _finalize_registry_pair(crop, disease)
            if disease in DISEASE_EXCLUDE or (crop, disease) in EXCLUDED_CLASSES:
                continue
            exp.add((crop, disease))
        return exp

    def _disk_pairs_for_dataset(dataset_name):
        src_tag = safe_name(dataset_name)[:12]
        pairs = set()
        if not os.path.isdir(IMAGES_DIR):
            return pairs
        for crop_dir in os.listdir(IMAGES_DIR):
            crop_path = os.path.join(IMAGES_DIR, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in os.listdir(crop_path):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                try:
                    has_any = any(
                        f.lower().endswith(IMAGE_EXT) and src_tag in f
                        for f in os.listdir(disease_path)
                    )
                except OSError:
                    has_any = False
                if not has_any:
                    continue
                crop, disease = _finalize_registry_pair(
                    crop_dir.replace("_", " ").title(),
                    disease_dir.replace("_", " ").title(),
                )
                if disease in DISEASE_EXCLUDE or (crop, disease) in EXCLUDED_CLASSES:
                    continue
                pairs.add((crop, disease))
        return pairs

    def _delete_tagged_images(dataset_name):
        src_tag = safe_name(dataset_name)[:12]
        if not os.path.isdir(IMAGES_DIR):
            return
        for crop_dir in list(os.listdir(IMAGES_DIR)):
            crop_path = os.path.join(IMAGES_DIR, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in list(os.listdir(crop_path)):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                try:
                    for f in list(os.listdir(disease_path)):
                        if f.lower().endswith(IMAGE_EXT) and src_tag in f:
                            try:
                                os.remove(os.path.join(disease_path, f))
                            except OSError:
                                pass
                except OSError:
                    pass
                # prune empty disease dir
                try:
                    if not any(fn.lower().endswith(IMAGE_EXT) for fn in os.listdir(disease_path)):
                        os.rmdir(disease_path)
                except OSError:
                    pass
            # prune empty crop dir
            try:
                if not any(os.path.isdir(os.path.join(crop_path, d)) for d in os.listdir(crop_path)):
                    os.rmdir(crop_path)
            except OSError:
                pass

    def _run_loader_with_resample(loader_fn, n):
        # First attempt (may hit cache)
        result = loader_fn(n)
        ds_name, saved, _bench_unused, classes, desc = result
        # Normalise to unified 5-tuple: (name, saved_df, saved_df, classes, desc)
        result = (ds_name, saved, saved, classes, desc)

        # For CDDM we never auto-resample. Its classes and splits are fixed
        # and relatively heavy; if you need to refresh it, do so manually.
        if ds_name == "CDDM":
            return result

        # Check for stale cache: disk has fewer classes than expected
        exp = _expected_pairs_for_dataset(ds_name, classes)
        disk = _disk_pairs_for_dataset(ds_name)
        if exp and len(disk) < len(exp):
            _log(f"  [RESAMPLE] {ds_name} — disk has {len(disk)} class(es), expected {len(exp)} — re-sampling.")
            _delete_tagged_images(ds_name)
            FORCE_RESAMPLE_DATASETS.add(ds_name)
            try:
                result = loader_fn(n)
                ds_name2, saved2, _b2, classes2, desc2 = result
                result = (ds_name2, saved2, saved2, classes2, desc2)
            finally:
                FORCE_RESAMPLE_DATASETS.discard(ds_name)
        return result

    def _update_outputs():
        # XLSX is always refreshed after each dataset
        generate_xlsx(all_datasets, OUTPUT_XLSX)

    def _leafnet_smoke_ok():
        """
        Lightweight connectivity check for LeafNet:
        try to touch a tiny split slice so we fail fast on auth issues.
        """
        if not HF_AVAILABLE:
            return False
        try:
            # Small sample to validate token / connectivity; avoid large download.
            _ = hf_load_dataset("enalis/LeafNet", split="train[:1]")
            _log("  [LEAFNET SMOKE] HuggingFace access OK.")
            return True
        except Exception as e:
            _log(f"  [LEAFNET SMOKE] Failed to load enalis/LeafNet: {e}")
            return False

    # ── Online datasets ───────────────────────────────────────────────────────
    _log("\n" + "="*60)
    _log("  ONLINE DATASETS")
    _log("="*60)

    leafnet_ok = _leafnet_smoke_ok()

    online_loaders = [
        load_SBRD, load_MangoLeaf, load_SoybeanPNAS,
        load_BeanLeaf, load_YellowRust,
        load_BananaLeaf, load_Cauliflower, load_Lettuce,
        load_Cucumber, load_DurianLeaf, load_EggplantDisease,
        load_CottonDisease, load_PumpkinLeaf, load_RoseLeaf,
        load_CoconutDisease,
        load_VanillaDisease,
        load_StrawberryDiseaseDetection,
        load_SugarLeafIDN, load_CucumberZenodo,
        load_NewPlantDiseases, load_PlantDoc,
        load_FUSARIUM22,
        load_RadyPlantDiseases,
        load_A2H0H0R1PlantDisease,
        load_AvinashPlantDisease,
        load_SakethPlantDisease,
        load_VQAPlantDisease,
        load_BDCropVegetable,
        load_LeafNet, load_PlantVillage,
    ]

    for loader in online_loaders:
        if loader is load_LeafNet and not leafnet_ok:
            _log("  [SKIP] LeafNet loader due to failing smoke-check.")
            continue
        _log(f"\n-- {loader.__name__} --")
        result = _run_loader_with_resample(loader, n)
        all_datasets.append(result)
        nm, saved, _bench, cls, _ = result
        _log(f"   {nm}: {len(saved)} images | {len(cls)} classes")
        _print_cropwise_classes(nm, cls)
        _update_outputs()

    # ── InternalData local datasets ───────────────────────────────────────────
    # Resolve path: use _BASE_DIR/InternalData or _BASE_DIR/data/InternalData if present
    _local_root = LOCAL_SOURCE_ROOT
    if _local_root and not os.path.isdir(_local_root):
        _fallback = os.path.join(DATA_ROOT, "InternalData")
        if os.path.isdir(_fallback):
            _local_root = _fallback
            _log(f"  [LOCAL] Using InternalData at data/InternalData: {_local_root}")
    if _local_root:
        if os.path.isdir(_local_root):
            print("\n" + "="*60)
            print("  LOCAL DATASETS  (InternalData)")
            print(f"  Source: {_local_root}")
            print("="*60)
            for cat, cls_map in build_local_class_map(_local_root).items():
                print(f"\n-- Local: {cat} --")
                result = _run_loader_with_resample(lambda _n: load_local_category(cat, cls_map, _n, _local_root), n)
                all_datasets.append(result)
                nm, saved, _bench, cls, _ = result
                print(f"   {nm}: {len(saved)} images | {len(cls)} classes")
                _update_outputs()
        else:
            print(f"\n[WARN] InternalData not found: {LOCAL_SOURCE_ROOT} or {os.path.join(DATA_ROOT, 'InternalData')}")

    # ── CDDM dataset ──────────────────────────────────────────────────────────
    # ── CDDM dataset (auto-downloads if not present) ─────────────────────────────
    print("\n" + "="*60)
    print("  LOCAL DATASET  (CDDM)")
    print(f"  Source: {CDDM_SOURCE_ROOT}  [auto-download enabled]")
    print("="*60)
    print("\n-- load_CDDM --")
    result = _run_loader_with_resample(load_CDDM, n)
    all_datasets.append(result)
    nm, saved, _bench, cls, _ = result
    print(f"   {nm}: {len(saved)} images | {len(cls)} classes")
    _update_outputs()


    # ── PlantWild v2 dataset ──────────────────────────────────────────────────
    # ── PlantWild v2 dataset (auto-downloads if not present) ───────────────────
    print("\n" + "="*60)
    print("  LOCAL DATASET  (PlantWild v2)")
    print(f"  Source: {PLANTWILD_SOURCE_ROOT}  [auto-download enabled]")
    print("="*60)
    print("\n-- load_PlantWild --")
    result = _run_loader_with_resample(load_PlantWild, n)
    all_datasets.append(result)
    nm, saved, _bench, cls, _ = result
    print(f"   {nm}: {len(saved)} images | {len(cls)} classes")
    _update_outputs()

    # ── Bugwood merged dataset ───────────────────────────────────────────────
    print("\n" + "="*60)
    print("  LOCAL DATASET  (Bugwood Merged)")
    print(f"  Source: {BUGWOOD_IMAGES_ROOT}")
    print("="*60)
    print("\n-- load_BugwoodMerged --")
    result = _run_loader_with_resample(load_BugwoodMerged, n)
    all_datasets.append(result)
    nm, saved, _bench, cls, _ = result
    print(f"   {nm}: {len(saved)} images | {len(cls)} classes")
    _update_outputs()


    # ── Final XLSX ───────────────────────────────────────────────────────────
    print("\n" + "="*60)
    print("  BUILDING XLSX (final)")
    print("="*60)
    # Merge/rename/drop on disk so the workbook matches folder layout.
    migrate_curated_image_tree(IMAGES_DIR)
    generate_xlsx(all_datasets, OUTPUT_XLSX)   # final XLSX (all data)

    # ── Manifest — count actual files on disk, flag shortfalls ───────────────
    def _count_saved(ds_name, split_base, src_tag):
        """Count images on disk tagged with this dataset's source tag."""
        total = 0
        if not os.path.isdir(split_base):
            return total
        for crop_dir in os.listdir(split_base):
            crop_path = os.path.join(split_base, crop_dir)
            if not os.path.isdir(crop_path):
                continue
            for disease_dir in os.listdir(crop_path):
                disease_path = os.path.join(crop_path, disease_dir)
                if not os.path.isdir(disease_path):
                    continue
                total += sum(
                    1 for f in os.listdir(disease_path)
                    if f.lower().endswith(IMAGE_EXT) and src_tag in f
                )
        return total

    def _shortfall_classes(ds_name, n_required):
        """Return list of (crop/disease, ref_count, bench_count) where count < n_required."""
        src_tag = safe_name(ds_name)[:12]
        issues  = []
        for split_base, split_label in [(IMAGES_DIR, "images")]:  # single dir
            if not os.path.isdir(split_base):
                continue
            for crop_dir in os.listdir(split_base):
                crop_path = os.path.join(split_base, crop_dir)
                if not os.path.isdir(crop_path):
                    continue
                for disease_dir in os.listdir(crop_path):
                    disease_path = os.path.join(crop_path, disease_dir)
                    if not os.path.isdir(disease_path):
                        continue
                    count = sum(
                        1 for f in os.listdir(disease_path)
                        if f.lower().endswith(IMAGE_EXT) and src_tag in f
                    )
                    if 0 < count < n_required:
                        issues.append(f"{crop_dir}/{disease_dir} [{split_label}]: "
                                      f"{count}/{n_required}")
        return issues

    ds_entries = []
    shortfall_report = []

    for r in all_datasets:
        ds_name = r[0]
        saved   = r[1]   # unified pool (ref/bench merged)
        classes = r[3]
        src_tag = safe_name(ds_name)[:12]

        n_saved   = _count_saved(ds_name, IMAGES_DIR, src_tag)
        expected  = (n * len(classes)) if n is not None else n_saved

        shortfalls = _shortfall_classes(ds_name, n) if (classes and n is not None) else []
        if shortfalls:
            shortfall_report.append({"dataset": ds_name, "classes_below_target": shortfalls})

        ds_entries.append({
            "name":         ds_name,
            "classes":     len(classes),
            "images":       n_saved,
            "expected":    expected,
            "complete":    n_saved == expected,
        })

    manifest = {
        "images_per_class": n,
        "download_all":     n is None,
        "datasets": ds_entries,
    }
    if shortfall_report:
        manifest["shortfalls"] = shortfall_report
        print("\n  [WARN] Some classes had fewer images than requested:")
        for s in shortfall_report:
            print(f"    {s['dataset']}:")
            for c in s["classes_below_target"]:
                print(f"      - {c}")

    mpath = os.path.join(SAMPLES_DIR, "manifest.json")
    with open(mpath, "w") as f:
        json.dump(manifest, f, indent=2)
    print(f"  Manifest -> {os.path.abspath(mpath)}\n")


if __name__ == "__main__":
    main()