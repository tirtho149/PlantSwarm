"""
scripts/registry_to_excel.py
============================
Convert a unified ``final_registry.json`` into a SINGLE-sheet Excel
workbook. One row per disease. Canonical KB on the left (the trunk of
the decision tree). State-specific deltas on the right (the branches),
collapsed into a single multiline cell — each delta names which
canonical field it refines, what canonical says about that field, what
the image shows, and a one-sentence visual quote.

This is the decision-tree shape: canonical owns the symptom slots,
regional only emits ADDS or CONTRADICTS — never a parallel extraction.

Usage:
    python scripts/registry_to_excel.py \\
        artifacts/pathome_kb/Soybean/final_registry.json \\
        [--out artifacts/pathome_kb/Soybean/final_registry.xlsx]
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
STATE_FONT = Font(bold=True, color="0B4F8A")


def _val(field):
    if not isinstance(field, dict):
        return ""
    v = field.get("value")
    if v is None:
        return ""
    if isinstance(v, list):
        return "; ".join(str(x) for x in v if x)
    return str(v)


def _format_deltas(per_state: dict) -> str:
    if not per_state:
        return ""
    lines = []
    for state in sorted(per_state.keys()):
        rec = per_state[state]
        if not isinstance(rec, dict):
            continue
        deltas = rec.get("deltas") or []
        if not deltas:
            lines.append(f"━━ {state} ━━  (image confirms canonical — no deltas)")
            lines.append("")
            continue
        lines.append(f"━━ {state} ━━")
        for d in deltas:
            field = d.get("field", "other")
            says = d.get("canonical_says", "(not specified)")
            shows = d.get("image_shows", "")
            quote = d.get("image_quote", "")
            lines.append(f"• [{field}]")
            lines.append(f"    canonical: {says}")
            lines.append(f"    image:     {shows}")
            if quote:
                lines.append(f"    evidence:  \"{quote}\"")
        lines.append("")
    return "\n".join(lines).rstrip()


def write_decision_tree_sheet(ws, crop, diseases):
    headers = [
        "disease",
        "pathogen",
        "type",
        "affected_parts",
        "canonical_summary",
        "diagnostic_features",
        "look_alikes",
        "treatments",
        "n_states",
        "n_deltas",
        "regional_deltas (decision-tree)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="top")

    row = 2
    for d in diseases:
        per_state = d.get("regional_observations") or {}
        n_states = len(per_state)
        n_deltas = sum(len((s.get("deltas") or [])) for s in per_state.values()
                       if isinstance(s, dict))

        vals = [
            d.get("disease_name", ""),
            _val(d.get("pathogen_scientific_name")),
            _val(d.get("type_of_disease")),
            _val(d.get("affected_parts")),
            _val((d.get("visual_symptoms") or {}).get("summary")),
            _val((d.get("visual_symptoms") or {}).get("diagnostic_features")),
            _val((d.get("visual_symptoms") or {}).get("look_alikes")),
            _val(d.get("treatments")),
            n_states,
            n_deltas,
            _format_deltas(per_state),
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.freeze_panes = "B2"
    widths = [28, 32, 14, 28, 70, 70, 36, 70, 8, 8, 110]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    ap.add_argument("registry", help="path to final_registry.json")
    ap.add_argument("--out", default=None, help="output xlsx path (default: final_registry.xlsx in same dir)")
    args = ap.parse_args()

    src = Path(args.registry)
    if not src.is_file():
        raise SystemExit(f"registry not found: {src}")
    data = json.load(open(src))
    crop = data.get("crop", "")
    diseases = data.get("diseases", []) or []

    dst = Path(args.out) if args.out else (src.parent / "final_registry.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = crop or "registry"
    write_decision_tree_sheet(ws, crop, diseases)
    wb.save(dst)

    n_states = sum(len(d.get("regional_observations") or {}) for d in diseases)
    n_deltas = 0
    for d in diseases:
        for s in (d.get("regional_observations") or {}).values():
            if isinstance(s, dict):
                n_deltas += len(s.get("deltas") or [])
    print(f"wrote {dst}")
    print(f"  diseases               : {len(diseases)}")
    print(f"  per-state observations : {n_states}")
    print(f"  state-specific deltas  : {n_deltas}")


if __name__ == "__main__":
    main()
